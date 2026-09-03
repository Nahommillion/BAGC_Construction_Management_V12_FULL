import os, sqlite3, calendar, datetime as dt, json
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

BASE=os.path.dirname(os.path.abspath(__file__))
# Keep application data outside the code directory when BAGC_DATA_DIR is configured.
# This prevents redeploying/replacing the source from replacing the database.
DATA_DIR=os.environ.get("BAGC_DATA_DIR", "/data" if os.path.isdir("/data") and os.access("/data", os.W_OK) else BASE).strip() or BASE
os.makedirs(DATA_DIR, exist_ok=True)
DB=os.path.join(DATA_DIR,"bagc.db")
UPLOADS=os.path.join(DATA_DIR,"uploads")
USER_PHOTOS=os.path.join(UPLOADS,"user_photos")
WORKFLOW_FILES=os.path.join(UPLOADS,"workflow")
os.makedirs(UPLOADS,exist_ok=True)
os.makedirs(USER_PHOTOS,exist_ok=True)
os.makedirs(WORKFLOW_FILES,exist_ok=True)
ALLOWED_PHOTO_EXT={"jpg","jpeg","png","webp"}
ALLOWED_FILE_EXT={"pdf","doc","docx","xls","xlsx","csv","txt","jpg","jpeg","png","webp","zip","rar"}
MAX_UPLOAD_MB=25
app=Flask(__name__)
app.secret_key=os.environ.get("SECRET_KEY","bagc-change-this-secret")

DEPARTMENTS=["Administration","Design","Machinery","Finance","HR","Store","Project","Consultant"]
HEAD_OFFICE_STRUCTURE=[
("General Manager",None,"Management"),
("Operational Manager","General Manager","Management"),
("Equipment & Store Department","Operational Manager","Department"),
("Equipment Management Team","Equipment & Store Department","Team"),
("Machinery Team","Equipment & Store Department","Team"),
("Fuel Team","Equipment & Store Department","Team"),
("Store Team","Equipment & Store Department","Team"),
("Contract Administration Department","Operational Manager","Department"),
("Contract Administration Team","Contract Administration Department","Team"),
("Procurement Team","Contract Administration Department","Team"),
("Engineering Department","Operational Manager","Department"),
("Engineering Team","Engineering Department","Team"),
("Project Management Team","Engineering Department","Team"),
("Design Team","Engineering Department","Team"),
("Quantity Survey & Cost Control Team","Engineering Department","Team"),
("Planning & Monitoring Team","Engineering Department","Team"),
("QA/QC Team","Engineering Department","Team"),
("Survey Team","Engineering Department","Team"),
("HSE / Safety Team","Engineering Department","Team"),
("HR Department","Operational Manager","Department"),
("HR Team","HR Department","Team"),
("Recruitment & Staff Administration Team","HR Department","Team"),
("Training & Performance Team","HR Department","Team"),
("Finance Department","Operational Manager","Department"),
("Finance Team","Finance Department","Team"),
("Accounting Team","Finance Department","Team"),
("Cash & Treasury Team","Finance Department","Team"),
("Budget & Cost Control Team","Finance Department","Team"),
("Administration & General Services Department","Operational Manager","Department"),
("Administration Team","Administration & General Services Department","Team"),
("Document Control Team","Administration & General Services Department","Team"),
("IT & Systems Team","Administration & General Services Department","Team"),
("Legal & Compliance Team","Operational Manager","Team")
]
UNIT_CATALOG=["m","m²","m³","mm","cm","kg","ton","litre","L","pcs","pc","no","set","lot","item","bag","roll","sheet","length","day","hour","hr","month","lump sum"]
MACHINE_TYPES=["Dozer","Excavator","Wheel Loader","Backhoe Loader","Motor Grader","Roller","Dump Truck","Fuel Truck","Water Truck","Shower Truck","Crane","Forklift","Concrete Mixer","Concrete Pump","Batching Plant","Crusher","Asphalt Plant","Asphalt Paver","Bitumen Distributor","Road Sweeper","Generator","Welding Machine","Vibrator","Air Compressor","Pickup","Other"]
MATERIAL_CATEGORIES=["Common Construction","Earthworks","Concrete","Rebar","Structural Steel / RHS","Formwork","Masonry","Roofing","Waterproofing","Finishing","Tiles","Natural Stone","Sanitary","Plumbing","Drainage","Electrical","Low Voltage / ICT","Aluminium","Glass","Road Works","Culverts","Landscaping","Fuel & Oil","Lubricants","Spare Parts","Welding / Cutting","PPE / Safety","Stationery & Cleaning","Tools","Other"]
MATERIAL_CATALOG=[
"Cement OPC","Cement PPC","Cement Rapid Hardening","Sand Fine","Sand Coarse","Quarry Dust","Selected Fill","Subbase","Base Course",
"Aggregate 5mm","Aggregate 10mm","Aggregate 14mm","Aggregate 20mm","Aggregate 25mm","Aggregate 40mm","Aggregate 50mm","Water",
"Rebar Ø6","Rebar Ø8","Rebar Ø10","Rebar Ø12","Rebar Ø14","Rebar Ø16","Rebar Ø20","Rebar Ø25","Rebar Ø32","Binding Wire 0.9mm","Binding Wire 1.2mm","Black Wire Roll",
"Welded Wire Mesh","Tie Wire","Anchor Bolt","Steel Plate","Angle Bar","Flat Bar","Round Bar","Channel","I-Beam","H-Beam","RHS 40x40","RHS 50x50","RHS 60x60","RHS 80x80","RHS 100x50","RHS 100x100","SHS","GI Sheet","Corrugated Sheet","Roofing Screw","Self Tapping Screw",
"Plywood 4x8","Plywood 12mm","Plywood 15mm","Plywood 18mm","Marine Plywood","Eucalyptus Pole","Eucalyptus Plank","Timber 2x4","Timber 2x6","Timber 2x8","Timber 4x4","Scaffold Tube","Scaffold Coupler","Form Tie","Form Oil","Release Agent","Nails 1in","Nails 2in","Nails 3in","Nails 4in","Nails 5in","Nails 6in",
"Concrete Block","Hollow Concrete Block 10cm","Hollow Concrete Block 15cm","Hollow Concrete Block 20cm","Solid Block","Brick","Kerbstone","Interlock Paver","Concrete Pipe","U-Ditch","Manhole Cover",
"PVC Pipe 25mm","PVC Pipe 32mm","PVC Pipe 50mm","PVC Pipe 75mm","PVC Pipe 110mm","PVC Pipe 160mm","PVC Pipe 200mm","PVC Pipe 250mm","HDPE Pipe","PPR Pipe","UPVC Fitting","Elbow","Tee","Reducer","Valve","Manhole Ring","Geotextile","Waterproofing Membrane",
"Electrical Cable 1.5mm2","Electrical Cable 2.5mm2","Electrical Cable 4mm2","Electrical Cable 6mm2","Electrical Cable 10mm2","Electrical Cable 16mm2","Armoured Cable","Control Cable","Conduit PVC","Flexible Conduit","Junction Box","Distribution Board","MCB","MCCB","RCD","Isolator","Switch","Socket","Industrial Socket","Light Fixture","LED Light","Street Light","Earthing Rod","Earthing Cable","Cable Lug","Cable Tray","Cable Tie","Electrical Tape",
"Aluminium Frame","Aluminium Sheet","Aluminium Composite Panel","Glass 5mm","Glass 6mm","Tempered Glass","Laminated Glass","Silicone","Rubber Gasket","Door Handle","Lockset","Hinge","Door Closer",
"Ceramic Tile","Porcelain Tile","Granite Tile","Marble","Natural Stone Cladding","Stone Veneer","Terrazzo","Skirting Tile","Tile Adhesive","Grout","Cement Mortar","Waterproofing Additive","Paint Primer","Emulsion Paint","Enamel Paint","Exterior Paint","Thinner","Putty","Gypsum Board","Gypsum Powder","Ceiling Tile","Acoustic Panel","Insulation","Bituminous Paint",
"Sanitary WC","Wash Basin","Urinal","Shower Mixer","Tap","Bib Tap","Floor Drain","Bottle Trap","P-Trap","S-Trap","Flush Valve","Flush Tank","Flexible Hose","Mirror","Soap Holder","Towel Rail","Sanitary Silicone",
"Diesel","Petrol","Engine Oil","Hydraulic Oil","Gear Oil","Brake Fluid","Coolant","Grease","AdBlue","Bitumen 60/70","Emulsion","Prime Coat","Tack Coat","Welding Rod 2.5mm","Welding Rod 3.2mm","Welding Rod 4mm","Oxygen","Acetylene","Grinding Disc","Cutting Disc","Drill Bit","Saw Blade",
"Safety Helmet","Safety Vest","Safety Boot","Safety Glove","Safety Goggles","Ear Plug","Dust Mask","Harness","Reflective Tape","Barricade","Warning Sign","Cones","Fire Extinguisher","First Aid Kit",
"Other"]
CREW_GROUPS=["Project Management","Key Staff","Earthwork Crew","Structure Crew","Road/Culvert Crew","Equipment / Machinery","Store","DL","Data Collector","Survey Team","Supporting Staff","Office Staff","Skilled Labour","Semi-Skilled Labour","Non-Skilled Labour","Security / General Support"]
POSITION_CATALOG=["Project Manager","Deputy Project Manager","Construction Manager","Site Engineer","Office Engineer","Quantity Surveyor","Planning Engineer","Design Engineer","QA/QC Engineer","Materials Engineer","Surveyor","Survey Assistant","HSE Officer","Foreman","Earthwork Foreman","Structure Foreman","Road Foreman","DL","Data Collector","Store Keeper","Store Assistant","Mechanic","Electrician","Plumber","Mason","Carpenter","Steel Fixer","Welder","Painter","Aluminium Worker","Equipment Operator","Driver","Labourer","Security Guard","Cleaner","Office Assistant","Document Controller","Accountant","Procurement Officer","Other","Office Head","Senior Store Officer","Senior Equipment Officer","Senior Fuel Officer","HR Head","Finance Head","Design Head"]
DESIGN_STATUSES=["Draft","Submitted","Under Review","Approved","Approved with Comments","Revise & Resubmit","Rejected","As-Built","Handed Over"]

# Strict module visibility: ordinary personnel only see the module owned by their department.
# Project personnel receive Project tools; cross-department work is handled through Workflow/Requests.
MODULE_DEPARTMENTS={
    "Administration":"Administration", "Design":"Design", "Machinery":"Machinery",
    "Finance":"Finance", "HR":"HR", "Store":"Store", "Project":"Project", "Consultant":"Consultant"
}
PERSONNEL_SCOPES=["HEAD_OFFICE","PROJECT"]
REPORT_SCOPES_BY_DEPARTMENT={"Machinery":{"MACHINERY","FUEL"},"Store":{"STORE"},"HR":{"MANPOWER"},"Finance":{"FINANCE"},"Design":{"DESIGN"},"Project":{"BOQ","PROBLEMS"},"Administration":{"FINANCE"},"Consultant":{"BOQ","PROBLEMS"}}
REQUEST_TYPES=["MATERIAL","FUEL","MACHINERY","MANPOWER","EXPENSE","DESIGN","PROCUREMENT","OTHER"]
REQUEST_CATEGORIES={
    "MATERIAL":"Store / Material", "FUEL":"Fuel", "MACHINERY":"Machinery", "MANPOWER":"Manpower / HR",
    "EXPENSE":"Finance / Expense", "DESIGN":"Design", "PROCUREMENT":"Procurement", "OTHER":"General"
}


@app.context_processor
def template_helpers():
    return {"dt": dt, "head_office_units": HEAD_OFFICE_STRUCTURE, "project_admin": project_admin}


def db():
    c=sqlite3.connect(DB, timeout=30, isolation_level="DEFERRED")
    c.row_factory=sqlite3.Row
    c.execute("PRAGMA busy_timeout=30000")
    c.execute("PRAGMA foreign_keys=ON")
    return c


DEPT_CODES={'Administration':'ADM','Design':'DSN','Machinery':'MCH','Finance':'FIN','HR':'HR','Store':'STR','Project':'PRJ'}
def make_staff_id(department, seq):
    code=DEPT_CODES.get(department,'STF')
    return f"BAGC-{code}-{dt.date.today().year}-{seq:04d}"


def init_db():
    if os.environ.get("RENDER") and not os.environ.get("BAGC_DATA_DIR"):
        app.logger.warning("BAGC_DATA_DIR is not set on Render. SQLite will be ephemeral; configure a persistent disk or external database for permanent users/reports/BOQ.")
    c=db()
    try:
        c.execute("PRAGMA journal_mode=WAL")
        c.execute("PRAGMA synchronous=NORMAL")
    except sqlite3.OperationalError:
        pass
    c.executescript('''
    CREATE TABLE IF NOT EXISTS users(id INTEGER PRIMARY KEY,full_name TEXT,username TEXT UNIQUE,password_hash TEXT,department TEXT,position TEXT,location TEXT,phone TEXT,email TEXT,role TEXT,active INTEGER DEFAULT 1,staff_id TEXT UNIQUE,photo_filename TEXT,last_login TEXT,org_unit_id INTEGER,reports_to_user_id INTEGER,created_at TEXT DEFAULT CURRENT_TIMESTAMP,personnel_scope TEXT DEFAULT 'PROJECT');
    CREATE TABLE IF NOT EXISTS projects(id INTEGER PRIMARY KEY,name TEXT UNIQUE,code TEXT,location TEXT,client TEXT,consultant TEXT,status TEXT DEFAULT 'Active',start_date TEXT,end_date TEXT);
    CREATE TABLE IF NOT EXISTS org_units(id INTEGER PRIMARY KEY,name TEXT UNIQUE,parent_id INTEGER,unit_type TEXT DEFAULT 'Team',active INTEGER DEFAULT 1,manager_user_id INTEGER,sort_order INTEGER DEFAULT 0);
    CREATE TABLE IF NOT EXISTS user_projects(user_id INTEGER,project_id INTEGER,UNIQUE(user_id,project_id));
    CREATE TABLE IF NOT EXISTS boq(id INTEGER PRIMARY KEY,project_id INTEGER,item_no TEXT,description TEXT,unit TEXT,rate REAL DEFAULT 0,contract_qty REAL DEFAULT 0,source_sheet TEXT,series TEXT DEFAULT '',title TEXT DEFAULT '',UNIQUE(project_id,item_no));
    CREATE TABLE IF NOT EXISTS boq_settings(id INTEGER PRIMARY KEY,project_id INTEGER UNIQUE,title TEXT DEFAULT '',revision TEXT DEFAULT '',effective_date TEXT);
    CREATE TABLE IF NOT EXISTS daily_work(id INTEGER PRIMARY KEY,project_id INTEGER,date TEXT,boq_id INTEGER,quantity REAL,unit TEXT,station_from TEXT,station_to TEXT,notes TEXT,user_id INTEGER);
    CREATE TABLE IF NOT EXISTS machines(id INTEGER PRIMARY KEY,project_id INTEGER,machine_type TEXT,code TEXT,ownership TEXT,hourly_rate REAL DEFAULT 0,rate_unit TEXT DEFAULT 'hr',expected_fuel REAL DEFAULT 0,active INTEGER DEFAULT 1);
    CREATE TABLE IF NOT EXISTS machine_logs(id INTEGER PRIMARY KEY,project_id INTEGER,machine_id INTEGER,date TEXT,work_hours REAL DEFAULT 0,idle_hours REAL DEFAULT 0,idle_reason TEXT,idle_payable INTEGER DEFAULT 0,down_hours REAL DEFAULT 0,down_reason TEXT,opening_gauge REAL DEFAULT 0,fuel_received REAL DEFAULT 0,closing_gauge REAL DEFAULT 0,notes TEXT,user_id INTEGER);
    CREATE TABLE IF NOT EXISTS materials(id INTEGER PRIMARY KEY,project_id INTEGER,category TEXT,name TEXT,unit TEXT,min_stock REAL DEFAULT 0,active INTEGER DEFAULT 1,UNIQUE(project_id,name));
    CREATE TABLE IF NOT EXISTS store_logs(id INTEGER PRIMARY KEY,project_id INTEGER,material_id INTEGER,date TEXT,received REAL DEFAULT 0,issued REAL DEFAULT 0,unit_cost REAL DEFAULT 0,physical_balance REAL,reference TEXT,notes TEXT,user_id INTEGER);
    CREATE TABLE IF NOT EXISTS manpower(id INTEGER PRIMARY KEY,project_id INTEGER,date TEXT,name TEXT,employment TEXT,position TEXT,present REAL DEFAULT 1,working_hours REAL DEFAULT 8,hourly_rate REAL DEFAULT 0,daily_rate REAL DEFAULT 0,normal_ot_hours REAL DEFAULT 0,normal_ot_rate REAL DEFAULT 0,night_ot_hours REAL DEFAULT 0,night_ot_rate REAL DEFAULT 0,sunday_ot_hours REAL DEFAULT 0,sunday_ot_rate REAL DEFAULT 0,holiday_ot_hours REAL DEFAULT 0,holiday_ot_rate REAL DEFAULT 0,overtime_hours REAL DEFAULT 0,overtime_rate REAL DEFAULT 0,notes TEXT,user_id INTEGER);
    CREATE TABLE IF NOT EXISTS design_items(id INTEGER PRIMARY KEY,project_id INTEGER,drawing_no TEXT,title TEXT,discipline TEXT,revision TEXT,status TEXT,submitted TEXT,approved TEXT,comments TEXT,user_id INTEGER);
    CREATE TABLE IF NOT EXISTS finance_logs(id INTEGER PRIMARY KEY,project_id INTEGER,date TEXT,category TEXT,kind TEXT,description TEXT,amount REAL DEFAULT 0,reference TEXT,user_id INTEGER);
    CREATE TABLE IF NOT EXISTS boq_uploads(id INTEGER PRIMARY KEY,project_id INTEGER,filename TEXT,uploaded_at TEXT,user_id INTEGER,rows_imported INTEGER DEFAULT 0);
    CREATE TABLE IF NOT EXISTS performance_rates(id INTEGER PRIMARY KEY,project_id INTEGER,work_type TEXT,worker_type TEXT,unit TEXT,qty_per_hour REAL DEFAULT 0,notes TEXT,UNIQUE(project_id,work_type,worker_type));
    CREATE TABLE IF NOT EXISTS daily_activities(id INTEGER PRIMARY KEY,project_id INTEGER,date TEXT,boq_id INTEGER,work_type TEXT,executed_qty REAL DEFAULT 0,unit TEXT,machine_id INTEGER,machine_hours REAL DEFAULT 0,manpower_position TEXT,manpower_qty REAL DEFAULT 0,manpower_hours REAL DEFAULT 0,material_id INTEGER,material_qty REAL DEFAULT 0,remarks TEXT,user_id INTEGER);
    CREATE TABLE IF NOT EXISTS problems(id INTEGER PRIMARY KEY,project_id INTEGER,date TEXT,problem TEXT,remark TEXT,user_id INTEGER);
    CREATE TABLE IF NOT EXISTS fuel_logs(id INTEGER PRIMARY KEY,project_id INTEGER,machine_id INTEGER,date TEXT,opening_gauge REAL DEFAULT 0,fuel_received REAL DEFAULT 0,closing_gauge REAL DEFAULT 0,fuel_price REAL DEFAULT 0,reference TEXT,notes TEXT,user_id INTEGER,source TEXT DEFAULT 'Fuel Register');
    CREATE TABLE IF NOT EXISTS project_crews(id INTEGER PRIMARY KEY,project_id INTEGER,date TEXT,group_name TEXT,position TEXT,name TEXT,employment TEXT,skill_level TEXT,working_hours REAL DEFAULT 0,hourly_rate REAL DEFAULT 0,notes TEXT,user_id INTEGER);
    CREATE TABLE IF NOT EXISTS crew_group_capacity(id INTEGER PRIMARY KEY,group_name TEXT UNIQUE,foreman_qty REAL DEFAULT 0,dl_qty REAL DEFAULT 0,surveyor_qty REAL DEFAULT 0,data_collector_qty REAL DEFAULT 0,time_keeper_qty REAL DEFAULT 0,other_qty REAL DEFAULT 0,total_qty REAL DEFAULT 0,active INTEGER DEFAULT 1);
    CREATE TABLE IF NOT EXISTS activity_machines(id INTEGER PRIMARY KEY,activity_id INTEGER,machine_log_id INTEGER,machine_id INTEGER,hours REAL DEFAULT 0);
    CREATE TABLE IF NOT EXISTS activity_manpower(id INTEGER PRIMARY KEY,activity_id INTEGER,manpower_id INTEGER,crew_id INTEGER,qty REAL DEFAULT 0,hours REAL DEFAULT 0);
    CREATE TABLE IF NOT EXISTS activity_store(id INTEGER PRIMARY KEY,activity_id INTEGER,store_log_id INTEGER,material_id INTEGER,qty REAL DEFAULT 0);
    CREATE TABLE IF NOT EXISTS activity_fuel(id INTEGER PRIMARY KEY,activity_id INTEGER,fuel_log_id INTEGER,litres REAL DEFAULT 0);
    CREATE TABLE IF NOT EXISTS activity_finance(id INTEGER PRIMARY KEY,activity_id INTEGER,finance_log_id INTEGER,amount REAL DEFAULT 0);
    CREATE TABLE IF NOT EXISTS crew_evaluations(id INTEGER PRIMARY KEY,activity_id INTEGER,crew_id INTEGER,evaluation TEXT,remarks TEXT,score REAL);
    CREATE TABLE IF NOT EXISTS variation_alerts(id INTEGER PRIMARY KEY,project_id INTEGER,boq_id INTEGER,date TEXT,contract_qty REAL,previous_qty REAL,period_qty REAL,to_date_qty REAL,excess_qty REAL,status TEXT DEFAULT 'OPEN',message TEXT,created_by INTEGER,created_at TEXT DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS crew_groups(id INTEGER PRIMARY KEY,name TEXT UNIQUE,active INTEGER DEFAULT 1);
    CREATE TABLE IF NOT EXISTS crew_positions(id INTEGER PRIMARY KEY,name TEXT UNIQUE,active INTEGER DEFAULT 1);
    CREATE TABLE IF NOT EXISTS report_settings(id INTEGER PRIMARY KEY,project_id INTEGER UNIQUE,contractor_role TEXT DEFAULT 'Main Contractor',phone TEXT,email TEXT,website TEXT,fax TEXT,address TEXT,logo_text TEXT);
    CREATE TABLE IF NOT EXISTS rfis(id INTEGER PRIMARY KEY,project_id INTEGER,rfi_no TEXT,date_requested TEXT,inspection_date TEXT,location TEXT,boq_id INTEGER,work_description TEXT,drawing_no TEXT,drawing_revision TEXT,specification TEXT,work_stage TEXT,submitted_by INTEGER,status TEXT DEFAULT 'PENDING INSPECTION',overall_comment TEXT,corrective_action TEXT,created_at TEXT DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS rfi_inspections(id INTEGER PRIMARY KEY,rfi_id INTEGER,inspector_user_id INTEGER,inspector_role TEXT,decision TEXT DEFAULT 'PENDING',comments TEXT,inspection_date TEXT,signed_at TEXT,UNIQUE(rfi_id,inspector_user_id));
    CREATE TABLE IF NOT EXISTS rfi_steps(id INTEGER PRIMARY KEY,rfi_id INTEGER,step_order INTEGER,stage TEXT,assigned_user_id INTEGER,decision TEXT DEFAULT 'PENDING',comments TEXT,inspection_date TEXT,signed_at TEXT,UNIQUE(rfi_id,step_order));
    CREATE TABLE IF NOT EXISTS saved_reports(id INTEGER PRIMARY KEY,project_id INTEGER,report_no TEXT,report_type TEXT,scope TEXT DEFAULT 'ALL',start_date TEXT,end_date TEXT,generated_by INTEGER,generated_at TEXT DEFAULT CURRENT_TIMESTAMP,snapshot_json TEXT,source_report_ids TEXT DEFAULT '[]',UNIQUE(project_id,report_type,scope,start_date,end_date));
    CREATE TABLE IF NOT EXISTS machine_assignments(id INTEGER PRIMARY KEY,machine_id INTEGER,project_id INTEGER,start_date TEXT,start_hour REAL DEFAULT 0,total_signed_hours REAL DEFAULT 0,hours_used REAL DEFAULT 0,end_date TEXT,end_hour REAL,status TEXT DEFAULT 'ACTIVE',assigned_by INTEGER,ended_by INTEGER,ended_at TEXT,notes TEXT);
    CREATE TABLE IF NOT EXISTS workflow_files(id INTEGER PRIMARY KEY,project_id INTEGER,from_user_id INTEGER,to_user_id INTEGER,to_org_unit_id INTEGER,file_name TEXT,stored_name TEXT,file_type TEXT,subject TEXT,message TEXT,category TEXT DEFAULT 'General Correspondence',status TEXT DEFAULT 'SENT',sent_at TEXT DEFAULT CURRENT_TIMESTAMP,received_at TEXT,created_by INTEGER);
    CREATE TABLE IF NOT EXISTS material_transfers(id INTEGER PRIMARY KEY,from_project_id INTEGER,to_project_id INTEGER,material_id INTEGER,date TEXT,quantity REAL DEFAULT 0,unit_cost REAL DEFAULT 0,reference TEXT,notes TEXT,sent_by INTEGER,received_by INTEGER,status TEXT DEFAULT 'SENT',sent_at TEXT DEFAULT CURRENT_TIMESTAMP,received_at TEXT);
    CREATE TABLE IF NOT EXISTS fuel_transfers(id INTEGER PRIMARY KEY,from_project_id INTEGER,to_project_id INTEGER,machine_id INTEGER,date TEXT,litres REAL DEFAULT 0,unit_cost REAL DEFAULT 0,reference TEXT,notes TEXT,sent_by INTEGER,received_by INTEGER,status TEXT DEFAULT 'SENT',sent_at TEXT DEFAULT CURRENT_TIMESTAMP,received_at TEXT);
    CREATE TABLE IF NOT EXISTS machine_transfers(id INTEGER PRIMARY KEY,from_project_id INTEGER,to_project_id INTEGER,machine_id INTEGER,date TEXT,notes TEXT,sent_by INTEGER,received_by INTEGER,status TEXT DEFAULT 'SENT',sent_at TEXT DEFAULT CURRENT_TIMESTAMP,received_at TEXT);
    CREATE TABLE IF NOT EXISTS expense_claims(id INTEGER PRIMARY KEY,project_id INTEGER,date TEXT,beneficiary_user_id INTEGER,beneficiary_name TEXT,category TEXT,description TEXT,amount REAL DEFAULT 0,paid_by_company INTEGER DEFAULT 1,receipt_file TEXT,receipt_name TEXT,submitted_by INTEGER,approved_by INTEGER,status TEXT DEFAULT 'SUBMITTED',created_at TEXT DEFAULT CURRENT_TIMESTAMP,approved_at TEXT);
    CREATE TABLE IF NOT EXISTS project_assignments(id INTEGER PRIMARY KEY,user_id INTEGER,project_id INTEGER,position TEXT,manager_user_id INTEGER,active INTEGER DEFAULT 1,UNIQUE(user_id,project_id));
    CREATE TABLE IF NOT EXISTS responsibilities(id INTEGER PRIMARY KEY,supervisor_user_id INTEGER,subordinate_user_id INTEGER,scope_type TEXT NOT NULL,project_id INTEGER,active INTEGER DEFAULT 1,source TEXT DEFAULT 'Manual',created_at TEXT DEFAULT CURRENT_TIMESTAMP,UNIQUE(supervisor_user_id,subordinate_user_id,scope_type,project_id));
    CREATE TABLE IF NOT EXISTS resource_requests(id INTEGER PRIMARY KEY,request_no TEXT UNIQUE,request_type TEXT,project_id INTEGER,requested_by INTEGER,requester_org_unit_id INTEGER,next_approver_user_id INTEGER,title TEXT,description TEXT,quantity REAL DEFAULT 0,unit TEXT,amount REAL DEFAULT 0,payload_json TEXT DEFAULT '{}',attachment_file TEXT,attachment_name TEXT,status TEXT DEFAULT 'SUBMITTED',approved_by INTEGER,approved_at TEXT,rejected_by INTEGER,rejected_at TEXT,rejection_reason TEXT,registered_table TEXT,registered_id INTEGER,created_at TEXT DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS request_steps(id INTEGER PRIMARY KEY,request_id INTEGER,step_order INTEGER,stage TEXT,assigned_user_id INTEGER,to_org_unit_id INTEGER,department TEXT,status TEXT DEFAULT 'PENDING',action TEXT,comments TEXT,acted_at TEXT,created_at TEXT DEFAULT CURRENT_TIMESTAMP,UNIQUE(request_id,step_order));
    CREATE TABLE IF NOT EXISTS project_responsibilities(id INTEGER PRIMARY KEY,project_id INTEGER,user_id INTEGER,responsibility_area TEXT DEFAULT 'General Project',source TEXT DEFAULT 'Manual',assigned_by INTEGER,active INTEGER DEFAULT 1,created_at TEXT DEFAULT CURRENT_TIMESTAMP,UNIQUE(project_id,user_id,responsibility_area));
    CREATE TABLE IF NOT EXISTS personnel_project_contacts(id INTEGER PRIMARY KEY,project_id INTEGER NOT NULL,head_office_user_id INTEGER NOT NULL,project_user_id INTEGER NOT NULL,responsibility_area TEXT DEFAULT 'General Project',active INTEGER DEFAULT 1,assigned_by INTEGER,created_at TEXT DEFAULT CURRENT_TIMESTAMP,UNIQUE(project_id,head_office_user_id,project_user_id,responsibility_area));
    ''')
    existing_users=[r['name'] for r in c.execute("PRAGMA table_info(users)").fetchall()]
    if 'personnel_scope' not in existing_users: c.execute("ALTER TABLE users ADD COLUMN personnel_scope TEXT DEFAULT 'PROJECT'")
    # Existing accounts are classified from their Head Office assignment; project-only accounts remain project personnel.
    c.execute("UPDATE users SET personnel_scope='HEAD_OFFICE' WHERE org_unit_id IS NOT NULL AND (personnel_scope IS NULL OR personnel_scope='PROJECT')")
    c.execute("UPDATE project_assignments SET active=0 WHERE user_id IN (SELECT id FROM users WHERE personnel_scope='HEAD_OFFICE')")
    c.execute("DELETE FROM user_projects WHERE user_id IN (SELECT id FROM users WHERE personnel_scope='HEAD_OFFICE')")
    existing_rr=[r['name'] for r in c.execute("PRAGMA table_info(resource_requests)").fetchall()]
    for col,typ in [('current_stage',"TEXT DEFAULT 'PROJECT'"),('origin_scope',"TEXT DEFAULT 'PROJECT'"),('head_office_sent_at','TEXT'),('finalized_at','TEXT')]:
        if col not in existing_rr: c.execute(f"ALTER TABLE resource_requests ADD COLUMN {col} {typ}")
    # Safe migrations for databases created by earlier BAGC versions.
    existing_bq=[r['name'] for r in c.execute("PRAGMA table_info(boq)").fetchall()]
    if 'series' not in existing_bq: c.execute("ALTER TABLE boq ADD COLUMN series TEXT DEFAULT ''")
    if 'title' not in existing_bq: c.execute("ALTER TABLE boq ADD COLUMN title TEXT DEFAULT ''")
    existing_sr=[r['name'] for r in c.execute("PRAGMA table_info(saved_reports)").fetchall()]
    if 'source_report_ids' not in existing_sr: c.execute("ALTER TABLE saved_reports ADD COLUMN source_report_ids TEXT DEFAULT '[]'")
    existing_u=[r['name'] for r in c.execute("PRAGMA table_info(users)").fetchall()]
    for col,typ in [('phone','TEXT'),('email','TEXT'),('last_login','TEXT'),('org_unit_id','INTEGER'),('reports_to_user_id','INTEGER')]:
        if col not in existing_u: c.execute(f"ALTER TABLE users ADD COLUMN {col} {typ}")
    if 'position' not in existing_u: c.execute("ALTER TABLE users ADD COLUMN position TEXT")
    if 'staff_id' not in existing_u: c.execute("ALTER TABLE users ADD COLUMN staff_id TEXT")
    if 'photo_filename' not in existing_u: c.execute("ALTER TABLE users ADD COLUMN photo_filename TEXT")
    # Workflow / project hierarchy migrations.
    existing_up=[r['name'] for r in c.execute("PRAGMA table_info(user_projects)").fetchall()]
    # project_assignments is additive and preserves legacy user_projects access.
    for ur in c.execute("SELECT user_id,project_id FROM user_projects").fetchall():
        c.execute("INSERT OR IGNORE INTO project_assignments(user_id,project_id,position) VALUES(?,?,?)",(ur['user_id'],ur['project_id'],''))

    for ur in c.execute("SELECT id,department,staff_id FROM users").fetchall():
        if not ur['staff_id'] or str(ur['staff_id']).startswith('BAGC-') and str(ur['staff_id'])[5:].isdigit(): c.execute("UPDATE users SET staff_id=? WHERE id=?",(make_staff_id(ur['department'],ur['id']),ur['id']))
    existing_dw=[r['name'] for r in c.execute("PRAGMA table_info(daily_work)").fetchall()]
    if 'unit' not in existing_dw: c.execute("ALTER TABLE daily_work ADD COLUMN unit TEXT")
    existing_da=[r['name'] for r in c.execute("PRAGMA table_info(daily_activities)").fetchall()]
    if 'unit' not in existing_da: c.execute("ALTER TABLE daily_activities ADD COLUMN unit TEXT")
    existing=[r['name'] for r in c.execute("PRAGMA table_info(machines)").fetchall()]
    for col,typ in [('plate_no','TEXT'),('engine_no','TEXT'),('fuel_price','REAL DEFAULT 0')]:
        if col not in existing: c.execute(f"ALTER TABLE machines ADD COLUMN {col} {typ}")
    existing_mp=[r['name'] for r in c.execute("PRAGMA table_info(manpower)").fetchall()]
    if 'crew_id' not in existing_mp: c.execute("ALTER TABLE manpower ADD COLUMN crew_id INTEGER")
    for col,typ in [('working_hours','REAL DEFAULT 8'),('hourly_rate','REAL DEFAULT 0'),('normal_ot_hours','REAL DEFAULT 0'),('normal_ot_rate','REAL DEFAULT 0'),('night_ot_hours','REAL DEFAULT 0'),('night_ot_rate','REAL DEFAULT 0'),('sunday_ot_hours','REAL DEFAULT 0'),('sunday_ot_rate','REAL DEFAULT 0'),('holiday_ot_hours','REAL DEFAULT 0'),('holiday_ot_rate','REAL DEFAULT 0'),('overtime_hours','REAL DEFAULT 0'),('overtime_rate','REAL DEFAULT 0')]:
        if col not in existing_mp: c.execute(f"ALTER TABLE manpower ADD COLUMN {col} {typ}")
    existing_ml=[r['name'] for r in c.execute("PRAGMA table_info(machine_logs)").fetchall()]
    if 'idle_payable' not in existing_ml: c.execute("ALTER TABLE machine_logs ADD COLUMN idle_payable INTEGER DEFAULT 0")
    existing_f=[r['name'] for r in c.execute("PRAGMA table_info(fuel_logs)").fetchall()]
    if 'source' not in existing_f: c.execute("ALTER TABLE fuel_logs ADD COLUMN source TEXT DEFAULT 'Fuel Register'")
    existing_m=[r['name'] for r in c.execute("PRAGMA table_info(machines)").fetchall()]
    for col,typ in [('assignment_start_date','TEXT'),('assignment_start_hour','REAL DEFAULT 0'),('assignment_end_date','TEXT'),('assignment_end_hour','REAL'),('total_signed_hours','REAL DEFAULT 0'),('hours_used','REAL DEFAULT 0'),('lifecycle_status',"TEXT DEFAULT 'ACTIVE'"),('assignment_signed_by','INTEGER'),('assignment_ended_by','INTEGER'),('assignment_ended_at','TEXT')]:
        if col not in existing_m: c.execute(f"ALTER TABLE machines ADD COLUMN {col} {typ}")
    existing_ma=[r['name'] for r in c.execute("PRAGMA table_info(machine_assignments)").fetchall()]
    for col,typ in [('total_signed_hours','REAL DEFAULT 0'),('hours_used','REAL DEFAULT 0')]:
        if col not in existing_ma: c.execute(f"ALTER TABLE machine_assignments ADD COLUMN {col} {typ}")
    # Final schema verification for legacy deployments. Never fail because a column already exists.
    existing_ma=[r['name'] for r in c.execute("PRAGMA table_info(machine_assignments)").fetchall()]
    if 'total_signed_hours' not in existing_ma: c.execute("ALTER TABLE machine_assignments ADD COLUMN total_signed_hours REAL DEFAULT 0")
    if 'hours_used' not in existing_ma: c.execute("ALTER TABLE machine_assignments ADD COLUMN hours_used REAL DEFAULT 0")
    existing_m=[r['name'] for r in c.execute("PRAGMA table_info(machines)").fetchall()]
    if 'rate_unit' not in existing_m: c.execute("ALTER TABLE machines ADD COLUMN rate_unit TEXT DEFAULT 'hr'")
    for col,typ in [('assignment_start_date','TEXT'),('assignment_start_hour','REAL DEFAULT 0'),('assignment_end_date','TEXT'),('assignment_end_hour','REAL'),('total_signed_hours','REAL DEFAULT 0'),('hours_used','REAL DEFAULT 0'),('lifecycle_status',"TEXT DEFAULT 'ACTIVE'"),('assignment_signed_by','INTEGER'),('assignment_ended_by','INTEGER'),('assignment_ended_at','TEXT')]:
        if col not in existing_m: c.execute(f"ALTER TABLE machines ADD COLUMN {col} {typ}")
    # New V26 responsibility/request workflow tables are created above; normalize legacy hierarchy into explicit responsibility links.
    for r in c.execute("SELECT id,reports_to_user_id,org_unit_id FROM users WHERE active=1 AND reports_to_user_id IS NOT NULL").fetchall():
        c.execute("INSERT OR IGNORE INTO responsibilities(supervisor_user_id,subordinate_user_id,scope_type,project_id,source) VALUES(?,?,?,NULL,?)",(r['reports_to_user_id'],r['id'],'HEAD_OFFICE','Hierarchy'))
    for r in c.execute("SELECT user_id,project_id,manager_user_id FROM project_assignments WHERE active=1 AND manager_user_id IS NOT NULL").fetchall():
        c.execute("INSERT OR IGNORE INTO responsibilities(supervisor_user_id,subordinate_user_id,scope_type,project_id,source) VALUES(?,?,?,?,?)",(r['manager_user_id'],r['user_id'],'PROJECT',r['project_id'],'Project Assignment'))

    for g in CREW_GROUPS: c.execute("INSERT OR IGNORE INTO crew_groups(name) VALUES(?)",(g,))
    for pos in POSITION_CATALOG: c.execute("INSERT OR IGNORE INTO crew_positions(name) VALUES(?)",(pos,))
    for g in CREW_GROUPS: c.execute("INSERT OR IGNORE INTO crew_group_capacity(group_name) VALUES(?)",(g,))
    existing_p=[r['name'] for r in c.execute("PRAGMA table_info(projects)").fetchall()]
    for col,typ in [('contractor_role',"TEXT DEFAULT 'Main Contractor'"),('contract_sign_date','TEXT'),('commencement_date','TEXT'),('contract_end_date','TEXT'),('contract_days','INTEGER DEFAULT 0'),('planned_income','REAL DEFAULT 0'),('planned_physical_pct','REAL DEFAULT 0'),('contract_value','REAL DEFAULT 0')]:
        if col not in existing_p: c.execute(f"ALTER TABLE projects ADD COLUMN {col} {typ}")
    # Head Office organization structure. Existing records are preserved.
    org_ids={}
    for idx,(name,parent_name,unit_type) in enumerate(HEAD_OFFICE_STRUCTURE,1):
        parent_id=org_ids.get(parent_name)
        c.execute("INSERT OR IGNORE INTO org_units(name,parent_id,unit_type,sort_order) VALUES(?,?,?,?)",(name,parent_id,unit_type,idx))
        row=c.execute("SELECT id FROM org_units WHERE name=?",(name,)).fetchone(); org_ids[name]=row["id"]
        # ENV-controlled Super Admin synchronization fixes an already-created SQLite DB.
    u=os.environ.get("ADMIN_USERNAME","admin").strip() or "admin"
    p=os.environ.get("ADMIN_PASSWORD","admin123")
    admin=c.execute("SELECT id FROM users WHERE role='SUPER_ADMIN' ORDER BY id LIMIT 1").fetchone()
    if not admin:
        c.execute("INSERT INTO users(full_name,username,password_hash,department,position,location,role) VALUES(?,?,?,?,?,?,?)",("System Administrator",u,generate_password_hash(p),"Administration","Super Admin","Head Office","SUPER_ADMIN"))
    else:
        c.execute("UPDATE users SET username=?,password_hash=?,active=1,department='Administration',position=COALESCE(position,'Super Admin'),location='Head Office' WHERE id=?",(u,generate_password_hash(p),admin["id"]))
    if not c.execute("SELECT id FROM projects").fetchone():
        c.execute("INSERT INTO projects(name,code,location,status) VALUES(?,?,?,?)",("Koye Feche","KOYE","Koye Feche","Active"))
    for ur in c.execute("SELECT id,department FROM users WHERE staff_id IS NULL OR staff_id=''").fetchall(): c.execute("UPDATE users SET staff_id=? WHERE id=?",(make_staff_id(ur['department'],ur['id']),ur['id']))
    c.commit();c.close()


def current_user():
    if not session.get("user_id"): return None
    c=db(); u=c.execute("SELECT * FROM users WHERE id=? AND active=1",(session["user_id"],)).fetchone(); c.close(); return u

@app.context_processor
def inject():
    return {"me":current_user(),"machine_types":MACHINE_TYPES,"unit_catalog":UNIT_CATALOG,"material_categories":MATERIAL_CATEGORIES,"material_catalog":MATERIAL_CATALOG,"design_statuses":DESIGN_STATUSES,"today":dt.date.today().isoformat(),"crew_groups":CREW_GROUPS,"position_catalog":POSITION_CATALOG,"request_types":REQUEST_TYPES,"request_categories":REQUEST_CATEGORIES,"can_module":can_module}


def login_required(f):
    @wraps(f)
    def w(*a,**k):
        if not current_user(): return redirect(url_for("login"))
        return f(*a,**k)
    return w


def admin_required(f):
    @wraps(f)
    def w(*a,**k):
        u=current_user()
        if not u or u["role"]!="SUPER_ADMIN":
            flash("👑 Super Admin access required.","error"); return redirect(url_for("dashboard"))
        return f(*a,**k)
    return w


def allowed_project(pid):
    u=current_user()
    if not u:return False
    if u["role"]=="SUPER_ADMIN":return True
    c=db()
    ok=c.execute("SELECT 1 FROM user_projects WHERE user_id=? AND project_id=?",(u["id"],pid)).fetchone()
    if not ok and (u["personnel_scope"] or "PROJECT")=="HEAD_OFFICE":
        ok=c.execute("SELECT 1 FROM project_responsibilities WHERE user_id=? AND project_id=? AND active=1",(u["id"],pid)).fetchone()
    c.close(); return bool(ok)


def project_admin(pid):
    u=current_user()
    if not u: return False
    if u["role"]=="SUPER_ADMIN": return True
    c=db(); row=c.execute("SELECT 1 FROM project_assignments WHERE user_id=? AND project_id=? AND active=1 AND lower(position) LIKE '%project manager%'",(u["id"],pid)).fetchone(); c.close()
    return bool(row)


def can_module(module):
    u=current_user()
    if not u:return False
    if u["role"]=="SUPER_ADMIN":return True
    return MODULE_DEPARTMENTS.get(module)==u["department"]


def responsibility_users(user_id, scope_type=None, project_id=None):
    c=db()
    q="SELECT r.*,u.full_name,u.staff_id,u.department,u.position,p.name project_name,o.name org_unit_name FROM responsibilities r JOIN users u ON u.id=r.subordinate_user_id LEFT JOIN projects p ON p.id=r.project_id LEFT JOIN org_units o ON o.id=u.org_unit_id WHERE r.supervisor_user_id=? AND r.active=1"
    args=[user_id]
    if scope_type: q+=" AND r.scope_type=?"; args.append(scope_type)
    if project_id is not None: q+=" AND r.project_id=?"; args.append(project_id)
    rows=c.execute(q+" ORDER BY r.scope_type,p.name,u.full_name",tuple(args)).fetchall(); c.close(); return rows

def user_can_approve_request(me, row):
    if me["role"]=="SUPER_ADMIN": return True
    return row["next_approver_user_id"]==me["id"]

def find_project_person(c,pid,term):
    rows=c.execute("SELECT pa.user_id,pa.position FROM project_assignments pa JOIN users u ON u.id=pa.user_id WHERE pa.project_id=? AND pa.active=1 AND u.active=1 ORDER BY pa.id",(pid,)).fetchall()
    for r in rows:
        if term.lower() in (r['position'] or '').lower(): return r['user_id']
    return None


def project_request_steps(c,pid,requester_id):
    out=[]; used=set()
    for stage,term in [('OFFICE_ENGINEER','Office Engineer'),('OFFICE_HEAD','Office Head'),('PROJECT_MANAGER','Project Manager')]:
        uid=find_project_person(c,pid,term)
        if uid and uid!=requester_id and uid not in used: out.append((stage,uid)); used.add(uid)
    return out


def current_request_step(c,rid):
    return c.execute("SELECT * FROM request_steps WHERE request_id=? AND status='PENDING' ORDER BY step_order LIMIT 1",(rid,)).fetchone()


def request_approver(c, requester_id, project_id=None):
    # Project requests stay under the project chain; Head Office requests stay under the Head Office chain.
    if project_id:
        pa=c.execute("SELECT manager_user_id FROM project_assignments WHERE user_id=? AND project_id=? AND active=1",(requester_id,project_id)).fetchone()
        if pa and pa["manager_user_id"]: return pa["manager_user_id"]
    r=c.execute("SELECT reports_to_user_id FROM users WHERE id=?",(requester_id,)).fetchone()
    if r and r["reports_to_user_id"]:
        return r["reports_to_user_id"]
    return c.execute("SELECT id FROM users WHERE role='SUPER_ADMIN' AND active=1 ORDER BY id LIMIT 1").fetchone()["id"]


def parse_float(v):
    try:return float(v or 0)
    except:return 0.0


def period_bounds(period, anchor):
    d=dt.date.fromisoformat(anchor)
    if period=="day": return d,d
    if period=="week":
        start=d-dt.timedelta(days=d.weekday()); return start,start+dt.timedelta(days=6)
    start=d.replace(day=1); end=d.replace(day=calendar.monthrange(d.year,d.month)[1]); return start,end


def money(v):return round(float(v or 0),2)

def snapshot_json(obj):
    return json.dumps(obj, default=lambda x: dict(x) if isinstance(x, sqlite3.Row) else str(x))

def report_dates(report_type, start, end):
    if start and end: return dt.date.fromisoformat(start), dt.date.fromisoformat(end)
    today=dt.date.today()
    if report_type=='DAILY': return today,today
    if report_type=='WEEKLY': return today-dt.timedelta(days=today.weekday()),today
    if report_type=='MONTHLY': return today.replace(day=1),today
    if report_type=='SEMI_ANNUAL': return (dt.date(today.year,1,1),dt.date(today.year,6,30)) if today.month<=6 else (dt.date(today.year,7,1),dt.date(today.year,12,31))
    if report_type=='ANNUAL': return dt.date(today.year,1,1),dt.date(today.year,12,31)
    return today,today

def allowed_report_scopes_for_user(pid=None):
    u=current_user()
    if not u: return set()
    if u["role"]=="SUPER_ADMIN" or (pid is not None and project_admin(pid)): return {"ALL","BOQ","MACHINERY","MANPOWER","STORE","FUEL","FINANCE","PROBLEMS","DESIGN"}
    return set(REPORT_SCOPES_BY_DEPARTMENT.get(u["department"],set()))

def build_report_snapshot(pid,start,end,scope='ALL'):
    c=db(); out={'project_id':pid,'start_date':start.isoformat(),'end_date':end.isoformat(),'scope':scope}
    if scope in ('ALL','BOQ'):
        rows=c.execute("SELECT b.*,COALESCE(SUM(CASE WHEN dw.date<? THEN dw.quantity ELSE 0 END),0) previous_qty,COALESCE(SUM(CASE WHEN dw.date BETWEEN ? AND ? THEN dw.quantity ELSE 0 END),0) period_qty,COALESCE(SUM(dw.quantity),0) todate_qty FROM boq b LEFT JOIN daily_work dw ON dw.boq_id=b.id WHERE b.project_id=? GROUP BY b.id ORDER BY b.item_no",(start.isoformat(),start.isoformat(),end.isoformat(),pid)).fetchall()
        out['boq']=[dict(r,previous_amount=r['previous_qty']*r['rate'],period_amount=r['period_qty']*r['rate'],todate_amount=r['todate_qty']*r['rate']) for r in rows]
    if scope in ('ALL','MACHINERY'):
        out['machinery']=[dict(r) for r in c.execute("SELECT ml.*,m.machine_type,m.code,m.plate_no,m.engine_no,m.ownership,m.hourly_rate,m.rate_unit,m.expected_fuel,(CASE WHEN m.rate_unit='day' THEN CASE WHEN (ml.work_hours+ml.idle_hours+ml.down_hours)>0 THEN m.hourly_rate ELSE 0 END WHEN m.rate_unit='month' THEN CASE WHEN (ml.work_hours+ml.idle_hours+ml.down_hours)>0 THEN m.hourly_rate/30.0 ELSE 0 END ELSE (ml.work_hours+CASE WHEN ml.idle_payable=1 THEN ml.idle_hours ELSE 0 END)*m.hourly_rate END) expense,(ml.opening_gauge+ml.fuel_received-ml.closing_gauge) actual_fuel,(ml.work_hours*m.expected_fuel) expected_fuel FROM machine_logs ml JOIN machines m ON m.id=ml.machine_id WHERE ml.project_id=? AND ml.date BETWEEN ? AND ? ORDER BY ml.date,ml.id",(pid,start.isoformat(),end.isoformat())).fetchall()]
    if scope in ('ALL','MANPOWER'):
        out['manpower']=[dict(r) for r in c.execute("SELECT * FROM manpower WHERE project_id=? AND date BETWEEN ? AND ? ORDER BY date,id",(pid,start.isoformat(),end.isoformat())).fetchall()]
    if scope in ('ALL','STORE'):
        out['store']=[dict(r) for r in c.execute("SELECT sl.*,m.name,m.category,m.unit FROM store_logs sl JOIN materials m ON m.id=sl.material_id WHERE sl.project_id=? AND sl.date BETWEEN ? AND ? ORDER BY sl.date,sl.id",(pid,start.isoformat(),end.isoformat())).fetchall()]
    if scope in ('ALL','FUEL'):
        out['fuel']=[dict(r) for r in c.execute("SELECT f.*,m.machine_type,m.code,m.plate_no,m.engine_no,m.ownership,m.expected_fuel,COALESCE((SELECT SUM(ml.work_hours) FROM machine_logs ml WHERE ml.machine_id=f.machine_id AND ml.date=f.date),0) work_hours,(f.opening_gauge+f.fuel_received-f.closing_gauge) consumption,(f.fuel_received*f.fuel_price) cost FROM fuel_logs f JOIN machines m ON m.id=f.machine_id WHERE f.project_id=? AND f.date BETWEEN ? AND ? ORDER BY f.date,f.id",(pid,start.isoformat(),end.isoformat())).fetchall()]
    if scope in ('ALL','DESIGN'):
        out['design']=[dict(r) for r in c.execute("SELECT * FROM design_items WHERE project_id=? AND (submitted IS NULL OR submitted BETWEEN ? AND ?) ORDER BY id DESC",(pid,start.isoformat(),end.isoformat())).fetchall()]
    if scope in ('ALL','FINANCE'):
        out['finance']=[dict(r) for r in c.execute("SELECT * FROM finance_logs WHERE project_id=? AND date BETWEEN ? AND ? ORDER BY date,id",(pid,start.isoformat(),end.isoformat())).fetchall()]
    if scope in ('ALL','PROBLEMS'):
        out['variations']=[dict(r) for r in c.execute("SELECT va.*,b.item_no,b.description,b.unit FROM variation_alerts va JOIN boq b ON b.id=va.boq_id WHERE va.project_id=? AND va.date BETWEEN ? AND ? ORDER BY va.date,va.id",(pid,start.isoformat(),end.isoformat())).fetchall()]
        out['problems']=[dict(r) for r in c.execute("SELECT * FROM problems WHERE project_id=? AND date BETWEEN ? AND ? ORDER BY date,id",(pid,start.isoformat(),end.isoformat())).fetchall()]
    # summary totals, always useful for every saved report
    out['summary']={
        'income': c.execute("SELECT COALESCE(SUM(dw.quantity*b.rate),0) FROM daily_work dw JOIN boq b ON b.id=dw.boq_id WHERE dw.project_id=? AND dw.date BETWEEN ? AND ?",(pid,start.isoformat(),end.isoformat())).fetchone()[0],
        'machinery_expense': c.execute("SELECT COALESCE(SUM(CASE WHEN m.rate_unit='day' THEN CASE WHEN (ml.work_hours+ml.idle_hours+ml.down_hours)>0 THEN m.hourly_rate ELSE 0 END WHEN m.rate_unit='month' THEN CASE WHEN (ml.work_hours+ml.idle_hours+ml.down_hours)>0 THEN m.hourly_rate/30.0 ELSE 0 END ELSE (ml.work_hours+CASE WHEN ml.idle_payable=1 THEN ml.idle_hours ELSE 0 END)*m.hourly_rate END),0) FROM machine_logs ml JOIN machines m ON m.id=ml.machine_id WHERE ml.project_id=? AND ml.date BETWEEN ? AND ?",(pid,start.isoformat(),end.isoformat())).fetchone()[0],
        'manpower_expense': c.execute("SELECT COALESCE(SUM(CASE WHEN hourly_rate>0 THEN present*working_hours*hourly_rate ELSE present*daily_rate END+normal_ot_hours*normal_ot_rate+night_ot_hours*night_ot_rate+sunday_ot_hours*sunday_ot_rate+holiday_ot_hours*holiday_ot_rate),0) FROM manpower WHERE project_id=? AND date BETWEEN ? AND ?",(pid,start.isoformat(),end.isoformat())).fetchone()[0],
        'store_expense': c.execute("SELECT COALESCE(SUM(issued*unit_cost),0) FROM store_logs WHERE project_id=? AND date BETWEEN ? AND ?",(pid,start.isoformat(),end.isoformat())).fetchone()[0],
        'other_expense': c.execute("SELECT COALESCE(SUM(amount),0) FROM finance_logs WHERE project_id=? AND kind='Expense' AND date BETWEEN ? AND ?",(pid,start.isoformat(),end.isoformat())).fetchone()[0],
        'fuel_cost': c.execute("SELECT COALESCE(SUM(fuel_received*fuel_price),0) FROM fuel_logs WHERE project_id=? AND date BETWEEN ? AND ?",(pid,start.isoformat(),end.isoformat())).fetchone()[0]
    }
    c.close(); return out

def save_report(pid, report_type, start, end, scope='ALL', user_id=None):
    snap=build_report_snapshot(pid,start,end,scope); c=db(); existing=c.execute("SELECT id FROM saved_reports WHERE project_id=? AND report_type=? AND scope=? AND start_date=? AND end_date=?",(pid,report_type,scope,start.isoformat(),end.isoformat())).fetchone()
    if existing:
        rid=existing['id']; c.execute("UPDATE saved_reports SET snapshot_json=?,generated_by=?,generated_at=CURRENT_TIMESTAMP WHERE id=?",(snapshot_json(snap),user_id or session.get('user_id'),rid))
    else:
        n=c.execute("SELECT COUNT(*) FROM saved_reports WHERE project_id=? AND report_type=?",(pid,report_type)).fetchone()[0]+1; no=f"{report_type[:3]}-{dt.date.today().year}-{n:04d}"
        c.execute("INSERT INTO saved_reports(project_id,report_no,report_type,scope,start_date,end_date,generated_by,snapshot_json) VALUES(?,?,?,?,?,?,?,?)",(pid,no,report_type,scope,start.isoformat(),end.isoformat(),user_id or session.get('user_id'),snapshot_json(snap))); rid=c.execute("SELECT last_insert_rowid()").fetchone()[0]
    # Explicit lineage: higher reports retain the saved lower-level reports they consolidate.
    source_types={'WEEKLY':['DAILY'],'MONTHLY':['DAILY','WEEKLY'],'SEMI_ANNUAL':['MONTHLY','WEEKLY'],'ANNUAL':['SEMI_ANNUAL','MONTHLY']} .get(report_type,[])
    src=[]
    for rt in source_types:
        src += [x['id'] for x in c.execute("SELECT id FROM saved_reports WHERE project_id=? AND report_type=? AND start_date<=? AND end_date>=? ORDER BY start_date",(pid,rt,end.isoformat(),start.isoformat())).fetchall()]
    c.execute("UPDATE saved_reports SET source_report_ids=? WHERE id=?",(json.dumps(src),rid)); c.commit(); c.close(); return rid



def dashboard_data(pid=None):
    c=db(); where="" if pid is None else " WHERE p.id=?"; args=() if pid is None else (pid,)
    projects=c.execute("SELECT p.* FROM projects p"+where+" ORDER BY p.name",args).fetchall()
    out=[]
    for p in projects:
        inc=c.execute("SELECT COALESCE(SUM(dw.quantity*b.rate),0) x FROM daily_work dw JOIN boq b ON b.id=dw.boq_id WHERE dw.project_id=?",(p["id"],)).fetchone()["x"]
        me=c.execute("SELECT COALESCE(SUM(CASE WHEN m.rate_unit='day' THEN CASE WHEN (ml.work_hours+ml.idle_hours+ml.down_hours)>0 THEN m.hourly_rate ELSE 0 END WHEN m.rate_unit='month' THEN CASE WHEN (ml.work_hours+ml.idle_hours+ml.down_hours)>0 THEN m.hourly_rate/30.0 ELSE 0 END ELSE (ml.work_hours+CASE WHEN ml.idle_payable=1 THEN ml.idle_hours ELSE 0 END)*m.hourly_rate END),0) x FROM machine_logs ml JOIN machines m ON m.id=ml.machine_id WHERE ml.project_id=?",(p["id"],)).fetchone()["x"]
        pe=c.execute("SELECT COALESCE(SUM((CASE WHEN mp.hourly_rate>0 THEN mp.present*mp.working_hours*mp.hourly_rate ELSE mp.present*mp.daily_rate END + mp.normal_ot_hours*mp.normal_ot_rate + mp.night_ot_hours*mp.night_ot_rate + mp.sunday_ot_hours*mp.sunday_ot_rate + mp.holiday_ot_hours*mp.holiday_ot_rate)),0) x FROM manpower mp WHERE mp.project_id=?",(p["id"],)).fetchone()["x"]
        se=c.execute("SELECT COALESCE(SUM(sl.issued*sl.unit_cost),0) x FROM store_logs sl WHERE sl.project_id=?",(p["id"],)).fetchone()["x"]
        other=c.execute("SELECT COALESCE(SUM(amount),0) x FROM finance_logs WHERE project_id=? AND kind='Expense'",(p["id"],)).fetchone()["x"]
        workers=c.execute("SELECT COUNT(*) x FROM manpower WHERE project_id=? AND present>0",(p["id"],)).fetchone()["x"]
        machines=c.execute("SELECT COUNT(*) x FROM machines WHERE project_id=? AND active=1",(p["id"],)).fetchone()["x"]
        total_exp=money(me+pe+se+other)
        daily_m=c.execute("SELECT ml.*,m.machine_type,m.code,m.plate_no,m.engine_no,m.ownership,((CASE WHEN m.rate_unit='day' THEN CASE WHEN (ml.work_hours+ml.idle_hours+ml.down_hours)>0 THEN m.hourly_rate ELSE 0 END WHEN m.rate_unit='month' THEN CASE WHEN (ml.work_hours+ml.idle_hours+ml.down_hours)>0 THEN m.hourly_rate/30.0 ELSE 0 END ELSE (ml.work_hours+CASE WHEN ml.idle_payable=1 THEN ml.idle_hours ELSE 0 END)*m.hourly_rate END)) expense,(ml.opening_gauge+ml.fuel_received-ml.closing_gauge) actual_fuel FROM machine_logs ml JOIN machines m ON m.id=ml.machine_id WHERE ml.project_id=? ORDER BY ml.date DESC,ml.id DESC LIMIT 8",(p["id"],)).fetchall()
        daily_mat=c.execute("SELECT sl.*,m.name,m.category,m.unit FROM store_logs sl JOIN materials m ON m.id=sl.material_id WHERE sl.project_id=? ORDER BY sl.date DESC,sl.id DESC LIMIT 8",(p["id"],)).fetchall()
        fuel_recent=c.execute("SELECT f.*,m.machine_type,m.code,m.plate_no,m.engine_no,m.ownership,(f.opening_gauge+f.fuel_received-f.closing_gauge) consumption,(f.fuel_received*f.fuel_price) cost FROM fuel_logs f JOIN machines m ON m.id=f.machine_id WHERE f.project_id=? ORDER BY f.date DESC,f.id DESC LIMIT 8",(p["id"],)).fetchall()
        contract_value=p["contract_value"] or c.execute("SELECT COALESCE(SUM(contract_qty*rate),0) FROM boq WHERE project_id=?",(p["id"],)).fetchone()[0]
        physical_pct=(inc/contract_value*100) if contract_value else 0
        start=p["commencement_date"] or p["start_date"]; end=p["contract_end_date"] or p["end_date"]; schedule_pct=0; time_var=0; days_remaining=None
        if start and end:
            sd=dt.date.fromisoformat(start); ed=dt.date.fromisoformat(end); total=max((ed-sd).days,0); elapsed=max(min((dt.date.today()-sd).days,total),0); schedule_pct=(elapsed/total*100) if total else 0; time_var=physical_pct-schedule_pct; days_remaining=(ed-dt.date.today()).days
        out.append({"p":p,"income":money(inc),"expense":total_exp,"expense_pct":money((total_exp/(inc+0.0001))*100) if inc else 0,"machine_expense":money(me),"manpower_expense":money(pe),"store_expense":money(se),"other_expense":money(other),"workers":workers,"machines":machines,"daily_machines":daily_m,"daily_materials":daily_mat,"fuel_recent":fuel_recent,"contract_value":money(contract_value),"physical_pct":money(physical_pct),"schedule_pct":money(schedule_pct),"time_variance_pct":money(time_var),"days_remaining":days_remaining,"planned_income":money(p["planned_income"] or 0)})
    company_exp=sum(x["expense"] for x in out) or 1
    out=[dict(x,company_expense_pct=money((x["expense"]/company_exp)*100)) for x in out]
    c.close();return out

@app.route("/")
def home():
    # Render the authenticated dashboard directly from / so the post-login
    # flow does not depend on a second redirect that can produce a 404.
    u=current_user()
    if not u:
        return redirect(url_for("login"))
    c=db()
    if u["role"]=="SUPER_ADMIN":
        projects=c.execute("SELECT p.* FROM projects p ORDER BY p.name").fetchall()
    else:
        projects=c.execute("SELECT p.* FROM projects p JOIN user_projects up ON up.project_id=p.id WHERE up.user_id=? ORDER BY p.name",(u["id"],)).fetchall()
    c.close()
    data=dashboard_data(); allowed_ids={p["id"] for p in projects}; data=[x for x in data if x["p"]["id"] in allowed_ids]
    totals={k:sum(x[k] for x in data) for k in ["income","expense","machine_expense","manpower_expense","store_expense","other_expense"]}
    c=db()
    if u["role"]=="SUPER_ADMIN":
        org_units=c.execute("SELECT o.*,p.name parent_name,mu.full_name manager_name FROM org_units o LEFT JOIN org_units p ON p.id=o.parent_id LEFT JOIN users mu ON mu.id=o.manager_user_id WHERE o.active=1 ORDER BY o.sort_order,o.name").fetchall()
        staff=c.execute("SELECT u.*,o.name org_name,m.full_name manager_name FROM users u LEFT JOIN org_units o ON o.id=u.org_unit_id LEFT JOIN users m ON m.id=u.reports_to_user_id WHERE u.active=1 ORDER BY o.sort_order,o.name,u.full_name").fetchall()
    else:
        org_units=c.execute("SELECT o.*,p.name parent_name,mu.full_name manager_name FROM org_units o LEFT JOIN org_units p ON p.id=o.parent_id LEFT JOIN users mu ON mu.id=o.manager_user_id WHERE o.active=1 AND o.id=? ORDER BY o.name",(u["org_unit_id"] or -1,)).fetchall()
        staff=c.execute("SELECT u.*,o.name org_name,m.full_name manager_name FROM users u LEFT JOIN org_units o ON o.id=u.org_unit_id LEFT JOIN users m ON m.id=u.reports_to_user_id WHERE u.active=1 AND (u.id=? OR u.reports_to_user_id=?) ORDER BY u.full_name",(u["id"],u["id"])).fetchall()
    req_total=c.execute("SELECT COUNT(*) n FROM resource_requests WHERE requested_by=? OR next_approver_user_id=?",(u["id"],u["id"])).fetchone()["n"]
    req_pending=c.execute("SELECT COUNT(*) n FROM resource_requests WHERE status='SUBMITTED' AND next_approver_user_id=?",(u["id"],)).fetchone()["n"]
    req_approved=c.execute("SELECT COUNT(*) n FROM resource_requests WHERE status='APPROVED' AND (requested_by=? OR approved_by=?)",(u["id"],u["id"])).fetchone()["n"]
    req_rejected=c.execute("SELECT COUNT(*) n FROM resource_requests WHERE status='REJECTED' AND (requested_by=? OR rejected_by=?)",(u["id"],u["id"])).fetchone()["n"]
    head_resp=c.execute("SELECT COUNT(*) n FROM responsibilities WHERE supervisor_user_id=? AND scope_type='HEAD_OFFICE' AND active=1",(u["id"],)).fetchone()["n"]
    proj_resp=c.execute("SELECT COUNT(*) n FROM responsibilities WHERE supervisor_user_id=? AND scope_type='PROJECT' AND active=1",(u["id"],)).fetchone()["n"]
    # Department-specific dashboard KPIs: personnel see metrics only for their assigned work area.
    if u['role']=='SUPER_ADMIN':
        module_metrics={'records':c.execute("SELECT COUNT(*) FROM resource_requests").fetchone()[0],'files':c.execute("SELECT COUNT(*) FROM workflow_files").fetchone()[0],'machines':c.execute("SELECT COUNT(*) FROM machines WHERE active=1").fetchone()[0],'materials':c.execute("SELECT COUNT(*) FROM materials WHERE active=1").fetchone()[0],'manpower':c.execute("SELECT COUNT(*) FROM manpower").fetchone()[0],'design':c.execute("SELECT COUNT(*) FROM design_items").fetchone()[0],'expenses':c.execute("SELECT COALESCE(SUM(amount),0) FROM finance_logs WHERE kind='Expense'").fetchone()[0]}
    else:
        d=u['department']; module_metrics={'records':req_total,'files':c.execute("SELECT COUNT(*) FROM workflow_files WHERE from_user_id=? OR to_user_id=? OR to_org_unit_id=?",(u['id'],u['id'],u['org_unit_id'] or -1)).fetchone()[0],'machines':0,'materials':0,'manpower':0,'design':0,'expenses':0}
        if d=='Machinery': module_metrics['machines']=c.execute("SELECT COUNT(*) FROM machines WHERE active=1 AND project_id IN (SELECT project_id FROM user_projects WHERE user_id=?)",(u['id'],)).fetchone()[0]
        elif d=='Store': module_metrics['materials']=c.execute("SELECT COUNT(*) FROM materials WHERE active=1 AND project_id IN (SELECT project_id FROM user_projects WHERE user_id=?)",(u['id'],)).fetchone()[0]
        elif d=='HR': module_metrics['manpower']=c.execute("SELECT COUNT(*) FROM manpower WHERE project_id IN (SELECT project_id FROM user_projects WHERE user_id=?)",(u['id'],)).fetchone()[0]
        elif d=='Design': module_metrics['design']=c.execute("SELECT COUNT(*) FROM design_items WHERE project_id IN (SELECT project_id FROM user_projects WHERE user_id=?)",(u['id'],)).fetchone()[0]
        elif d=='Finance': module_metrics['expenses']=c.execute("SELECT COALESCE(SUM(amount),0) FROM finance_logs WHERE kind='Expense' AND project_id IN (SELECT project_id FROM user_projects WHERE user_id=?)",(u['id'],)).fetchone()[0]
    c.close()
    dashboard_workflow={"total":req_total,"pending":req_pending,"approved":req_approved,"rejected":req_rejected,"head_responsibilities":head_resp,"project_responsibilities":proj_resp}
    return render_template("dashboard.html",data=data,totals=totals,org_units=org_units,staff=staff,dashboard_workflow=dashboard_workflow,module_metrics=module_metrics)

@app.route("/login", methods=["GET", "POST"])
def login():
    # Database/schema initialization is performed once at application startup.
    # Re-running migrations on every login can create unnecessary SQLite write contention.
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        c = db()
        u = c.execute("SELECT * FROM users WHERE username=? AND active=1", (username,)).fetchone()
        c.close()
        if u and check_password_hash(u["password_hash"], password):
            session.clear()
            session["user_id"] = int(u["id"])
            session.permanent = True
            lc=db(); lc.execute("UPDATE users SET last_login=CURRENT_TIMESTAMP WHERE id=?",(int(u["id"]),)); lc.commit(); lc.close()
            # Use a normal redirect to the canonical dashboard endpoint.
            # The dashboard route itself rebuilds the project scope from the session.
            return redirect(url_for("dashboard"))
        flash("🔐 Invalid username or password.", "error")
    return render_template("login.html")

@app.route("/logout")
def logout():session.clear();return redirect(url_for("login"))

@app.route("/dashboard")
@login_required
def dashboard():
    return home()

@app.route("/home")
@login_required
def home_alias():
    return home()

@app.route("/projects/<int:pid>")
@login_required
def project(pid):
    if not allowed_project(pid): flash("🚫 You do not have access to this project.","error"); return redirect(url_for("dashboard"))
    c=db(); p=c.execute("SELECT * FROM projects WHERE id=?",(pid,)).fetchone()
    if not p: c.close(); return redirect(url_for("dashboard"))
    boq_count=c.execute("SELECT COUNT(*) n FROM boq WHERE project_id=?",(pid,)).fetchone()["n"]
    machine_count=c.execute("SELECT COUNT(*) n FROM machines WHERE project_id=? AND active=1",(pid,)).fetchone()["n"]
    mat_count=c.execute("SELECT COUNT(*) n FROM materials WHERE project_id=? AND active=1",(pid,)).fetchone()["n"]
    actual_income=c.execute("SELECT COALESCE(SUM(dw.quantity*b.rate),0) x FROM daily_work dw JOIN boq b ON b.id=dw.boq_id WHERE dw.project_id=?",(pid,)).fetchone()["x"]
    actual_expense=c.execute("SELECT COALESCE(SUM((ml.work_hours + CASE WHEN ml.idle_payable=1 THEN ml.idle_hours ELSE 0 END)*m.hourly_rate),0) FROM machine_logs ml JOIN machines m ON m.id=ml.machine_id WHERE ml.project_id=?",(pid,)).fetchone()[0]
    actual_expense+=c.execute("SELECT COALESCE(SUM((CASE WHEN mp.hourly_rate>0 THEN mp.present*mp.working_hours*mp.hourly_rate ELSE mp.present*mp.daily_rate END + mp.normal_ot_hours*mp.normal_ot_rate + mp.night_ot_hours*mp.night_ot_rate + mp.sunday_ot_hours*mp.sunday_ot_rate + mp.holiday_ot_hours*mp.holiday_ot_rate)),0) FROM manpower mp WHERE mp.project_id=?",(pid,)).fetchone()[0]
    actual_expense+=c.execute("SELECT COALESCE(SUM(issued*unit_cost),0) FROM store_logs WHERE project_id=?",(pid,)).fetchone()[0]
    actual_expense+=c.execute("SELECT COALESCE(SUM(amount),0) FROM finance_logs WHERE project_id=? AND kind='Expense'",(pid,)).fetchone()[0]
    contract_value=p["contract_value"] or c.execute("SELECT COALESCE(SUM(contract_qty*rate),0) FROM boq WHERE project_id=?",(pid,)).fetchone()[0]
    physical_pct=(actual_income/contract_value*100) if contract_value else 0
    today=dt.date.today(); start=p["commencement_date"] or p["start_date"]; end=p["contract_end_date"] or p["end_date"]
    elapsed=0; schedule_pct=0; time_variance_pct=0; days_remaining=None
    if start and end:
        sd=dt.date.fromisoformat(start); ed=dt.date.fromisoformat(end); total=max((ed-sd).days,0); elapsed=max(min((today-sd).days,total),0); schedule_pct=(elapsed/total*100) if total else 0; time_variance_pct=physical_pct-schedule_pct; days_remaining=(ed-today).days
    planned_income=p["planned_income"] or 0; income_variance=actual_income-planned_income
    crew_count=c.execute("SELECT COUNT(*) n FROM project_crews WHERE project_id=?",(pid,)).fetchone()["n"]
    # Project page also needs the same responsibility/contact context as the Project Team page.
    project_users=c.execute("SELECT u.id,u.full_name,u.department,u.position,o.name org_unit_name FROM users u JOIN project_assignments pa ON pa.user_id=u.id AND pa.project_id=? AND pa.active=1 LEFT JOIN org_units o ON o.id=u.org_unit_id WHERE u.active=1 AND u.personnel_scope='PROJECT' ORDER BY u.full_name",(pid,)).fetchall()
    head_users=c.execute("SELECT u.id,u.full_name,u.department,u.position,o.name org_unit_name FROM users u LEFT JOIN org_units o ON o.id=u.org_unit_id WHERE u.active=1 AND u.personnel_scope='HEAD_OFFICE' ORDER BY u.full_name").fetchall()
    responsibility_map={}
    for area in ['General Project','Store','Machinery','Fuel','Manpower / HR','Finance','Design','Procurement','Project Management','HSE / Safety','QA/QC','Survey','Contract Administration']:
        responsibility_map[area]={r['user_id'] for r in c.execute("SELECT user_id FROM project_responsibilities WHERE project_id=? AND responsibility_area=? AND active=1",(pid,area)).fetchall()}
    c.close()
    return render_template("project.html",pid=pid,p=p,boq_count=boq_count,machine_count=machine_count,mat_count=mat_count,actual_income=actual_income,actual_expense=actual_expense,contract_value=contract_value,physical_pct=physical_pct,schedule_pct=schedule_pct,time_variance_pct=time_variance_pct,days_remaining=days_remaining,planned_income=planned_income,income_variance=income_variance,crew_count=crew_count,project_users=project_users,head_users=head_users,responsibility_map=responsibility_map)


def variation_check(c,pid,boq_id,additional_qty,d):
    b=c.execute("SELECT * FROM boq WHERE id=? AND project_id=?",(boq_id,pid)).fetchone()
    if not b: raise ValueError("Selected BOQ item was not found in this project.")
    prev=c.execute("SELECT COALESCE(SUM(quantity),0) FROM daily_work WHERE boq_id=? AND date<?",(boq_id,d)).fetchone()[0]
    period=c.execute("SELECT COALESCE(SUM(quantity),0) FROM daily_work WHERE boq_id=? AND date BETWEEN ? AND ?",(boq_id,d,d)).fetchone()[0]
    tod=prev+period
    new_tod=tod+additional_qty
    excess=max(new_tod-(b['contract_qty'] or 0),0)
    if excess>0:
        msg=f"Variation Order required: BOQ {b['item_no']} will exceed contract quantity by {excess:g} {b['unit']}."
        c.execute("INSERT INTO variation_alerts(project_id,boq_id,date,contract_qty,previous_qty,period_qty,to_date_qty,excess_qty,status,message,created_by) VALUES(?,?,?,?,?,?,?,?,?,?,?)",(pid,boq_id,d,b['contract_qty'] or 0,prev,period, new_tod, excess,'OPEN',msg,current_user()['id']))
        return msg
    return None

@app.errorhandler(Exception)
def application_error(error):
    # Never expose a generic blank 500 to field staff; log the actual exception and return a useful message.
    import traceback
    app.logger.error("Unhandled BAGC application error: %s\n%s", error, traceback.format_exc())
    if request.path.startswith('/projects/'):
        flash("⚠️ Could not save this entry. Nothing was intentionally deleted. Please check the fields and try again. Error: "+str(error),"error")
        return redirect(request.referrer or url_for('dashboard'))
    return ("Internal Server Error: "+str(error),500)

@app.route("/projects/<int:pid>/daily",methods=["GET","POST"])
@login_required
def daily(pid):
    u=current_user()
    if not allowed_project(pid) or (u["role"]!="SUPER_ADMIN" and u["position"]!="Office Engineer" and not project_admin(pid)):
        flash("🚫 Daily Report is controlled by the Office Engineer; Project Manager has project-admin access. Super Admin has override access.","error")
        return redirect(url_for("project",pid=pid))
    c=db(); default_date=request.args.get("date",dt.date.today().isoformat())
    try:
        if request.method=="POST":
            d=request.form.get("date") or default_date
            action=request.form.get("action","save_daily")
            if action != "save_daily":
                raise ValueError("Invalid Daily Report action.")

            # One transaction: every BOQ work package and every linked resource is saved
            # together. The Office Engineer presses Save only once for the whole day.
            boq_ids=request.form.getlist("wp_boq_id[]")
            work_types=request.form.getlist("wp_work_type[]")
            quantities=request.form.getlist("wp_qty[]")
            station_froms=request.form.getlist("wp_station_from[]")
            station_tos=request.form.getlist("wp_station_to[]")
            remarks=request.form.getlist("wp_remarks[]")
            machine_lists=request.form.getlist("wp_machine_log_ids[]")
            manpower_lists=request.form.getlist("wp_manpower_ids[]")
            crew_lists=request.form.getlist("wp_crew_ids[]")
            store_lists=request.form.getlist("wp_store_log_ids[]")
            fuel_lists=request.form.getlist("wp_fuel_log_ids[]")
            finance_lists=request.form.getlist("wp_finance_log_ids[]")
            evaluations=request.form.getlist("wp_evaluation[]")
            scores=request.form.getlist("wp_score[]")
            eval_remarks=request.form.getlist("wp_evaluation_remarks[]")

            # Browser sends one hidden marker for each row. JSON is used for the multi-selects
            # so a single form can contain an arbitrary number of work packages.
            import json as _json
            machine_data=request.form.getlist("wp_machine_json[]")
            manpower_data=request.form.getlist("wp_manpower_json[]")
            crew_data=request.form.getlist("wp_crew_json[]")
            store_data=request.form.getlist("wp_store_json[]")
            fuel_data=request.form.getlist("wp_fuel_json[]")
            finance_data=request.form.getlist("wp_finance_json[]")
            units=request.form.getlist("wp_unit[]")

            # Backward-compatible fallback for a simple single row POST.
            row_count=len(boq_ids)
            if row_count==0 and request.form.get("boq_id"):
                boq_ids=[request.form.get("boq_id")]; work_types=[request.form.get("work_type","")]; quantities=[request.form.get("executed_qty","0")]
                station_froms=[request.form.get("station_from","")]; station_tos=[request.form.get("station_to","")]; remarks=[request.form.get("remarks","")]
                machine_data=[_json.dumps(request.form.getlist("machine_log_ids"))]; manpower_data=[_json.dumps(request.form.getlist("manpower_ids"))]
                crew_data=[_json.dumps(request.form.getlist("crew_ids"))]; store_data=[_json.dumps(request.form.getlist("store_log_ids"))]
                fuel_data=[_json.dumps(request.form.getlist("fuel_log_ids"))]; finance_data=[_json.dumps(request.form.getlist("finance_log_ids"))]
                evaluations=[request.form.get("evaluation","")]; scores=[request.form.get("score","")]; eval_remarks=[request.form.get("evaluation_remarks","")]
                row_count=1

            if row_count==0:
                raise ValueError("Add at least one BOQ work package before saving the Daily Report.")

            def arr(items,i,default=""):
                return items[i] if i < len(items) else default
            def json_alloc(items,i):
                raw=arr(items,i,"[]")
                try: value=_json.loads(raw) if raw else []
                except Exception: value=[]
                out=[]
                if isinstance(value,dict):
                    for k,v in value.items():
                        if str(k).strip().isdigit(): out.append((int(k), v if isinstance(v,dict) else {"qty":parse_float(v)}))
                elif isinstance(value,list):
                    for x in value:
                        if isinstance(x,dict) and str(x.get("id","")).isdigit(): out.append((int(x["id"]),x))
                        elif str(x).strip().isdigit(): out.append((int(x),{}))
                return out

            saved=0
            for i in range(row_count):
                bid_s=arr(boq_ids,i).strip()
                if not bid_s: continue
                bid=int(bid_s); qty=parse_float(arr(quantities,i))
                if qty<=0: continue
                # Variation warning is raised only when cumulative executed quantity exceeds BOQ quantity.
                b=c.execute("SELECT unit FROM boq WHERE id=? AND project_id=?",(bid,pid)).fetchone()
                selected_unit=arr(units,i,(b["unit"] if b else "")) or (b["unit"] if b else "")
                msg=variation_check(c,pid,bid,qty,d)
                c.execute("INSERT INTO daily_work(project_id,date,boq_id,quantity,unit,station_from,station_to,notes,user_id) VALUES(?,?,?,?,?,?,?,?,?)",(pid,d,bid,qty,selected_unit,arr(station_froms,i),arr(station_tos,i),arr(remarks,i),u["id"]))
                c.execute("INSERT INTO daily_activities(project_id,date,boq_id,work_type,executed_qty,unit,machine_id,machine_hours,manpower_position,manpower_qty,manpower_hours,material_id,material_qty,remarks,user_id) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",(pid,d,bid,arr(work_types,i),qty,selected_unit,None,0,"",0,0,None,0,arr(remarks,i),u["id"]))
                aid=c.execute("SELECT last_insert_rowid()").fetchone()[0]

                for lid,meta in json_alloc(machine_data,i):
                    rec=c.execute("SELECT work_hours,idle_hours,down_hours FROM machine_logs WHERE id=? AND project_id=?",(lid,pid)).fetchone()
                    if not rec: continue
                    hours=parse_float(meta.get("hours")) if isinstance(meta,dict) else 0
                    max_hours=parse_float(rec["work_hours"])+parse_float(rec["idle_hours"])+parse_float(rec["down_hours"])
                    prior=c.execute("SELECT COALESCE(SUM(hours),0) x FROM activity_machines WHERE machine_log_id=?",(lid,)).fetchone()["x"]
                    if hours<0 or prior+hours>max_hours+0.0001: raise ValueError(f"Machinery allocation exceeds available hours for machine log {lid}. Available remaining: {max_hours-prior:.2f} h.")
                    c.execute("INSERT INTO activity_machines(activity_id,machine_log_id,machine_id,hours) SELECT ?,id,machine_id,? FROM machine_logs WHERE id=? AND project_id=?",(aid,hours,lid,pid))
                for mid,meta in json_alloc(manpower_data,i):
                    q=parse_float(meta.get("qty")) if isinstance(meta,dict) else 0; hrs=parse_float(meta.get("hours")) if isinstance(meta,dict) else 0
                    rec=c.execute("SELECT present,working_hours FROM manpower WHERE id=? AND project_id=?",(mid,pid)).fetchone()
                    if not rec: continue
                    if q<0 or hrs<0 or q>parse_float(rec["present"])+0.0001 or hrs>parse_float(rec["working_hours"])+0.0001: raise ValueError(f"Manpower allocation exceeds the selected attendance record for ID {mid}.")
                    c.execute("INSERT INTO activity_manpower(activity_id,manpower_id,crew_id,qty,hours) SELECT ?,id,crew_id,?,? FROM manpower WHERE id=? AND project_id=?",(aid,q,hrs,mid,pid))
                for cid,meta in json_alloc(crew_data,i):
                    c.execute("INSERT INTO crew_evaluations(activity_id,crew_id,evaluation,remarks,score) SELECT ?,id,?,?,? FROM project_crews WHERE id=? AND project_id=?",(aid,arr(evaluations,i),arr(eval_remarks,i),parse_float(arr(scores,i)),cid,pid))
                for sid,meta in json_alloc(store_data,i):
                    q=parse_float(meta.get("qty")) if isinstance(meta,dict) else 0
                    rec=c.execute("SELECT issued FROM store_logs WHERE id=? AND project_id=?",(sid,pid)).fetchone()
                    if not rec: continue
                    prior=c.execute("SELECT COALESCE(SUM(qty),0) x FROM activity_store WHERE store_log_id=?",(sid,)).fetchone()["x"]
                    if q<0 or prior+q>parse_float(rec["issued"])+0.0001: raise ValueError(f"Store allocation exceeds the issued quantity for store record {sid}. Remaining: {parse_float(rec['issued'])-prior:.3f}.")
                    c.execute("INSERT INTO activity_store(activity_id,store_log_id,material_id,qty) SELECT ?,id,material_id,? FROM store_logs WHERE id=? AND project_id=?",(aid,q,sid,pid))
                for fid,meta in json_alloc(fuel_data,i):
                    q=parse_float(meta.get("litres")) if isinstance(meta,dict) else 0
                    rec=c.execute("SELECT opening_gauge+fuel_received-closing_gauge actual FROM fuel_logs WHERE id=? AND project_id=?",(fid,pid)).fetchone()
                    if not rec: continue
                    prior=c.execute("SELECT COALESCE(SUM(litres),0) x FROM activity_fuel WHERE fuel_log_id=?",(fid,)).fetchone()["x"]
                    if q<0 or prior+q>parse_float(rec["actual"])+0.0001: raise ValueError(f"Fuel allocation exceeds the actual fuel in record {fid}. Remaining: {parse_float(rec['actual'])-prior:.2f} L.")
                    c.execute("INSERT INTO activity_fuel(activity_id,fuel_log_id,litres) SELECT ?,id,? FROM fuel_logs WHERE id=? AND project_id=?",(aid,q,fid,pid))
                for xid,meta in json_alloc(finance_data,i):
                    q=parse_float(meta.get("amount")) if isinstance(meta,dict) else 0
                    rec=c.execute("SELECT amount FROM finance_logs WHERE id=? AND project_id=?",(xid,pid)).fetchone()
                    if not rec: continue
                    prior=c.execute("SELECT COALESCE(SUM(amount),0) x FROM activity_finance WHERE finance_log_id=?",(xid,)).fetchone()["x"]
                    if q<0 or prior+q>parse_float(rec["amount"])+0.0001: raise ValueError(f"Finance allocation exceeds the amount in record {xid}. Remaining: {parse_float(rec['amount'])-prior:.2f}.")
                    c.execute("INSERT INTO activity_finance(activity_id,finance_log_id,amount) SELECT ?,id,? FROM finance_logs WHERE id=? AND project_id=?",(aid,q,xid,pid))
                if msg: flash("🚨 "+msg,"error")
                saved+=1

            if saved==0:
                raise ValueError("No valid BOQ work package was found. Enter a BOQ item and executed quantity.")

            # Commit the work/resources first. Do NOT open a second SQLite connection while
            # this write transaction is still open; that was a major cause of database-lock errors.
            c.commit()
            # Release this request's SQLite connection before the snapshot service writes.
            c.close(); c=None
            # Then create/update the permanent Daily snapshot using the normal report service.
            try:
                save_report(pid,"DAILY",dt.date.fromisoformat(d),dt.date.fromisoformat(d),"ALL",u["id"])
            except Exception as snapshot_error:
                app.logger.exception("Daily snapshot failed after successful work save: %s", snapshot_error)
                flash("⚠️ Daily work was saved, but the archive snapshot could not be refreshed: "+str(snapshot_error),"error")
            flash(f"✅ Daily Report saved once: {saved} BOQ work package(s) with all selected machinery, manpower, crews, store, fuel and finance linked.","success")
            c=db()
        if c is None: c=db()
        boq=c.execute("SELECT * FROM boq WHERE project_id=? ORDER BY series,item_no,id",(pid,)).fetchall()
        units=sorted(set(UNIT_CATALOG) | {str(r["unit"]).strip() for r in boq if r["unit"]})
        linked_machines=c.execute("SELECT ml.*,m.machine_type,m.code,m.plate_no,m.ownership,(ml.work_hours+ml.idle_hours+ml.down_hours-COALESCE((SELECT SUM(am.hours) FROM activity_machines am WHERE am.machine_log_id=ml.id),0)) remaining_hours FROM machine_logs ml JOIN machines m ON m.id=ml.machine_id WHERE ml.project_id=? AND ml.date=? ORDER BY m.machine_type,m.code",(pid,default_date)).fetchall()
        linked_manpower=c.execute("SELECT mp.*,pc.group_name,pc.position crew_position,pc.name crew_name,MAX(0,mp.present-COALESCE((SELECT SUM(ap.qty) FROM activity_manpower ap WHERE ap.manpower_id=mp.id),0)) remaining_qty,MAX(0,mp.working_hours-COALESCE((SELECT SUM(ap.hours) FROM activity_manpower ap WHERE ap.manpower_id=mp.id),0)) remaining_hours FROM manpower mp LEFT JOIN project_crews pc ON pc.id=mp.crew_id WHERE mp.project_id=? AND mp.date=? ORDER BY pc.group_name,mp.position,mp.name",(pid,default_date)).fetchall()
        linked_fuel=c.execute("SELECT f.*,m.machine_type,m.code,m.plate_no,MAX(0,(f.opening_gauge+f.fuel_received-f.closing_gauge)-COALESCE((SELECT SUM(af.litres) FROM activity_fuel af WHERE af.fuel_log_id=f.id),0)) remaining_litres FROM fuel_logs f JOIN machines m ON m.id=f.machine_id WHERE f.project_id=? AND f.date=? ORDER BY m.machine_type,m.code,f.id",(pid,default_date)).fetchall()
        linked_store=c.execute("SELECT sl.*,m.name,m.unit,MAX(0,sl.issued-COALESCE((SELECT SUM(ast.qty) FROM activity_store ast WHERE ast.store_log_id=sl.id),0)) remaining_qty FROM store_logs sl JOIN materials m ON m.id=sl.material_id WHERE sl.project_id=? AND sl.date=? ORDER BY m.name",(pid,default_date)).fetchall()
        linked_finance=c.execute("SELECT fl.*,MAX(0,fl.amount-COALESCE((SELECT SUM(af.amount) FROM activity_finance af WHERE af.finance_log_id=fl.id),0)) remaining_amount FROM finance_logs fl WHERE fl.project_id=? AND fl.date=? ORDER BY fl.id",(pid,default_date)).fetchall()
        crews=c.execute("SELECT * FROM project_crews WHERE project_id=? ORDER BY group_name,position,name",(pid,)).fetchall()
        alerts=c.execute("SELECT va.*,b.item_no,b.description FROM variation_alerts va JOIN boq b ON b.id=va.boq_id WHERE va.project_id=? ORDER BY va.id DESC LIMIT 30",(pid,)).fetchall()
        recent=c.execute("SELECT dw.*,b.item_no,b.description,b.unit,b.rate,b.series,(dw.quantity*b.rate) amount FROM daily_work dw JOIN boq b ON b.id=dw.boq_id WHERE dw.project_id=? ORDER BY dw.date DESC,dw.id DESC LIMIT 50",(pid,)).fetchall()
        return render_template("daily.html",pid=pid,date=default_date,boq=boq,units=units,linked_machines=linked_machines,linked_manpower=linked_manpower,linked_fuel=linked_fuel,linked_store=linked_store,linked_finance=linked_finance,crews=crews,alerts=alerts,recent=recent)
    except Exception as e:
        try: c.rollback()
        except Exception: pass
        flash("Daily Report save failed: "+str(e),"error")
        return redirect(url_for("daily",pid=pid,date=default_date))
    finally:
        try: c.close()
        except Exception: pass

@app.route("/projects/<int:pid>/fuel",methods=["GET","POST"])
@login_required
def fuel(pid):
    if not allowed_project(pid) or (not can_module("Machinery") and not project_admin(pid)): flash("🚫 Machinery/Fuel access is not assigned.","error"); return redirect(url_for("project",pid=pid))
    c=db()
    try:
        if request.method=="POST":
            mid=int(request.form["machine_id"]); d=request.form["date"]
            if not c.execute("SELECT 1 FROM machines WHERE id=? AND project_id=? AND active=1",(mid,pid)).fetchone(): raise ValueError('Selected machine is not active in this project.')
            vals=(pid,mid,d,parse_float(request.form.get("opening_gauge")),parse_float(request.form.get("fuel_received")),parse_float(request.form.get("closing_gauge")),parse_float(request.form.get("fuel_price")),request.form.get("reference",""),request.form.get("notes",""),current_user()["id"],"Fuel Register")
            c.execute("INSERT INTO fuel_logs(project_id,machine_id,date,opening_gauge,fuel_received,closing_gauge,fuel_price,reference,notes,user_id,source) VALUES(?,?,?,?,?,?,?,?,?,?,?)",vals)
            c.commit(); flash("⛽ Fuel log saved.","success")
    except Exception as e:
        c.rollback(); flash("Fuel log failed: "+str(e),"error")
    machines=c.execute("SELECT * FROM machines WHERE project_id=? AND active=1 ORDER BY machine_type,code",(pid,)).fetchall()
    logs=c.execute("SELECT f.*,m.machine_type,m.code,m.plate_no,m.engine_no,m.ownership,(f.opening_gauge+f.fuel_received-f.closing_gauge) consumption,(f.fuel_received*f.fuel_price) cost,COALESCE((SELECT SUM(ml.work_hours) FROM machine_logs ml WHERE ml.machine_id=f.machine_id AND ml.date=f.date),0) work_hours,COALESCE((SELECT SUM(ml.work_hours) FROM machine_logs ml WHERE ml.machine_id=f.machine_id AND ml.date=f.date),0)*m.expected_fuel expected_consumption FROM fuel_logs f JOIN machines m ON m.id=f.machine_id WHERE f.project_id=? ORDER BY f.date DESC,f.id DESC LIMIT 100",(pid,)).fetchall()
    total_l=c.execute("SELECT COALESCE(SUM(fuel_received),0) FROM fuel_logs WHERE project_id=?",(pid,)).fetchone()[0]; total_cost=c.execute("SELECT COALESCE(SUM(fuel_received*fuel_price),0) FROM fuel_logs WHERE project_id=?",(pid,)).fetchone()[0]
    c.close(); return render_template("fuel.html",pid=pid,machines=machines,logs=logs,total_l=total_l,total_cost=total_cost,today=dt.date.today().isoformat())

@app.route("/projects/<int:pid>/performance",methods=["GET","POST"])
@login_required
def performance(pid):
    if not allowed_project(pid): return redirect(url_for("dashboard"))
    c=db()
    if request.method=="POST":
        c.execute("INSERT INTO performance_rates(project_id,work_type,worker_type,unit,qty_per_hour,notes) VALUES(?,?,?,?,?,?) ON CONFLICT(project_id,work_type,worker_type) DO UPDATE SET unit=excluded.unit,qty_per_hour=excluded.qty_per_hour,notes=excluded.notes",(pid,request.form["work_type"],request.form["worker_type"],request.form["unit"],parse_float(request.form["qty_per_hour"]),request.form.get("notes",""))); c.commit(); flash("📈 Performance rate saved.","success")
    rows=c.execute("SELECT * FROM performance_rates WHERE project_id=? ORDER BY work_type,worker_type",(pid,)).fetchall(); c.close(); return render_template("performance.html",pid=pid,rows=rows)

@app.route("/projects/<int:pid>/print-report")
@login_required
def print_report(pid):
    if not allowed_project(pid): return redirect(url_for("dashboard"))
    date=request.args.get("date",dt.date.today().isoformat()); c=db()
    p=c.execute("SELECT * FROM projects WHERE id=?",(pid,)).fetchone(); settings=c.execute("SELECT * FROM report_settings WHERE project_id=?",(pid,)).fetchone()
    boq=c.execute("SELECT b.item_no,b.description,b.unit,b.rate,COALESCE(SUM(dw.quantity),0) qty FROM boq b LEFT JOIN daily_work dw ON dw.boq_id=b.id AND dw.date=? WHERE b.project_id=? GROUP BY b.id ORDER BY b.item_no",(date,pid)).fetchall()
    activities=c.execute("SELECT a.*,b.item_no,b.description FROM daily_activities a LEFT JOIN boq b ON b.id=a.boq_id WHERE a.project_id=? AND a.date=? ORDER BY a.id",(pid,date)).fetchall()
    activity_links={}
    for a in activities:
        aid=a['id']
        activity_links[aid]={
            'machines':c.execute("SELECT am.*,m.machine_type,m.code,m.plate_no FROM activity_machines am JOIN machines m ON m.id=am.machine_id WHERE am.activity_id=?",(aid,)).fetchall(),
            'manpower':c.execute("SELECT ap.*,mp.name,mp.position,pc.group_name,pc.name crew_name FROM activity_manpower ap JOIN manpower mp ON mp.id=ap.manpower_id LEFT JOIN project_crews pc ON pc.id=ap.crew_id WHERE ap.activity_id=?",(aid,)).fetchall(),
            'store':c.execute("SELECT ast.*,m.name,m.unit FROM activity_store ast JOIN materials m ON m.id=ast.material_id WHERE ast.activity_id=?",(aid,)).fetchall(),
            'fuel':c.execute("SELECT af.*,f.opening_gauge,f.fuel_received,f.closing_gauge,m.machine_type,m.code FROM activity_fuel af JOIN fuel_logs f ON f.id=af.fuel_log_id JOIN machines m ON m.id=f.machine_id WHERE af.activity_id=?",(aid,)).fetchall(),
            'finance':c.execute("SELECT ax.*,f.kind,f.description FROM activity_finance ax JOIN finance_logs f ON f.id=ax.finance_log_id WHERE ax.activity_id=?",(aid,)).fetchall(),
            'evaluations':c.execute("SELECT ce.*,pc.name,pc.position,pc.group_name FROM crew_evaluations ce JOIN project_crews pc ON pc.id=ce.crew_id WHERE ce.activity_id=?",(aid,)).fetchall()
        }
    machines=c.execute("SELECT ml.*,m.machine_type,m.code,m.plate_no,m.ownership FROM machine_logs ml JOIN machines m ON m.id=ml.machine_id WHERE ml.project_id=? AND ml.date=?",(pid,date)).fetchall()
    manpower=c.execute("SELECT * FROM manpower WHERE project_id=? AND date=?",(pid,date)).fetchall(); store=c.execute("SELECT sl.*,m.name,m.unit FROM store_logs sl JOIN materials m ON m.id=sl.material_id WHERE sl.project_id=? AND sl.date=?",(pid,date)).fetchall(); problems=c.execute("SELECT * FROM problems WHERE project_id=? AND date=?",(pid,date)).fetchall(); c.close()
    return render_template("print_report.html",p=p,settings=settings,date=date,boq=boq,activities=activities,activity_links=activity_links,machines=machines,manpower=manpower,store=store,problems=problems)

@app.route("/projects/<int:pid>/machinery/assign", methods=["POST"])
@login_required
def assign_machine(pid):
    if not allowed_project(pid) or (not can_module("Machinery") and not project_admin(pid)):
        flash("🚫 Machinery access is not assigned.", "error"); return redirect(url_for("project", pid=pid))
    c=db()
    try:
        mid=int(request.form.get("machine_id")); m=c.execute("SELECT * FROM machines WHERE id=? AND project_id=? AND active=1",(mid,pid)).fetchone()
        if not m: raise ValueError("Machine not found in this project fleet.")
        active=c.execute("SELECT id FROM machine_assignments WHERE machine_id=? AND project_id=? AND status='ACTIVE'",(mid,pid)).fetchone()
        if active: raise ValueError("This machine already has an active signed registration.")
        raw=request.form.get("total_hours","").strip()
        if not raw: raise ValueError("Enter the Total Signed Hours before registering the machine.")
        total=parse_float(raw)
        if total<=0: raise ValueError("Total Signed Hours must be greater than zero.")
        c.execute("INSERT INTO machine_assignments(machine_id,project_id,total_signed_hours,hours_used,status,assigned_by,notes) VALUES(?,?,?,0,'ACTIVE',?,?)",(mid,pid,total,current_user()["id"],request.form.get("notes","")))
        c.execute("UPDATE machines SET total_signed_hours=?,hours_used=0,lifecycle_status='ACTIVE',assignment_signed_by=?,assignment_ended_by=NULL,assignment_ended_at=NULL,assignment_end_date=NULL,assignment_end_hour=NULL WHERE id=? AND project_id=?",(total,current_user()["id"],mid,pid))
        c.commit(); flash(f"✍️ Machine registered for {total:g} signed hours. The system will alert and stop the assignment when the limit is reached.","success")
    except Exception as e:
        c.rollback(); flash("Could not sign machine assignment: "+str(e),"error")
    finally: c.close()
    return redirect(url_for("machinery",pid=pid))

@app.route("/projects/<int:pid>/machinery/end-assignment", methods=["POST"])
@login_required
def end_machine_assignment(pid):
    if not allowed_project(pid) or (not can_module("Machinery") and not project_admin(pid)):
        flash("🚫 Machinery access is not assigned.", "error"); return redirect(url_for("project", pid=pid))
    c=db()
    try:
        mid=int(request.form.get("machine_id")); a=c.execute("SELECT * FROM machine_assignments WHERE machine_id=? AND project_id=? AND status='ACTIVE' ORDER BY id DESC LIMIT 1",(mid,pid)).fetchone()
        if not a: raise ValueError("No active assignment exists for this machine.")
        end_date=request.form.get("end_date") or dt.date.today().isoformat(); end_hour=parse_float(request.form.get("end_hour"))
        c.execute("UPDATE machine_assignments SET status='ENDED',end_date=?,end_hour=?,ended_by=?,ended_at=CURRENT_TIMESTAMP WHERE id=?",(end_date,end_hour,current_user()["id"],a["id"]))
        c.execute("UPDATE machines SET lifecycle_status='ENDED',assignment_end_date=?,assignment_end_hour=?,assignment_ended_by=?,assignment_ended_at=CURRENT_TIMESTAMP WHERE id=? AND project_id=?",(end_date,end_hour,current_user()["id"],mid,pid))
        c.commit(); flash("🛑 Assignment ended. A new signed assignment is required before logging this machine again.","success")
    except Exception as e:
        c.rollback(); flash("Could not end machine assignment: "+str(e),"error")
    finally: c.close()
    return redirect(url_for("machinery",pid=pid))

@app.route("/projects/<int:pid>/machinery",methods=["GET","POST"])
@login_required
def machinery(pid):
    if not allowed_project(pid) or (not can_module("Machinery") and not project_admin(pid)):flash("🚫 Machinery access is not assigned.","error");return redirect(url_for("project",pid=pid))
    c=db()
    if request.method=="POST":
        action=request.form.get("action")
        if action=="add":
            machine_type=request.form["machine_type"]
            allowed_rate_unit=request.form.get("rate_unit","hr") if machine_type in ["Dump Truck","Fuel Truck","Water Truck","Shower Truck"] else "hr"
            c.execute("INSERT INTO machines(project_id,machine_type,code,plate_no,engine_no,ownership,hourly_rate,rate_unit,expected_fuel,fuel_price,lifecycle_status) VALUES(?,?,?,?,?,?,?,?,?,?,?)",(pid,machine_type,request.form["code"],request.form.get("plate_no",request.form["code"]),request.form.get("engine_no",""),request.form["ownership"],parse_float(request.form["hourly_rate"]),allowed_rate_unit,parse_float(request.form["expected_fuel"]),parse_float(request.form.get("fuel_price")),"UNASSIGNED"))
            flash("🚜 Machine added to this project's fleet.","success")
        elif action=="remove":c.execute("UPDATE machines SET active=0 WHERE id=? AND project_id=?",(request.form["machine_id"],pid));flash("Machine removed from active fleet.","success")
        elif action=="log":
            active_assignment=c.execute("SELECT id FROM machine_assignments WHERE machine_id=? AND project_id=? AND status='ACTIVE' ORDER BY id DESC LIMIT 1",(request.form['machine_id'],pid)).fetchone()
            if not active_assignment: raise ValueError('Machine has no active signed assignment. Machinery Admin must sign a new start date/hour first.')
            vals=(pid,request.form["machine_id"],request.form["date"],parse_float(request.form.get("work_hours")),parse_float(request.form.get("idle_hours")),request.form.get("idle_reason",""),1 if request.form.get("idle_payable")=="1" else 0,parse_float(request.form.get("down_hours")),request.form.get("down_reason",""),parse_float(request.form.get("opening_gauge")),parse_float(request.form.get("fuel_received")),parse_float(request.form.get("closing_gauge")),request.form.get("notes",""),current_user()["id"])
            c.execute("INSERT INTO machine_logs(project_id,machine_id,date,work_hours,idle_hours,idle_reason,idle_payable,down_hours,down_reason,opening_gauge,fuel_received,closing_gauge,notes,user_id) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",vals)
            if any(parse_float(request.form.get(k)) for k in ("opening_gauge","fuel_received","closing_gauge")):
                mp=c.execute("SELECT fuel_price FROM machines WHERE id=?",(request.form['machine_id'],)).fetchone()
                c.execute("INSERT INTO fuel_logs(project_id,machine_id,date,opening_gauge,fuel_received,closing_gauge,fuel_price,reference,notes,user_id,source) VALUES(?,?,?,?,?,?,?,?,?,?,?)",(pid,request.form['machine_id'],request.form['date'],parse_float(request.form.get('opening_gauge')),parse_float(request.form.get('fuel_received')),parse_float(request.form.get('closing_gauge')),mp['fuel_price'] if mp else 0,'MCH-'+request.form['date'],request.form.get('notes',''),current_user()['id'],'Machinery Log'))
            used=parse_float(request.form.get("work_hours"))+parse_float(request.form.get("idle_hours"))+parse_float(request.form.get("down_hours"))
            ma=c.execute("SELECT id,total_signed_hours,hours_used FROM machine_assignments WHERE machine_id=? AND project_id=? AND status='ACTIVE' ORDER BY id DESC LIMIT 1",(request.form['machine_id'],pid)).fetchone()
            if ma:
                limit=ma['total_signed_hours'] or 0; new_used=(ma['hours_used'] or 0)+used; reached=bool(limit>0 and new_used >= limit)
                c.execute("UPDATE machine_assignments SET hours_used=?,status=CASE WHEN ? THEN 'ENDED' ELSE status END,end_date=CASE WHEN ? THEN ? ELSE end_date END,end_hour=CASE WHEN ? THEN NULL ELSE end_hour END WHERE id=?",(new_used,reached,reached,request.form['date'],reached,ma['id']))
                c.execute("UPDATE machines SET hours_used=?,lifecycle_status=CASE WHEN ? THEN 'ENDED' ELSE lifecycle_status END,assignment_end_date=CASE WHEN ? THEN ? ELSE assignment_end_date END,assignment_end_hour=NULL WHERE id=?",(new_used,reached,reached,request.form['date'],request.form['machine_id']))
                if reached:
                    flash(f"🚨 SIGNED HOURS REACHED: Machine has used {new_used:g} of {limit:g} signed hours. A new signature is required before further use.","error")
                else:
                    flash(f"⏱️ Machine log saved. Signed hours used: {new_used:g} / {limit:g} h.","success")
            else:
                flash("⏱️ Machine hours / idle / down / gauge saved successfully.","success")
        try: c.commit()
        except Exception as e: c.rollback(); flash("Machinery save failed: "+str(e),"error")
    machines=c.execute("SELECT * FROM machines WHERE project_id=? AND active=1 ORDER BY machine_type,code",(pid,)).fetchall();assignments=c.execute("SELECT ma.*,m.machine_type,m.code,m.plate_no FROM machine_assignments ma JOIN machines m ON m.id=ma.machine_id WHERE ma.project_id=? ORDER BY ma.id DESC LIMIT 100",(pid,)).fetchall();logs=c.execute("SELECT ml.*,m.machine_type,m.code,m.ownership,m.hourly_rate,m.rate_unit,m.expected_fuel,((CASE WHEN m.rate_unit='day' THEN CASE WHEN (ml.work_hours+ml.idle_hours+ml.down_hours)>0 THEN m.hourly_rate ELSE 0 END WHEN m.rate_unit='month' THEN CASE WHEN (ml.work_hours+ml.idle_hours+ml.down_hours)>0 THEN m.hourly_rate/30.0 ELSE 0 END ELSE (ml.work_hours+CASE WHEN ml.idle_payable=1 THEN ml.idle_hours ELSE 0 END)*m.hourly_rate END)) expense,(ml.opening_gauge+ml.fuel_received-ml.closing_gauge) actual_fuel,(ml.work_hours*m.expected_fuel) expected_fuel_qty,CASE WHEN (ml.work_hours+ml.idle_hours+ml.down_hours)>0 THEN ml.work_hours*100.0/(ml.work_hours+ml.idle_hours+ml.down_hours) ELSE 0 END utilization,CASE WHEN (ml.work_hours+ml.idle_hours+ml.down_hours)>0 THEN (ml.work_hours+ml.idle_hours)*100.0/(ml.work_hours+ml.idle_hours+ml.down_hours) ELSE 0 END availability,(ml.opening_gauge+ml.fuel_received-ml.closing_gauge)-(ml.work_hours*m.expected_fuel) fuel_discrepancy FROM machine_logs ml JOIN machines m ON m.id=ml.machine_id WHERE ml.project_id=? ORDER BY ml.date DESC,ml.id DESC LIMIT 50",(pid,)).fetchall();c.close()
    return render_template("machinery.html",pid=pid,machines=machines,assignments=assignments,logs=logs)

@app.route("/projects/<int:pid>/manpower",methods=["GET","POST"])
@login_required
def manpower(pid):
    if not allowed_project(pid) or (not can_module("HR") and not project_admin(pid)): flash("🚫 HR/manpower access is not assigned.","error"); return redirect(url_for("project",pid=pid))
    c=db()
    try:
        if request.method=="POST":
            vals=(pid,request.form["date"],request.form["name"],request.form["employment"],request.form["position"],request.form.get("crew_id") or None,parse_float(request.form.get("present")),parse_float(request.form.get("working_hours",8)),parse_float(request.form.get("hourly_rate")),parse_float(request.form.get("daily_rate")),parse_float(request.form.get("normal_ot_hours")),parse_float(request.form.get("normal_ot_rate")),parse_float(request.form.get("night_ot_hours")),parse_float(request.form.get("night_ot_rate")),parse_float(request.form.get("sunday_ot_hours")),parse_float(request.form.get("sunday_ot_rate")),parse_float(request.form.get("holiday_ot_hours")),parse_float(request.form.get("holiday_ot_rate")),0,0,request.form.get("notes",""),current_user()["id"])
            c.execute("INSERT INTO manpower(project_id,date,name,employment,position,crew_id,present,working_hours,hourly_rate,daily_rate,normal_ot_hours,normal_ot_rate,night_ot_hours,night_ot_rate,sunday_ot_hours,sunday_ot_rate,holiday_ot_hours,holiday_ot_rate,overtime_hours,overtime_rate,notes,user_id) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",vals); c.commit(); flash("👷 Manpower record saved and linked to the selected crew.","success")
    except Exception as e: c.rollback(); flash("Manpower save failed: "+str(e),"error")
    rows=c.execute("SELECT mp.*,pc.group_name,pc.position crew_position,pc.name crew_name FROM manpower mp LEFT JOIN project_crews pc ON pc.id=mp.crew_id WHERE mp.project_id=? ORDER BY mp.date DESC,mp.id DESC LIMIT 100",(pid,)).fetchall(); crews=c.execute("SELECT * FROM project_crews WHERE project_id=? ORDER BY group_name,position,name",(pid,)).fetchall(); c.close(); return render_template("manpower.html",pid=pid,rows=rows,crews=crews)

@app.route("/projects/<int:pid>/store",methods=["GET","POST"])
@login_required
def store(pid):
    if not allowed_project(pid) or (not can_module("Store") and not project_admin(pid)):flash("🚫 Store access is not assigned.","error");return redirect(url_for("project",pid=pid))
    c=db()
    if request.method=="POST":
        if request.form.get("action")=="add":c.execute("INSERT INTO materials(project_id,category,name,unit,min_stock) VALUES(?,?,?,?,?)",(pid,request.form["category"],request.form["name"],request.form["unit"],parse_float(request.form["min_stock"])))
        else:
            physical=request.form.get("physical_balance");physical=parse_float(physical) if physical not in (None,"") else None
            c.execute("INSERT INTO store_logs(project_id,material_id,date,received,issued,unit_cost,physical_balance,reference,notes,user_id) VALUES(?,?,?,?,?,?,?,?,?,?)",(pid,request.form["material_id"],request.form["date"],parse_float(request.form["received"]),parse_float(request.form["issued"]),parse_float(request.form["unit_cost"]),physical,request.form.get("reference",""),request.form.get("notes",""),current_user()["id"]))
        c.commit();flash("📦 Store record saved.","success")
    mats=c.execute("SELECT m.*,COALESCE(SUM(sl.received),0) rec,COALESCE(SUM(sl.issued),0) iss,COALESCE(SUM(sl.received)-SUM(sl.issued),0) balance,(SELECT sl2.physical_balance FROM store_logs sl2 WHERE sl2.material_id=m.id AND sl2.physical_balance IS NOT NULL ORDER BY sl2.date DESC,sl2.id DESC LIMIT 1) physical FROM materials m LEFT JOIN store_logs sl ON sl.material_id=m.id WHERE m.project_id=? AND m.active=1 GROUP BY m.id ORDER BY m.category,m.name",(pid,)).fetchall();logs=c.execute("SELECT sl.*,m.name,m.unit FROM store_logs sl JOIN materials m ON m.id=sl.material_id WHERE sl.project_id=? ORDER BY sl.date DESC,sl.id DESC LIMIT 100",(pid,)).fetchall();c.close();return render_template("store.html",pid=pid,materials=mats,logs=logs)


@app.route("/projects/<int:pid>/crew",methods=["GET","POST"])
@login_required
def crew(pid):
    if not allowed_project(pid) or (not can_module("HR") and not project_admin(pid)): flash("🚫 Crew access is not assigned.","error"); return redirect(url_for("project",pid=pid))
    c=db()
    try:
        if request.method=="POST":
            action=request.form.get("action","add")
            if action=="capacity":
                name=request.form.get("capacity_group","").strip()
                if not name: raise ValueError("Select or enter a crew group.")
                c.execute("INSERT OR IGNORE INTO crew_groups(name) VALUES(?)",(name,))
                c.execute("INSERT INTO crew_group_capacity(group_name,foreman_qty,dl_qty,surveyor_qty,data_collector_qty,time_keeper_qty,other_qty,total_qty) VALUES(?,?,?,?,?,?,?,?) ON CONFLICT(group_name) DO UPDATE SET foreman_qty=excluded.foreman_qty,dl_qty=excluded.dl_qty,surveyor_qty=excluded.surveyor_qty,data_collector_qty=excluded.data_collector_qty,time_keeper_qty=excluded.time_keeper_qty,other_qty=excluded.other_qty,total_qty=excluded.total_qty",(name,parse_float(request.form.get('foreman_qty')),parse_float(request.form.get('dl_qty')),parse_float(request.form.get('surveyor_qty')),parse_float(request.form.get('data_collector_qty')),parse_float(request.form.get('time_keeper_qty')),parse_float(request.form.get('other_qty')),parse_float(request.form.get('total_qty'))))
                c.commit(); flash("👥 Crew capacity saved for future daily crew planning.","success")
            elif action=="add_group":
                name=request.form.get("new_group","").strip()
                if name: c.execute("INSERT OR IGNORE INTO crew_groups(name) VALUES(?)",(name,)); c.execute("INSERT OR IGNORE INTO crew_group_capacity(group_name) VALUES(?)",(name,)); c.commit(); flash("👥 Group added for future selection.","success")
            elif action=="add_position":
                name=request.form.get("new_position","").strip()
                if name: c.execute("INSERT OR IGNORE INTO crew_positions(name) VALUES(?)",(name,)); c.commit(); flash("👷 Position added for future selection.","success")
            else:
                group=request.form.get("group_name","").strip(); position=request.form.get("position","").strip()
                if group: c.execute("INSERT OR IGNORE INTO crew_groups(name) VALUES(?)",(group,)); c.execute("INSERT OR IGNORE INTO crew_group_capacity(group_name) VALUES(?)",(group,))
                if position: c.execute("INSERT OR IGNORE INTO crew_positions(name) VALUES(?)",(position,))
                c.execute("INSERT INTO project_crews(project_id,date,group_name,position,name,employment,skill_level,working_hours,hourly_rate,notes,user_id) VALUES(?,?,?,?,?,?,?,?,?,?,?)",(pid,request.form.get("date",dt.date.today().isoformat()),group,position,request.form["name"],request.form["employment"],request.form.get("skill_level","Skilled"),parse_float(request.form.get("working_hours")),parse_float(request.form.get("hourly_rate")),request.form.get("notes",""),current_user()["id"])); c.commit(); flash("👷 Crew member registered.","success")
    except Exception as e: c.rollback(); flash("Crew save failed: "+str(e),"error")
    groups=[r["name"] for r in c.execute("SELECT name FROM crew_groups WHERE active=1 ORDER BY name").fetchall()]; positions=[r["name"] for r in c.execute("SELECT name FROM crew_positions WHERE active=1 ORDER BY name").fetchall()]; rows=c.execute("SELECT * FROM project_crews WHERE project_id=? ORDER BY group_name,position,name",(pid,)).fetchall(); capacities=c.execute("SELECT * FROM crew_group_capacity WHERE active=1 ORDER BY group_name").fetchall(); c.close()
    return render_template("crew.html",pid=pid,rows=rows,crew_groups=groups,position_catalog=positions,capacities=capacities)

@app.route("/projects/<int:pid>/report-settings",methods=["GET","POST"])
@login_required
def report_settings(pid):
    if not project_admin(pid): return redirect(url_for("project",pid=pid))
    c=db(); s=c.execute("SELECT * FROM report_settings WHERE project_id=?",(pid,)).fetchone()
    if request.method=="POST":
        vals=(pid,request.form.get("contractor_role","Main Contractor"),request.form.get("phone",""),request.form.get("email",""),request.form.get("website",""),request.form.get("fax",""),request.form.get("address",""),request.form.get("logo_text","BAGC"))
        c.execute("INSERT INTO report_settings(project_id,contractor_role,phone,email,website,fax,address,logo_text) VALUES(?,?,?,?,?,?,?,?) ON CONFLICT(project_id) DO UPDATE SET contractor_role=excluded.contractor_role,phone=excluded.phone,email=excluded.email,website=excluded.website,fax=excluded.fax,address=excluded.address,logo_text=excluded.logo_text",vals); c.commit(); flash("🖨️ Report header/footer saved.","success"); s=c.execute("SELECT * FROM report_settings WHERE project_id=?",(pid,)).fetchone()
    c.close(); return render_template("report_settings.html",pid=pid,s=s)

@app.route("/projects/<int:pid>/design",methods=["GET","POST"])
@login_required
def design(pid):
    if not allowed_project(pid) or (not can_module("Design") and not project_admin(pid)):flash("🚫 Design access is not assigned.","error");return redirect(url_for("project",pid=pid))
    c=db()
    if request.method=="POST":
        c.execute("INSERT INTO design_items(project_id,drawing_no,title,discipline,revision,status,submitted,approved,comments,user_id) VALUES(?,?,?,?,?,?,?,?,?,?)",(pid,request.form["drawing_no"],request.form["title"],request.form["discipline"],request.form["revision"],request.form["status"],request.form.get("submitted",""),request.form.get("approved",""),request.form.get("comments",""),current_user()["id"]));c.commit();flash("🎨 Design record saved.","success")
    rows=c.execute("SELECT * FROM design_items WHERE project_id=? ORDER BY id DESC",(pid,)).fetchall();c.close();return render_template("design.html",pid=pid,rows=rows)

@app.route("/projects/<int:pid>/finance",methods=["GET","POST"])
@login_required
def finance(pid):
    if not allowed_project(pid) or (not can_module("Finance") and not project_admin(pid)):flash("🚫 Finance access is not assigned.","error");return redirect(url_for("project",pid=pid))
    c=db()
    if request.method=="POST":c.execute("INSERT INTO finance_logs(project_id,date,category,kind,description,amount,reference,user_id) VALUES(?,?,?,?,?,?,?,?)",(pid,request.form["date"],request.form["category"],request.form["kind"],request.form["description"],parse_float(request.form["amount"]),request.form.get("reference",""),current_user()["id"]));c.commit();flash("💰 Finance record saved.","success")
    rows=c.execute("SELECT * FROM finance_logs WHERE project_id=? ORDER BY date DESC,id DESC LIMIT 100",(pid,)).fetchall();c.close();return render_template("finance.html",pid=pid,rows=rows)

@app.route("/projects/<int:pid>/reports")
@login_required
def reports(pid):
    if not allowed_project(pid): return redirect(url_for("dashboard"))
    report_type=request.args.get('report_type','MONTHLY').upper()
    scope=request.args.get('scope','ALL').upper()
    allowed_scopes=allowed_report_scopes_for_user(pid)
    if scope not in allowed_scopes:
        scope='MACHINERY' if 'MACHINERY' in allowed_scopes else ('FUEL' if 'FUEL' in allowed_scopes else (next(iter(sorted(allowed_scopes))) if allowed_scopes else 'BOQ'))
        flash('📊 You can only view reports belonging to your assigned department/project responsibility.','error')
    start_s=request.args.get('start',''); end_s=request.args.get('end','')
    try: start,end=report_dates(report_type,start_s,end_s)
    except Exception: start,end=report_dates(report_type,'','')
    c=db(); p=c.execute("SELECT * FROM projects WHERE id=?",(pid,)).fetchone()
    saved_all=c.execute("SELECT sr.*,u.full_name generated_name FROM saved_reports sr LEFT JOIN users u ON u.id=sr.generated_by WHERE sr.project_id=? ORDER BY sr.generated_at DESC,sr.id DESC LIMIT 100",(pid,)).fetchall()
    saved=[r for r in saved_all if r['scope'] in allowed_scopes]
    c.close()
    snapshot=build_report_snapshot(pid,start,end,scope)
    if request.args.get('save')=='1':
        rid=save_report(pid,report_type,start,end,scope,current_user()['id']); flash(f'📚 {report_type} report saved as a permanent report record.','success'); return redirect(url_for('reports',pid=pid,report_type=report_type,scope=scope,start=start.isoformat(),end=end.isoformat()))
    return render_template('reports.html',pid=pid,p=p,report_type=report_type,scope=scope,start=start,end=end,snapshot=snapshot,saved=saved,allowed_report_scopes=allowed_scopes)

@app.route("/projects/<int:pid>/reports/save",methods=['POST'])
@login_required
def save_report_route(pid):
    if not allowed_project(pid): return redirect(url_for('dashboard'))
    rt=request.form.get('report_type','MONTHLY').upper(); scope=request.form.get('scope','ALL').upper()
    if scope not in allowed_report_scopes_for_user(pid):
        flash('🚫 You are not allowed to save this report section.','error'); return redirect(url_for('reports',pid=pid))
    start=dt.date.fromisoformat(request.form['start']); end=dt.date.fromisoformat(request.form['end'])
    rid=save_report(pid,rt,start,end,scope,current_user()['id']); flash(f'📚 Report saved permanently (Record #{rid}).','success')
    return redirect(url_for('reports',pid=pid,report_type=rt,scope=scope,start=start.isoformat(),end=end.isoformat()))

@app.route("/projects/<int:pid>/reports/<int:rid>")
@login_required
def saved_report(pid,rid):
    if not allowed_project(pid): return redirect(url_for('dashboard'))
    c=db(); r=c.execute("SELECT sr.*,p.name project_name,p.client,p.consultant,p.contractor_role,u.full_name generated_name FROM saved_reports sr JOIN projects p ON p.id=sr.project_id LEFT JOIN users u ON u.id=sr.generated_by WHERE sr.id=? AND sr.project_id=?",(rid,pid)).fetchone()
    if not r: c.close(); return ('Report not found',404)
    if r['scope'] not in allowed_report_scopes_for_user(pid): c.close(); return ('Report access denied',403)
    source_ids=json.loads(r['source_report_ids'] or '[]'); sources=[]
    if source_ids:
        marks=','.join('?'*len(source_ids)); sources=c.execute(f"SELECT id,report_no,report_type,scope,start_date,end_date FROM saved_reports WHERE id IN ({marks}) ORDER BY start_date",source_ids).fetchall()
    c.close(); return render_template('saved_report.html',r=r,snapshot=json.loads(r['snapshot_json']),sources=sources)

@app.route("/projects/<int:pid>/rfi", methods=["GET","POST"])
@login_required
def rfi(pid):
    if not allowed_project(pid): flash("🚫 You do not have access to this project.","error"); return redirect(url_for("dashboard"))
    c=db(); project=c.execute("SELECT * FROM projects WHERE id=?",(pid,)).fetchone(); boqs=c.execute("SELECT * FROM boq WHERE project_id=? ORDER BY item_no",(pid,)).fetchall()
    users=c.execute("SELECT id,full_name,username,department,position,role FROM users WHERE active=1 AND (role='SUPER_ADMIN' OR id IN (SELECT user_id FROM user_projects WHERE project_id=?)) ORDER BY full_name",(pid,)).fetchall()
    if request.method=='POST':
        f=request.form
        roles=[('Site Engineer',f.get('site_engineer')),('Office Engineer',f.get('office_engineer')),('Project Manager',f.get('project_manager'))]
        if any(not uid for _,uid in roles):
            flash('👷 Select Site Engineer → Office Engineer → Project Manager in sequence.','error'); c.close(); return render_template('rfi.html',p=project,boqs=boqs,users=users,rfis=[],selected_inspectors=[])
        count=c.execute("SELECT COUNT(*) n FROM rfis WHERE project_id=?",(pid,)).fetchone()['n']+1; rfi_no=f"RFI-{dt.date.today().year}-{count:03d}"
        c.execute("INSERT INTO rfis(project_id,rfi_no,date_requested,inspection_date,location,boq_id,work_description,drawing_no,drawing_revision,specification,work_stage,submitted_by,status,overall_comment,corrective_action) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",(pid,rfi_no,f.get('date_requested') or dt.date.today().isoformat(),f.get('inspection_date') or None,f.get('location',''),f.get('boq_id') or None,f.get('work_description',''),f.get('drawing_no',''),f.get('drawing_revision',''),f.get('specification',''),f.get('work_stage',''),session['user_id'],'PENDING SITE ENGINEER','', ''))
        rid=c.execute('SELECT last_insert_rowid()').fetchone()[0]
        for order,(stage,uid) in enumerate(roles,1):
            c.execute("INSERT INTO rfi_steps(rfi_id,step_order,stage,assigned_user_id) VALUES(?,?,?,?)",(rid,order,stage,int(uid)))
        # retain legacy participant table for compatibility/printing
        for stage,uid in roles: c.execute("INSERT OR IGNORE INTO rfi_inspections(rfi_id,inspector_user_id,inspector_role) VALUES(?,?,?)",(rid,int(uid),stage))
        c.commit(); c.close(); flash(f'📋 {rfi_no} submitted. Workflow: Site Engineer → Office Engineer → Project Manager.','success'); return redirect(url_for('rfi',pid=pid))
    rfis=c.execute("SELECT r.*,u.full_name AS submitter,(SELECT COUNT(*) FROM rfi_steps s WHERE s.rfi_id=r.id) step_count,(SELECT COUNT(*) FROM rfi_steps s WHERE s.rfi_id=r.id AND s.decision IN ('APPROVED','APPROVED WITH COMMENTS')) approved_count FROM rfis r LEFT JOIN users u ON u.id=r.submitted_by WHERE r.project_id=? ORDER BY r.id DESC",(pid,)).fetchall(); c.close()
    return render_template('rfi.html',p=project,boqs=boqs,users=users,rfis=rfis,selected_inspectors=[])

@app.route("/projects/<int:pid>/rfi/<int:rid>", methods=["GET","POST"])
@login_required
def rfi_detail(pid,rid):
    if not allowed_project(pid): return redirect(url_for('dashboard'))
    c=db(); r=c.execute("SELECT r.*,p.name project_name,p.client,p.consultant,u.full_name submitter FROM rfis r JOIN projects p ON p.id=r.project_id LEFT JOIN users u ON u.id=r.submitted_by WHERE r.id=? AND r.project_id=?",(rid,pid)).fetchone()
    if not r: c.close(); return ('RFI not found',404)
    if request.method=='POST':
        sid=int(request.form.get('step_id','0')); decision=request.form.get('decision','PENDING'); comment=request.form.get('comments',''); me=c.execute('SELECT * FROM users WHERE id=?',(session.get('user_id'),)).fetchone()
        step=c.execute('SELECT * FROM rfi_steps WHERE id=? AND rfi_id=?',(sid,rid)).fetchone()
        if not step or (me['role']!='SUPER_ADMIN' and step['assigned_user_id']!=me['id']): flash('🚫 This RFI step is not assigned to your account.','error')
        else:
            previous=c.execute("SELECT COUNT(*) n FROM rfi_steps WHERE rfi_id=? AND step_order<? AND decision NOT IN ('APPROVED','APPROVED WITH COMMENTS')",(rid,step['step_order'])).fetchone()['n']
            if previous and me['role']!='SUPER_ADMIN': flash('⏳ The previous approval stage must be completed first.','error')
            else:
                c.execute("UPDATE rfi_steps SET decision=?,comments=?,inspection_date=?,signed_at=CURRENT_TIMESTAMP WHERE id=?",(decision,comment,request.form.get('inspection_date') or dt.date.today().isoformat(),sid))
                c.execute("UPDATE rfi_inspections SET decision=?,comments=?,inspection_date=?,signed_at=CURRENT_TIMESTAMP WHERE rfi_id=? AND inspector_user_id=?",(decision,comment,request.form.get('inspection_date') or dt.date.today().isoformat(),rid,step['assigned_user_id']))
                if decision=='REJECTED': status='REJECTED'
                elif decision=='RESUBMISSION REQUIRED': status='RESUBMISSION REQUIRED'
                else:
                    remaining=c.execute("SELECT COUNT(*) n FROM rfi_steps WHERE rfi_id=? AND decision NOT IN ('APPROVED','APPROVED WITH COMMENTS')",(rid,)).fetchone()['n']
                    if remaining==0: status='APPROVED'
                    else:
                        nxt=c.execute("SELECT stage FROM rfi_steps WHERE rfi_id=? AND decision NOT IN ('APPROVED','APPROVED WITH COMMENTS') ORDER BY step_order LIMIT 1",(rid,)).fetchone(); status=f"PENDING {nxt['stage'].upper()}" if nxt else 'PENDING INSPECTION'
                c.execute('UPDATE rfis SET status=? WHERE id=?',(status,rid)); c.commit(); flash('✅ RFI approval stage signed and workflow advanced.','success')
    r=c.execute("SELECT r.*,p.name project_name,p.client,p.consultant,u.full_name submitter FROM rfis r JOIN projects p ON p.id=r.project_id LEFT JOIN users u ON u.id=r.submitted_by WHERE r.id=?",(rid,)).fetchone(); steps=c.execute("SELECT s.*,u.full_name,u.username,u.department FROM rfi_steps s LEFT JOIN users u ON u.id=s.assigned_user_id WHERE s.rfi_id=? ORDER BY s.step_order",(rid,)).fetchall(); c.close(); return render_template('rfi_detail.html',r=r,inspections=steps)

@app.route("/projects/<int:pid>/rfi/<int:rid>/print")
@login_required
def rfi_print(pid,rid):
    if not allowed_project(pid): return redirect(url_for("dashboard"))
    c=db(); r=c.execute("SELECT r.*,p.name project_name,p.client,p.consultant,p.contractor_role,u.full_name submitter FROM rfis r JOIN projects p ON p.id=r.project_id LEFT JOIN users u ON u.id=r.submitted_by WHERE r.id=? AND r.project_id=?",(rid,pid)).fetchone(); inspections=c.execute("SELECT s.*,u.full_name,u.department,u.username FROM rfi_steps s LEFT JOIN users u ON u.id=s.assigned_user_id WHERE s.rfi_id=? ORDER BY s.step_order",(rid,)).fetchall(); c.close()
    if not r: return ("RFI not found",404)
    return render_template("rfi_print.html",r=r,inspections=inspections)


@app.route("/admin/users/<int:uid>/project-assignment", methods=["POST"])
@admin_required
def assign_project_role(uid):
    c=db()
    try:
        pid=int(request.form["project_id"]); position=request.form.get("project_position","").strip(); manager=request.form.get("project_manager_id") or None
        userrow=c.execute("SELECT personnel_scope FROM users WHERE id=?",(uid,)).fetchone()
        if not userrow: raise ValueError("User not found.")
        if (userrow['personnel_scope'] or "PROJECT")!="PROJECT": raise ValueError("Head Office personnel cannot be added as Project Team personnel. Assign them as Head Office responsible personnel instead.")
        c.execute("INSERT INTO user_projects(user_id,project_id) VALUES(?,?) ON CONFLICT(user_id,project_id) DO NOTHING",(uid,pid))
        c.execute("INSERT INTO project_assignments(user_id,project_id,position,manager_user_id,active) VALUES(?,?,?,?,1) ON CONFLICT(user_id,project_id) DO UPDATE SET position=excluded.position,manager_user_id=excluded.manager_user_id,active=1",(uid,pid,position,manager))
        c.execute("DELETE FROM responsibilities WHERE subordinate_user_id=? AND scope_type='PROJECT' AND project_id=?",(uid,pid))
        if manager: c.execute("INSERT OR IGNORE INTO responsibilities(supervisor_user_id,subordinate_user_id,scope_type,project_id,source) VALUES(?,?,?,?,?)",(manager,uid,'PROJECT',pid,'Project Assignment'))
        c.commit(); flash("🏗️ Project assignment, position and reporting line saved.","success")
    except Exception as e: c.rollback(); flash("Project assignment failed: "+str(e),"error")
    c.close(); return redirect(url_for("users"))

@app.route("/admin/users/<int:uid>/project-assignment/<int:aid>/remove", methods=["POST"])
@admin_required
def remove_project_assignment(uid,aid):
    c=db(); c.execute("UPDATE project_assignments SET active=0 WHERE id=? AND user_id=?",(aid,uid)); pa=c.execute("SELECT project_id FROM project_assignments WHERE id=?",(aid,)).fetchone(); c.execute("DELETE FROM user_projects WHERE user_id=? AND project_id=?",(uid,pa["project_id"] if pa else -1)); c.execute("UPDATE responsibilities SET active=0 WHERE subordinate_user_id=? AND scope_type='PROJECT' AND project_id=?",(uid,pa["project_id"] if pa else -1)); c.commit(); c.close(); flash("🏗️ Project assignment removed; staff record remains permanent.","success"); return redirect(url_for("users"))


@app.route("/projects/<int:pid>/team")
@login_required
def project_team(pid):
    if not allowed_project(pid): return redirect(url_for("dashboard"))
    c=db(); p=c.execute("SELECT * FROM projects WHERE id=?",(pid,)).fetchone()
    rows=c.execute("SELECT pa.*,u.full_name,u.staff_id,u.department,u.phone,u.email,u.org_unit_id,u.personnel_scope,ou.name org_unit_name,m.full_name manager_name FROM project_assignments pa JOIN users u ON u.id=pa.user_id LEFT JOIN org_units ou ON ou.id=u.org_unit_id LEFT JOIN users m ON m.id=pa.manager_user_id WHERE pa.project_id=? AND pa.active=1 AND u.active=1 ORDER BY u.full_name",(pid,)).fetchall()
    project_users=c.execute("SELECT u.id,u.full_name,u.department,u.position,o.name org_unit_name FROM users u LEFT JOIN org_units o ON o.id=u.org_unit_id WHERE u.active=1 AND u.personnel_scope='PROJECT' ORDER BY u.full_name").fetchall()
    head_users=c.execute("SELECT u.id,u.full_name,u.department,u.position,o.name org_unit_name FROM users u LEFT JOIN org_units o ON o.id=u.org_unit_id WHERE u.active=1 AND u.personnel_scope='HEAD_OFFICE' ORDER BY u.full_name").fetchall()
    responsibility_map={}
    for area in ['General Project','Store','Machinery','Fuel','Manpower / HR','Finance','Design','Procurement','Project Management','HSE / Safety','QA/QC','Survey','Contract Administration']:
        responsibility_map[area]={r['user_id'] for r in c.execute("SELECT user_id FROM project_responsibilities WHERE project_id=? AND responsibility_area=? AND active=1",(pid,area)).fetchall()}
    c.close(); return render_template("project_team.html",pid=pid,p=p,rows=rows,project_users=project_users,head_users=head_users,responsibility_map=responsibility_map)

def _request_scope_allowed(pid=None):
    me=current_user()
    if not me: return False
    if pid is None: return True
    return allowed_project(pid)


def _request_rows_for_user(c, me, pid=None):
    where=[]; args=[]
    if me['role']=='SUPER_ADMIN':
        if pid is not None:
            where.append('rr.project_id=?'); args.append(pid)
    else:
        if pid is not None and project_admin(pid):
            where.append('rr.project_id=?'); args.append(pid)
        else:
            scope=['rr.requested_by=?','rr.next_approver_user_id=?']
            scope_args=[me['id'],me['id']]
            if me['org_unit_id']:
                scope.append('rr.requester_org_unit_id=?'); scope_args.append(me['org_unit_id'])
            where.append('('+' OR '.join(scope)+')'); args.extend(scope_args)
            if pid is not None:
                where.append('rr.project_id=?'); args.append(pid)
    w=(' WHERE '+' AND '.join(where)) if where else ''
    return c.execute("SELECT rr.*,p.name project_name,u.full_name requester,ou.name requester_unit,au.full_name approver FROM resource_requests rr LEFT JOIN projects p ON p.id=rr.project_id LEFT JOIN users u ON u.id=rr.requested_by LEFT JOIN org_units ou ON ou.id=rr.requester_org_unit_id LEFT JOIN users au ON au.id=rr.next_approver_user_id"+w+" ORDER BY rr.created_at DESC,rr.id DESC",tuple(args)).fetchall()


def register_approved_request(c, row):
    payload=json.loads(row['payload_json'] or '{}')
    typ=row['request_type']; pid=row['project_id']; uid=row['approved_by'] or row['requested_by']
    if typ in ('MATERIAL','PROCUREMENT'):
        if not pid: return None,None
        name=payload.get('material_name') or row['title']; unit=row['unit'] or payload.get('unit') or 'pcs'; category=payload.get('category','Other')
        m=c.execute("SELECT id FROM materials WHERE project_id=? AND name=?",(pid,name)).fetchone()
        if not m:
            c.execute("INSERT INTO materials(project_id,category,name,unit,min_stock) VALUES(?,?,?,?,?)",(pid,category,name,unit,parse_float(payload.get('min_stock')))); mid=c.execute("SELECT last_insert_rowid() id").fetchone()['id']
        else: mid=m['id']
        c.execute("INSERT INTO store_logs(project_id,material_id,date,received,issued,unit_cost,physical_balance,reference,notes,user_id) VALUES(?,?,?,?,?,?,?,?,?,?)",(pid,mid,payload.get('date',dt.date.today().isoformat()),row['quantity'],0,row['amount']/row['quantity'] if row['quantity'] else parse_float(payload.get('unit_cost')),None,row['request_no'] or f'REQ-{row["id"]}',row['description'],uid))
        return 'store_logs',c.execute("SELECT last_insert_rowid() id").fetchone()['id']
    if typ=='FUEL':
        if not pid: return None,None
        mid=payload.get('machine_id')
        if not mid: raise ValueError('Fuel request requires a machine.')
        c.execute("INSERT INTO fuel_logs(project_id,machine_id,date,opening_gauge,fuel_received,closing_gauge,fuel_price,reference,notes,user_id,source) VALUES(?,?,?,?,?,?,?,?,?,?,?)",(pid,mid,payload.get('date',dt.date.today().isoformat()),parse_float(payload.get('opening_gauge')),row['quantity'],parse_float(payload.get('closing_gauge')),parse_float(payload.get('fuel_price')),row['request_no'] or f'REQ-{row["id"]}',row['description'],uid,'Approved Request'))
        return 'fuel_logs',c.execute("SELECT last_insert_rowid() id").fetchone()['id']
    if typ=='MACHINERY':
        if not pid: return None,None
        c.execute("INSERT INTO machines(project_id,machine_type,code,plate_no,engine_no,ownership,hourly_rate,rate_unit,expected_fuel,fuel_price,active) VALUES(?,?,?,?,?,?,?,?,?,?,1)",(pid,payload.get('machine_type','Other'),payload.get('code',''),payload.get('plate_no',''),payload.get('engine_no',''),payload.get('ownership','Own'),parse_float(payload.get('rate')),payload.get('rate_unit','hr'),parse_float(payload.get('expected_fuel')),parse_float(payload.get('fuel_price'))))
        return 'machines',c.execute("SELECT last_insert_rowid() id").fetchone()['id']
    if typ=='MANPOWER':
        if not pid: return None,None
        c.execute("INSERT INTO manpower(project_id,date,name,employment,position,present,working_hours,hourly_rate,daily_rate,notes,user_id) VALUES(?,?,?,?,?,?,?,?,?,?,?)",(pid,payload.get('date',dt.date.today().isoformat()),payload.get('name',row['title']),payload.get('employment','Temporary'),payload.get('position','Other'),1,parse_float(payload.get('working_hours',8)),parse_float(payload.get('hourly_rate')),parse_float(payload.get('daily_rate')),row['description'],uid))
        return 'manpower',c.execute("SELECT last_insert_rowid() id").fetchone()['id']
    if typ=='EXPENSE':
        if not pid: return None,None
        c.execute("INSERT INTO finance_logs(project_id,date,category,kind,description,amount,reference,user_id) VALUES(?,?,?,?,?,?,?,?)",(pid,payload.get('date',dt.date.today().isoformat()),payload.get('category','Company Expense'),'Expense',row['description'],row['amount'],row['reference'] or row['request_no'],uid))
        return 'finance_logs',c.execute("SELECT last_insert_rowid() id").fetchone()['id']
    if typ=='DESIGN':
        if not pid: return None,None
        c.execute("INSERT INTO design_items(project_id,drawing_no,title,discipline,revision,status,submitted,comments,user_id) VALUES(?,?,?,?,?,?,?,?,?)",(pid,payload.get('drawing_no',''),row['title'],payload.get('discipline','General'),payload.get('revision',''),payload.get('status','Submitted'),payload.get('submitted',dt.date.today().isoformat()),row['description'],uid))
        return 'design_items',c.execute("SELECT last_insert_rowid() id").fetchone()['id']
    return 'resource_requests',row['id']


@app.route('/requests', methods=['GET','POST'])
@app.route('/projects/<int:pid>/requests', methods=['GET','POST'])
@login_required
def resource_requests(pid=None):
    me=current_user()
    if pid is not None and not allowed_project(pid): return redirect(url_for('dashboard'))
    c=db()
    projects=c.execute("SELECT p.* FROM projects p ORDER BY p.name").fetchall() if me['role']=='SUPER_ADMIN' else c.execute("SELECT p.* FROM projects p JOIN user_projects up ON up.project_id=p.id WHERE up.user_id=? ORDER BY p.name",(me['id'],)).fetchall()
    machines=c.execute("SELECT * FROM machines WHERE active=1 AND (? IS NULL OR project_id=?) ORDER BY machine_type,code",(pid,pid)).fetchall()
    units=c.execute("SELECT id,name,unit_type FROM org_units WHERE active=1 ORDER BY sort_order,name").fetchall()
    users=c.execute("SELECT u.id,u.full_name,u.department,u.position,u.org_unit_id,o.name org_unit_name,o.unit_type,o.parent_id FROM users u LEFT JOIN org_units o ON o.id=u.org_unit_id WHERE u.active=1 ORDER BY u.full_name").fetchall()
    if request.method=='POST':
        try:
            typ=request.form.get('request_type','OTHER'); rpid=request.form.get('project_id') or pid or None
            if typ not in REQUEST_TYPES: raise ValueError('Invalid request type.')
            dept=me['department']
            permitted={'MATERIAL':{'Store','Project'},'PROCUREMENT':{'Store','Project','Administration'},'FUEL':{'Machinery','Project'},'MACHINERY':{'Machinery','Project'},'MANPOWER':{'HR','Project'},'EXPENSE':{'Finance','Project','Consultant'},'DESIGN':{'Design','Project'},'OTHER':set(DEPARTMENTS)}
            if me['role']!='SUPER_ADMIN' and dept not in permitted.get(typ,set()): raise ValueError(f'{typ} requests are not enabled for the {dept} department.')
            if rpid and not allowed_project(int(rpid)): raise ValueError('You are not assigned to the selected project.')
            title=request.form.get('title','').strip(); desc=request.form.get('description','').strip()
            if not title: raise ValueError('Request title is required.')
            attachment=request.files.get('attachment'); stored=None; original=None
            if attachment and attachment.filename:
                ext=secure_filename(attachment.filename).rsplit('.',1)[-1].lower() if '.' in attachment.filename else ''
                if ext not in ALLOWED_FILE_EXT: raise ValueError('Unsupported attachment type.')
                stored=f"request_{dt.datetime.now().strftime('%Y%m%d%H%M%S')}_{me['id']}_{secure_filename(attachment.filename)}"; attachment.save(os.path.join(WORKFLOW_FILES,stored)); original=attachment.filename
            payload={k:v for k,v in request.form.items() if k not in {'request_type','project_id','title','description','quantity','unit','amount'}}
            steps=project_request_steps(c,int(rpid),me['id']) if rpid else []
            approver=steps[0][1] if steps else request_approver(c,me['id'],int(rpid) if rpid else None)
            count=c.execute("SELECT COUNT(*) FROM resource_requests").fetchone()[0]+1; no=f"REQ-{dt.date.today().year}-{count:05d}"
            c.execute("INSERT INTO resource_requests(request_no,request_type,project_id,requested_by,requester_org_unit_id,next_approver_user_id,title,description,quantity,unit,amount,payload_json,attachment_file,attachment_name,current_stage,origin_scope) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",(no,typ,int(rpid) if rpid else None,me['id'],me['org_unit_id'],approver,title,desc,parse_float(request.form.get('quantity')),request.form.get('unit',''),parse_float(request.form.get('amount')),json.dumps(payload),stored,original,'PROJECT' if rpid else 'HEAD_OFFICE','PROJECT' if rpid else 'HEAD_OFFICE'))
            rid=c.execute("SELECT last_insert_rowid() id").fetchone()['id']
            for i,(stage,uid) in enumerate(steps,1): c.execute("INSERT INTO request_steps(request_id,step_order,stage,assigned_user_id,status) VALUES(?,?,?,?,?)",(rid,i,stage,uid,'PENDING'))
            if not steps and approver: c.execute("INSERT INTO request_steps(request_id,step_order,stage,assigned_user_id,status) VALUES(?,?,?,?,?)",(rid,1,'HEAD_OFFICE_MANAGER',approver,'PENDING'))
            c.commit(); flash(f'📨 {no} submitted. It entered the approval chain.','success')
        except Exception as e: c.rollback(); flash('Request failed: '+str(e),'error')
    rows=_request_rows_for_user(c,me,pid)
    pending_count=c.execute("SELECT COUNT(*) n FROM resource_requests WHERE status='SUBMITTED' AND (requested_by=? OR next_approver_user_id=? OR ?='SUPER_ADMIN')",(me['id'],me['id'],me['role'])).fetchone()['n']
    c.close(); return render_template('requests.html',pid=pid,projects=projects,machines=machines,rows=rows,pending_count=pending_count,users=users,units=units,departments=DEPARTMENTS)


@app.route('/requests/<int:rid>/<action>', methods=['POST'])
@login_required
def process_resource_request(rid,action):
    me=current_user(); c=db(); row=c.execute("SELECT * FROM resource_requests WHERE id=?",(rid,)).fetchone()
    if not row: c.close(); flash('Request not found.','error'); return redirect(url_for('resource_requests'))
    try:
        if action not in ('approve','reject'): raise ValueError('Invalid request action.')
        if not user_can_approve_request(me,row): raise ValueError('Only the currently assigned responsible person can approve or reject this request.')
        step=current_request_step(c,rid)
        if action=='reject':
            if step: c.execute("UPDATE request_steps SET status='REJECTED',action='REJECTED',comments=?,acted_at=CURRENT_TIMESTAMP WHERE id=?",(request.form.get('reason','Rejected by approver'),step['id']))
            c.execute("UPDATE resource_requests SET status='REJECTED',rejected_by=?,rejected_at=CURRENT_TIMESTAMP,rejection_reason=?,next_approver_user_id=NULL WHERE id=?",(me['id'],dt.datetime.now().isoformat(timespec='seconds'),request.form.get('reason','Rejected by approver'),rid))
            flash(f'❌ {row["request_no"]} rejected.','success')
        else:
            if step: c.execute("UPDATE request_steps SET status='APPROVED',action='APPROVED',comments=?,acted_at=CURRENT_TIMESTAMP WHERE id=?",(request.form.get('comments',''),step['id']))
            nxt=current_request_step(c,rid)
            if nxt:
                c.execute("UPDATE resource_requests SET next_approver_user_id=?,current_stage=?,status='SUBMITTED' WHERE id=?",(nxt['assigned_user_id'],nxt['stage'],rid)); flash(f'✅ {row["request_no"]} approved at this stage and sent to the next project approver.','success')
            elif row['project_id'] and me['role']!='SUPER_ADMIN' and (row['current_stage'] or '')!='HEAD_OFFICE_REVIEW':
                c.execute("UPDATE resource_requests SET next_approver_user_id=NULL,current_stage='AWAITING_HEAD_OFFICE',status='AWAITING_HEAD_OFFICE' WHERE id=?",(rid,)); flash(f'✅ {row["request_no"]} passed the project approval chain. Project Manager must now send it to Head Office.','success')
            else:
                c.execute("UPDATE resource_requests SET status='APPROVED',approved_by=?,approved_at=CURRENT_TIMESTAMP,finalized_at=CURRENT_TIMESTAMP,next_approver_user_id=NULL WHERE id=?",(me['id'],rid)); row=c.execute("SELECT * FROM resource_requests WHERE id=?",(rid,)).fetchone(); table,regid=register_approved_request(c,row); c.execute("UPDATE resource_requests SET registered_table=?,registered_id=? WHERE id=?",(table,regid,rid)); flash(f'✅ {row["request_no"]} finally approved and registered in {table}.','success')
        c.commit()
    except Exception as e: c.rollback(); flash('Approval failed: '+str(e),'error')
    c.close(); return redirect(url_for('resource_requests',pid=row['project_id']) if row['project_id'] else url_for('resource_requests'))


@app.route('/requests/files/<path:filename>')
@login_required
def resource_request_file(filename):
    c=db(); me=current_user(); row=c.execute("SELECT * FROM resource_requests WHERE attachment_file=? AND (requested_by=? OR next_approver_user_id=? OR approved_by=? OR ?='SUPER_ADMIN')",(filename,me['id'],me['id'],me['id'],me['role'])).fetchone(); c.close()
    if not row: return ('File not found or access denied',404)
    return send_from_directory(WORKFLOW_FILES,filename,as_attachment=True,download_name=row['attachment_name'] or filename)


@app.route('/responsibilities')
@login_required
def responsibilities():
    me=current_user(); c=db()
    head=responsibility_users(me['id'],'HEAD_OFFICE')
    project=responsibility_users(me['id'],'PROJECT')
    project_function=c.execute("SELECT pr.*,p.name project_name,u.full_name,u.staff_id,u.department,u.position FROM project_responsibilities pr JOIN projects p ON p.id=pr.project_id JOIN users u ON u.id=pr.user_id WHERE pr.user_id=? AND pr.active=1 ORDER BY p.name,pr.responsibility_area,u.full_name",(me['id'],)).fetchall()
    my_projects=c.execute("SELECT pa.*,p.name project_name FROM project_assignments pa JOIN projects p ON p.id=pa.project_id WHERE pa.user_id=? AND pa.active=1 ORDER BY p.name",(me['id'],)).fetchall()
    c.close(); return render_template('responsibilities.html',head=head,project=project,project_function=project_function,my_projects=my_projects)


def visible_workflow_contacts(c, me):
    if me['role']=='SUPER_ADMIN':
        users=c.execute("SELECT id,full_name,department,position,org_unit_id FROM users WHERE active=1 ORDER BY full_name").fetchall()
        units=c.execute("SELECT id,name,unit_type FROM org_units WHERE active=1 ORDER BY sort_order,name").fetchall()
        return users,units
    ids={me['id']}
    if me['reports_to_user_id']: ids.add(me['reports_to_user_id'])
    for r in c.execute("SELECT subordinate_user_id FROM responsibilities WHERE supervisor_user_id=? AND active=1",(me['id'],)).fetchall(): ids.add(r['subordinate_user_id'])
    if me['org_unit_id']:
        for r in c.execute("SELECT id FROM users WHERE active=1 AND org_unit_id=?",(me['org_unit_id'],)).fetchall(): ids.add(r['id'])
    for r in c.execute("SELECT DISTINCT up2.user_id FROM user_projects up1 JOIN user_projects up2 ON up2.project_id=up1.project_id WHERE up1.user_id=? AND up2.user_id<>?",(me['id'],me['id'])).fetchall(): ids.add(r['user_id'])
    # Project responsibility contacts include Head Office supervisors assigned to this project.
    for r in c.execute("SELECT DISTINCT pr.user_id FROM project_responsibilities pr WHERE pr.project_id IN (SELECT project_id FROM user_projects WHERE user_id=? UNION SELECT project_id FROM project_responsibilities WHERE user_id=?) AND pr.active=1",(me['id'],me['id'])).fetchall(): ids.add(r['user_id'])
    if ids:
        ph=','.join('?'*len(ids)); users=c.execute(f"SELECT id,full_name,department,position,org_unit_id FROM users WHERE active=1 AND id IN ({ph}) ORDER BY full_name",tuple(ids)).fetchall()
    else: users=[]
    unit_ids=set()
    if me['org_unit_id']: unit_ids.add(me['org_unit_id'])
    for r in users:
        if r['org_unit_id']: unit_ids.add(r['org_unit_id'])
    if me['org_unit_id']:
        parent=c.execute("SELECT parent_id FROM org_units WHERE id=?",(me['org_unit_id'],)).fetchone()
        if parent and parent['parent_id']: unit_ids.add(parent['parent_id'])
    if unit_ids:
        ph=','.join('?'*len(unit_ids)); units=c.execute(f"SELECT id,name,unit_type FROM org_units WHERE active=1 AND id IN ({ph}) ORDER BY sort_order,name",tuple(unit_ids)).fetchall()
    else: units=[]
    return users,units


@app.route("/workflow", methods=["GET","POST"])
@login_required
def workflow():
    c=db(); me=current_user(); projects=c.execute("SELECT p.* FROM projects p JOIN user_projects up ON up.project_id=p.id WHERE up.user_id=? ORDER BY p.name",(me["id"],)).fetchall() if me["role"]!="SUPER_ADMIN" else c.execute("SELECT * FROM projects ORDER BY name").fetchall()
    users,units=visible_workflow_contacts(c,me)
    if request.method=="POST":
        allowed_user_ids={r['id'] for r in users}; allowed_unit_ids={r['id'] for r in units}
        try:
            to_user=request.form.get("to_user_id") or None; to_unit=request.form.get("to_org_unit_id") or None; pid=request.form.get("project_id") or None
            if to_user and int(to_user) not in allowed_user_ids: raise ValueError('Recipient is outside your permitted contact hierarchy.')
            if to_unit and int(to_unit) not in allowed_unit_ids: raise ValueError('Recipient team is outside your permitted contact hierarchy.')
            if not to_user and not to_unit: raise ValueError('Select a recipient person or department/team.')
            subject=request.form.get("subject","").strip(); message=request.form.get("message","").strip(); category=request.form.get("category","General Correspondence")
            f=request.files.get("file")
            if not subject: raise ValueError("Subject is required.")
            if not f or not f.filename: raise ValueError("Select a file to send.")
            ext=secure_filename(f.filename).rsplit('.',1)[-1].lower() if '.' in f.filename else ''
            if ext not in ALLOWED_FILE_EXT: raise ValueError("Unsupported file type.")
            if request.content_length and request.content_length>MAX_UPLOAD_MB*1024*1024: raise ValueError(f"File must be {MAX_UPLOAD_MB} MB or smaller.")
            stored=f"{dt.datetime.now().strftime('%Y%m%d%H%M%S')}_{me['id']}_{secure_filename(f.filename)}"; f.save(os.path.join(WORKFLOW_FILES,stored))
            c.execute("INSERT INTO workflow_files(project_id,from_user_id,to_user_id,to_org_unit_id,file_name,stored_name,file_type,subject,message,category,created_by) VALUES(?,?,?,?,?,?,?,?,?,?,?)",(pid,me["id"],to_user,to_unit,f.filename,stored,ext,subject,message,category,me["id"]))
            c.commit(); flash("📤 File sent through the company hierarchy.","success")
        except Exception as e: c.rollback(); flash("File workflow failed: "+str(e),"error")
    sent=c.execute("SELECT wf.*,fu.full_name sender,tu.full_name receiver,ou.name receiver_unit,p.name project_name FROM workflow_files wf LEFT JOIN users fu ON fu.id=wf.from_user_id LEFT JOIN users tu ON tu.id=wf.to_user_id LEFT JOIN org_units ou ON ou.id=wf.to_org_unit_id LEFT JOIN projects p ON p.id=wf.project_id WHERE wf.from_user_id=? ORDER BY wf.sent_at DESC LIMIT 100",(me["id"],)).fetchall()
    received=c.execute("SELECT wf.*,fu.full_name sender,tu.full_name receiver,ou.name receiver_unit,p.name project_name FROM workflow_files wf LEFT JOIN users fu ON fu.id=wf.from_user_id LEFT JOIN users tu ON tu.id=wf.to_user_id LEFT JOIN org_units ou ON ou.id=wf.to_org_unit_id LEFT JOIN projects p ON p.id=wf.project_id WHERE wf.to_user_id=? OR wf.to_org_unit_id=? ORDER BY wf.sent_at DESC LIMIT 100",(me["id"],me["org_unit_id"] or -1)).fetchall()
    c.close(); return render_template("workflow.html",projects=projects,users=users,units=units,sent=sent,received=received)

@app.route("/workflow/files/<path:filename>")
@login_required
def workflow_file(filename):
    c=db(); me=current_user(); row=c.execute("SELECT * FROM workflow_files WHERE stored_name=? AND (from_user_id=? OR to_user_id=? OR to_org_unit_id=? OR ?='SUPER_ADMIN')",(filename,me["id"],me["id"],me["org_unit_id"] or -1,me["role"])).fetchone(); c.close()
    if not row: return ("File not found or access denied",404)
    return send_from_directory(WORKFLOW_FILES,filename,as_attachment=True,download_name=row["file_name"])

@app.route("/workflow/files/<int:wfid>/receive",methods=["POST"])
@login_required
def receive_workflow_file(wfid):
    c=db(); me=current_user(); c.execute("UPDATE workflow_files SET status='RECEIVED',received_at=? WHERE id=? AND (to_user_id=? OR to_org_unit_id=?)",(dt.datetime.now().isoformat(timespec="seconds"),wfid,me["id"],me["org_unit_id"] or -1)); c.commit(); c.close(); flash("📥 File marked as received.","success"); return redirect(url_for("workflow"))

@app.route("/projects/<int:pid>/transfers",methods=["GET","POST"])
@login_required
def transfers(pid):
    if not allowed_project(pid): return redirect(url_for("dashboard"))
    c=db(); me=current_user(); projects=c.execute("SELECT * FROM projects WHERE id<>? ORDER BY name",(pid,)).fetchall(); materials=c.execute("SELECT * FROM materials WHERE project_id=? AND active=1 ORDER BY category,name",(pid,)).fetchall(); machines=c.execute("SELECT * FROM machines WHERE project_id=? AND active=1 ORDER BY machine_type,code",(pid,)).fetchall()
    if request.method=="POST":
        action=request.form.get("action")
        try:
            to_pid=int(request.form["to_project_id"]); date=request.form.get("date",dt.date.today().isoformat()); ref=request.form.get("reference",""); notes=request.form.get("notes","")
            if action=="material":
                mid=int(request.form["material_id"]); qty=parse_float(request.form.get("quantity")); unit_cost=parse_float(request.form.get("unit_cost"))
                if qty<=0: raise ValueError('Transfer quantity must be greater than zero.')
                bal=c.execute("SELECT COALESCE(SUM(received-issued),0) FROM store_logs WHERE project_id=? AND material_id=?",(pid,mid)).fetchone()[0]
                if qty>bal: raise ValueError(f'Insufficient source stock. Available balance: {bal:g}.')
                c.execute("INSERT INTO material_transfers(from_project_id,to_project_id,material_id,date,quantity,unit_cost,reference,notes,sent_by) VALUES(?,?,?,?,?,?,?,?,?)",(pid,to_pid,mid,date,qty,unit_cost,ref,notes,me["id"]))
                c.execute("INSERT INTO store_logs(project_id,material_id,date,received,issued,unit_cost,reference,notes,user_id) VALUES(?,?,?,?,?,?,?,?,?)",(pid,mid,date,0,qty,unit_cost,ref or f'Transfer to project {to_pid}',notes or 'Inter-project material transfer sent',me['id']))
            elif action=="fuel": c.execute("INSERT INTO fuel_transfers(from_project_id,to_project_id,machine_id,date,litres,unit_cost,reference,notes,sent_by) VALUES(?,?,?,?,?,?,?,?,?)",(pid,to_pid,request.form.get("machine_id") or None,date,parse_float(request.form.get("litres")),parse_float(request.form.get("unit_cost")),ref,notes,me["id"]))
            elif action=="machine": c.execute("INSERT INTO machine_transfers(from_project_id,to_project_id,machine_id,date,notes,sent_by) VALUES(?,?,?,?,?,?)",(pid,to_pid,request.form["machine_id"],date,notes,me["id"]))
            c.commit(); flash("🔄 Transfer sent to the receiving project store/team for confirmation.","success")
        except Exception as e: c.rollback(); flash("Transfer failed: "+str(e),"error")
    outgoing=c.execute("SELECT mt.*,m.name material_name,m.unit,t.name to_project,fu.full_name sender FROM material_transfers mt JOIN materials m ON m.id=mt.material_id JOIN projects t ON t.id=mt.to_project_id LEFT JOIN users fu ON fu.id=mt.sent_by WHERE mt.from_project_id=? ORDER BY mt.sent_at DESC",(pid,)).fetchall()
    incoming=c.execute("SELECT mt.*,m.name material_name,m.unit,f.name from_project,fu.full_name sender FROM material_transfers mt JOIN materials m ON m.id=mt.material_id JOIN projects f ON f.id=mt.from_project_id LEFT JOIN users fu ON fu.id=mt.sent_by WHERE mt.to_project_id=? ORDER BY mt.sent_at DESC",(pid,)).fetchall()
    fuel_out=c.execute("SELECT ft.*,p.name to_project,fu.full_name sender,m.code,m.machine_type FROM fuel_transfers ft JOIN projects p ON p.id=ft.to_project_id LEFT JOIN users fu ON fu.id=ft.sent_by LEFT JOIN machines m ON m.id=ft.machine_id WHERE ft.from_project_id=? ORDER BY ft.sent_at DESC",(pid,)).fetchall()
    fuel_in=c.execute("SELECT ft.*,p.name from_project,fu.full_name sender,m.code,m.machine_type FROM fuel_transfers ft JOIN projects p ON p.id=ft.from_project_id LEFT JOIN users fu ON fu.id=ft.sent_by LEFT JOIN machines m ON m.id=ft.machine_id WHERE ft.to_project_id=? ORDER BY ft.sent_at DESC",(pid,)).fetchall()
    mach_out=c.execute("SELECT mt.*,p.name to_project,m.code,m.machine_type FROM machine_transfers mt JOIN projects p ON p.id=mt.to_project_id JOIN machines m ON m.id=mt.machine_id WHERE mt.from_project_id=? ORDER BY mt.sent_at DESC",(pid,)).fetchall()
    mach_in=c.execute("SELECT mt.*,p.name from_project,m.code,m.machine_type FROM machine_transfers mt JOIN projects p ON p.id=mt.from_project_id JOIN machines m ON m.id=mt.machine_id WHERE mt.to_project_id=? ORDER BY mt.sent_at DESC",(pid,)).fetchall()
    c.close(); return render_template("transfers.html",pid=pid,projects=projects,materials=materials,machines=machines,outgoing=outgoing,incoming=incoming,fuel_out=fuel_out,fuel_in=fuel_in,mach_out=mach_out,mach_in=mach_in)

@app.route("/projects/<int:pid>/transfers/<string:kind>/<int:tid>/receive",methods=["POST"])
@login_required
def receive_transfer(pid,kind,tid):
    if not allowed_project(pid): return redirect(url_for("dashboard"))
    c=db(); me=current_user(); table={'material':'material_transfers','fuel':'fuel_transfers','machine':'machine_transfers'}.get(kind)
    if not table: c.close(); return ("Invalid transfer",400)
    required={'material':'Store','fuel':'Machinery','machine':'Machinery'}[kind]
    if me['role']!='SUPER_ADMIN' and me['department'] not in (required,'Project'):
        c.close(); flash(f"🚫 Only {required} / Project personnel can receive this transfer.","error"); return redirect(url_for("transfers",pid=pid))
    now=dt.datetime.now().isoformat(timespec="seconds")
    row=c.execute(f"SELECT * FROM {table} WHERE id=? AND to_project_id=? AND status='SENT'",(tid,pid)).fetchone()
    if not row: c.close(); flash("Transfer not found, already received, or not addressed to this project.","error"); return redirect(url_for("transfers",pid=pid))
    c.execute(f"UPDATE {table} SET status='RECEIVED',received_by=?,received_at=? WHERE id=?",(me["id"],now,tid))
    if kind=='material':
        src=c.execute("SELECT * FROM materials WHERE id=?",(row['material_id'],)).fetchone()
        if src:
            target=c.execute("SELECT id FROM materials WHERE project_id=? AND name=?",(pid,src['name'])).fetchone()
            if not target:
                c.execute("INSERT INTO materials(project_id,category,name,unit,min_stock) VALUES(?,?,?,?,?)",(pid,src['category'],src['name'],src['unit'],src['min_stock'])); target=c.execute("SELECT last_insert_rowid() id").fetchone()
            c.execute("INSERT INTO store_logs(project_id,material_id,date,received,issued,unit_cost,reference,notes,user_id) VALUES(?,?,?,?,?,?,?,?,?)",(pid,target['id'],row['date'],row['quantity'],0,row['unit_cost'],row['reference'] or f'Transfer from project {row["from_project_id"]}',row['notes'] or 'Inter-project material transfer received',me['id']))
    elif kind=='machine':
        c.execute("UPDATE machines SET project_id=? WHERE id=?",(pid,row['machine_id']))
    c.commit(); c.close(); flash("📥 Transfer received and recorded in the receiving project.","success"); return redirect(url_for("transfers",pid=pid))

@app.route("/projects/<int:pid>/expenses/claim",methods=["GET","POST"])
@login_required
def expense_claim(pid):
    if not allowed_project(pid): return redirect(url_for("dashboard"))
    c=db(); me=current_user(); users=c.execute("SELECT id,full_name,department,position FROM users WHERE active=1 ORDER BY full_name").fetchall()
    if request.method=="POST":
        try:
            receipt=request.files.get("receipt"); stored=None; original=None
            if receipt and receipt.filename:
                ext=secure_filename(receipt.filename).rsplit('.',1)[-1].lower() if '.' in receipt.filename else ''
                if ext not in ALLOWED_FILE_EXT: raise ValueError("Unsupported receipt file type.")
                stored=f"receipt_{dt.datetime.now().strftime('%Y%m%d%H%M%S')}_{me['id']}_{secure_filename(receipt.filename)}"; receipt.save(os.path.join(WORKFLOW_FILES,stored)); original=receipt.filename
            c.execute("INSERT INTO expense_claims(project_id,date,beneficiary_user_id,beneficiary_name,category,description,amount,paid_by_company,receipt_file,receipt_name,submitted_by) VALUES(?,?,?,?,?,?,?,?,?,?,?)",(pid,request.form.get('date',dt.date.today().isoformat()),request.form.get('beneficiary_user_id') or None,request.form.get('beneficiary_name',''),request.form.get('category','Consultant / Staff'),request.form.get('description',''),parse_float(request.form.get('amount')),1,stored,original,me['id']))
            c.commit(); flash("💳 Company-paid expense submitted with receipt tracking.","success")
        except Exception as e: c.rollback(); flash("Expense claim failed: "+str(e),"error")
    rows=c.execute("SELECT ec.*,u.full_name beneficiary, s.full_name submitter FROM expense_claims ec LEFT JOIN users u ON u.id=ec.beneficiary_user_id LEFT JOIN users s ON s.id=ec.submitted_by WHERE ec.project_id=? ORDER BY ec.created_at DESC",(pid,)).fetchall(); c.close(); return render_template("expense_claim.html",pid=pid,users=users,rows=rows)


@app.route("/projects/<int:pid>/expenses/<int:eid>/<action>",methods=["POST"])
@login_required
def process_expense_claim(pid,eid,action):
    me=current_user()
    if me['role']!='SUPER_ADMIN' and me['department']!='Finance' and not project_admin(pid):
        flash("🚫 Finance / Super Admin approval is required.","error"); return redirect(url_for("expense_claim",pid=pid))
    c=db(); row=c.execute("SELECT * FROM expense_claims WHERE id=? AND project_id=?",(eid,pid)).fetchone()
    if not row: c.close(); flash("Expense claim not found.","error"); return redirect(url_for("expense_claim",pid=pid))
    status='APPROVED' if action=='approve' else ('REJECTED' if action=='reject' else None)
    if not status: c.close(); return ("Invalid action",400)
    now=dt.datetime.now().isoformat(timespec='seconds')
    c.execute("UPDATE expense_claims SET status=?,approved_by=?,approved_at=? WHERE id=?",(status,me['id'],now,eid))
    if status=='APPROVED':
        c.execute("INSERT INTO finance_logs(project_id,date,category,kind,description,amount,reference,user_id) VALUES(?,?,?,?,?,?,?,?)",(pid,row['date'],row['category'],'Expense',f"Company-paid: {row['description']} | Beneficiary: {row['beneficiary_name'] or row['beneficiary_user_id'] or ''}",row['amount'],row['receipt_name'] or '',me['id']))
    c.commit(); c.close(); flash(f"💳 Expense claim {status.lower()}.","success"); return redirect(url_for("expense_claim",pid=pid))

@app.route("/workflow/receipts/<path:filename>")
@login_required
def receipt_file(filename):
    c=db(); me=current_user(); row=c.execute("SELECT * FROM expense_claims WHERE receipt_file=? AND (submitted_by=? OR beneficiary_user_id=? OR ?='SUPER_ADMIN' OR ?='Finance')",(filename,me['id'],me['id'],me['role'],me['department'])).fetchone(); c.close()
    if not row:return ("Receipt not found or access denied",404)
    return send_from_directory(WORKFLOW_FILES,filename,as_attachment=True,download_name=row['receipt_name'] or filename)

@app.route("/uploads/user_photos/<path:filename>")
def user_photo(filename):
    return send_from_directory(USER_PHOTOS, filename)

@app.route("/admin/head-office")
@admin_required
def head_office():
    c=db()
    units=c.execute("SELECT o.*,p.name parent_name,mu.full_name manager_name FROM org_units o LEFT JOIN org_units p ON p.id=o.parent_id LEFT JOIN users mu ON mu.id=o.manager_user_id WHERE o.active=1 ORDER BY o.sort_order,o.name").fetchall()
    staff=c.execute("SELECT u.*,o.name org_name,m.full_name manager_name FROM users u LEFT JOIN org_units o ON o.id=u.org_unit_id LEFT JOIN users m ON m.id=u.reports_to_user_id WHERE u.active=1 ORDER BY COALESCE(o.sort_order,9999),o.name,u.full_name").fetchall()
    users_all=c.execute("SELECT id,full_name,position,department FROM users WHERE active=1 ORDER BY full_name").fetchall()
    unit_staff={}
    for u in units: unit_staff[u["id"]]=[x for x in staff if x["org_unit_id"]==u["id"]]
    c.close()
    return render_template("head_office.html",units=units,staff=staff,users_all=users_all,unit_staff=unit_staff)

@app.route("/admin/head-office/unit",methods=["POST"])
@admin_required
def add_org_unit():
    c=db()
    try:
        name=request.form.get("name","").strip(); parent_id=request.form.get("parent_id") or None; unit_type=request.form.get("unit_type","Team"); manager=request.form.get("manager_user_id") or None
        if not name: raise ValueError("Organization unit name is required.")
        c.execute("INSERT INTO org_units(name,parent_id,unit_type,manager_user_id,sort_order) VALUES(?,?,?,?,COALESCE((SELECT MAX(sort_order)+1 FROM org_units),1))",(name,parent_id,unit_type,manager))
        c.commit(); flash("🏢 Head Office structure unit added.","success")
    except Exception as e: c.rollback(); flash("Could not add structure unit: "+str(e),"error")
    c.close(); return redirect(url_for("head_office"))

@app.route("/admin/head-office/unit/<int:oid>",methods=["POST"])
@admin_required
def edit_org_unit(oid):
    c=db()
    try:
        c.execute("UPDATE org_units SET name=?,parent_id=?,unit_type=?,manager_user_id=? WHERE id=?",(request.form.get("name","").strip(),request.form.get("parent_id") or None,request.form.get("unit_type","Team"),request.form.get("manager_user_id") or None,oid)); c.commit(); flash("🏢 Structure updated.","success")
    except Exception as e: c.rollback(); flash("Could not update structure: "+str(e),"error")
    c.close(); return redirect(url_for("head_office"))

@app.route("/admin/head-office/staff/<int:uid>",methods=["POST"])
@admin_required
def assign_head_office_staff(uid):
    c=db()
    try:
        c.execute("UPDATE users SET personnel_scope='HEAD_OFFICE',org_unit_id=?,reports_to_user_id=? WHERE id=?",(request.form.get('org_unit_id') or None,request.form.get('reports_to_user_id') or None,uid)); c.execute("UPDATE project_assignments SET active=0 WHERE user_id=?",(uid,)); c.execute("DELETE FROM user_projects WHERE user_id=?",(uid,)); c.execute("DELETE FROM responsibilities WHERE subordinate_user_id=? AND scope_type='HEAD_OFFICE'",(uid,)); mgr=request.form.get("reports_to_user_id") or None;
        if mgr: c.execute("INSERT OR IGNORE INTO responsibilities(supervisor_user_id,subordinate_user_id,scope_type,project_id,source) VALUES(?,?,?,NULL,?)",(mgr,uid,'HEAD_OFFICE','Head Office Hierarchy')); c.commit(); flash("👤 Staff hierarchy updated.","success")
    except Exception as e: c.rollback(); flash("Could not update staff hierarchy: "+str(e),"error")
    c.close(); return redirect(url_for("head_office"))

@app.route("/admin/users")
@admin_required
def users():
    c=db(); users=c.execute("SELECT u.*,o.name org_unit_name FROM users u LEFT JOIN org_units o ON o.id=u.org_unit_id ORDER BY u.full_name").fetchall(); projects=c.execute("SELECT * FROM projects ORDER BY name").fetchall(); units=c.execute("SELECT * FROM org_units WHERE active=1 ORDER BY sort_order,name").fetchall()
    assign={u["id"]:[r["project_id"] for r in c.execute("SELECT project_id FROM user_projects WHERE user_id=?",(u["id"],)).fetchall()] for u in users}
    project_assignments={u["id"]:[dict(r) for r in c.execute("SELECT pa.*,p.name project_name,m.full_name manager_name FROM project_assignments pa JOIN projects p ON p.id=pa.project_id LEFT JOIN users m ON m.id=pa.manager_user_id WHERE pa.user_id=? AND pa.active=1 ORDER BY p.name",(u["id"],)).fetchall()] for u in users}
    responsibility_counts={u['id']:{'head':c.execute("SELECT COUNT(*) n FROM responsibilities WHERE supervisor_user_id=? AND scope_type='HEAD_OFFICE' AND active=1",(u['id'],)).fetchone()['n'],'project':c.execute("SELECT COUNT(*) n FROM responsibilities WHERE supervisor_user_id=? AND scope_type='PROJECT' AND active=1",(u['id'],)).fetchone()['n']+c.execute("SELECT COUNT(*) n FROM project_responsibilities WHERE user_id=? AND active=1",(u['id'],)).fetchone()['n']} for u in users}
    responsibility_people={}
    for u in users:
        responsibility_people[u["id"]]=c.execute("SELECT r.scope_type,u.full_name,p.name project_name FROM responsibilities r JOIN users u ON u.id=r.subordinate_user_id LEFT JOIN projects p ON p.id=r.project_id WHERE r.supervisor_user_id=? AND r.active=1 ORDER BY r.scope_type,p.name,u.full_name",(u["id"],)).fetchall()
    # Project personnel available to each project, used by the Head Office contact assignment UI.
    project_contact_people={}
    for p in projects:
        project_contact_people[p['id']]=c.execute("SELECT u.id,u.full_name,u.department,u.position FROM project_assignments pa JOIN users u ON u.id=pa.user_id WHERE pa.project_id=? AND pa.active=1 AND u.active=1 AND u.personnel_scope='PROJECT' ORDER BY u.full_name",(p['id'],)).fetchall()
    ho_contact_links={}
    for u in users:
        if u['personnel_scope']=='HEAD_OFFICE':
            ho_contact_links[u['id']]=c.execute("SELECT pc.project_id,pc.project_user_id,pc.responsibility_area,p.name project_name,pu.full_name project_user_name FROM personnel_project_contacts pc JOIN projects p ON p.id=pc.project_id JOIN users pu ON pu.id=pc.project_user_id WHERE pc.head_office_user_id=? AND pc.active=1 ORDER BY p.name,pc.responsibility_area,pu.full_name",(u['id'],)).fetchall()
    c.close()
    return render_template("users.html",users=users,projects=projects,assign=assign,project_assignments=project_assignments,org_units=units,responsibility_counts=responsibility_counts,responsibility_people=responsibility_people,project_contact_people=project_contact_people,ho_contact_links=ho_contact_links)

@app.route("/admin/users/add",methods=["POST"])
@admin_required
def add_user():
    # SQLite can briefly be busy when a machinery/daily/fuel transaction is committing.
    # Use a short retry loop so Super Admin can create staff without seeing "database is locked".
    import time
    c=db(); photo_path=None
    try:
        photo=request.files.get("photo")
        if not photo or not photo.filename:
            raise ValueError("Staff photo is required. Upload a passport-style JPG, PNG or WEBP photo.")
        ext=secure_filename(photo.filename).rsplit('.',1)[-1].lower() if '.' in photo.filename else ''
        if ext not in ALLOWED_PHOTO_EXT: raise ValueError("Photo must be JPG, JPEG, PNG or WEBP.")
        role=request.form.get("role") if request.form.get("role") in ("STAFF","CONSULTANT","SUPER_ADMIN") else "STAFF"
        scope=request.form.get('personnel_scope','PROJECT') if request.form.get('personnel_scope') in PERSONNEL_SCOPES else 'PROJECT'
        org_id=request.form.get('org_unit_id') or None
        if scope=='HEAD_OFFICE' and not org_id: raise ValueError('Head Office personnel must be assigned to a Head Office department/team.')
        if scope=='PROJECT': org_id=None
        vals=(request.form['full_name'].strip(),request.form['username'].strip(),generate_password_hash(request.form['password']),request.form['department'],request.form.get('position','Other'),request.form.get('location','').strip(),request.form.get('phone','').strip(),request.form.get('email','').strip(),role,org_id,request.form.get('reports_to_user_id') or None,scope)
        for attempt in range(5):
            try:
                c.execute("INSERT INTO users(full_name,username,password_hash,department,position,location,phone,email,role,org_unit_id,reports_to_user_id,personnel_scope,photo_filename) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",vals+ (None,))
                uid=c.execute("SELECT id FROM users WHERE username=?",(request.form["username"].strip(),)).fetchone()["id"]
                staff_id=make_staff_id(request.form["department"],uid)
                filename=f"{staff_id}_{uid}.{ext}"
                photo_path=os.path.join(USER_PHOTOS,filename)
                photo.save(photo_path)
                c.execute("UPDATE users SET staff_id=?,photo_filename=? WHERE id=?",(staff_id,filename,uid))
                if scope=='PROJECT':
                    for pid2 in request.form.getlist('project_ids'):
                        c.execute('INSERT OR IGNORE INTO user_projects(user_id,project_id) VALUES(?,?)',(uid,pid2))
                c.commit()
                flash(f"👤 Staff registered permanently: {staff_id}. The account can be disabled later, but it is never deleted.","success")
                break
            except sqlite3.OperationalError as e:
                if "locked" not in str(e).lower() or attempt==4: raise
                try: c.rollback()
                except Exception: pass
                time.sleep(0.5*(attempt+1))
    except Exception as e:
        try: c.rollback()
        except Exception: pass
        if photo_path and os.path.isfile(photo_path):
            try: os.remove(photo_path)
            except Exception: pass
        flash("Could not create user: "+str(e),"error")
    finally:
        c.close()
    return redirect(url_for("users"))

@app.route("/admin/users/<int:uid>/photo",methods=["POST"])
@admin_required
def update_user_photo(uid):
    photo=request.files.get("photo")
    c=db(); u=c.execute("SELECT * FROM users WHERE id=?",(uid,)).fetchone()
    if not u: c.close(); flash("User not found.","error"); return redirect(url_for("users"))
    try:
        if not photo or not photo.filename: raise ValueError("Select a photo.")
        ext=secure_filename(photo.filename).rsplit('.',1)[-1].lower() if '.' in photo.filename else ''
        if ext not in ALLOWED_PHOTO_EXT: raise ValueError("Photo must be JPG, JPEG, PNG or WEBP.")
        if u['photo_filename']:
            old=os.path.join(USER_PHOTOS,u['photo_filename'])
            if os.path.isfile(old): os.remove(old)
        filename=f"{u['staff_id']}_{uid}.{ext}"
        photo.save(os.path.join(USER_PHOTOS,filename)); c.execute("UPDATE users SET photo_filename=? WHERE id=?",(filename,uid)); c.commit(); flash("📷 Staff photo updated.","success")
    except Exception as e: c.rollback(); flash("Photo update failed: "+str(e),"error")
    c.close(); return redirect(url_for("users"))


CODE128_B_PATTERNS=["212222","222122","222221","121223","121322","131222","122213","122312","132212","221213","221312","231212","112232","122132","122231","113222","123122","123221","223211","221132","221231","213212","223112","312131","311222","321122","321221","312212","322112","322211","212123","212321","232121","111323","131123","131321","112313","132113","132311","211313","231113","231311","112133","112331","132131","113123","113321","133121","313121","211331","231131","213113","213311","213131","311123","311321","331121","312113","312311","332111","314111","221411","431111","111224","111422","121124","121421","141122","141221","112214","112412","122114","122411","142112","142211","241211","221114","413111","241112","134111","111242","121142","121241","114212","124112","124211","411212","421112","421211","212141","214121","412121","111143","111341","131141","114113","114311","411113","411311","113141","114131","311141","411131","211412","211214","211232","2331112"]
def code128_svg(value):
    value=str(value or "STAFF"); vals=[104]+[ord(ch)-32 if 32<=ord(ch)<=127 else 0 for ch in value]; checksum=(104+sum((i+1)*v for i,v in enumerate(vals[1:],1)))%103; vals.append(checksum); vals.append(106)
    x=4; scale=1.35; h=34; parts=[]
    for code in vals:
        pat=CODE128_B_PATTERNS[code]
        black=True
        for n in map(int,pat):
            if black: parts.append(f'<rect x="{x:.2f}" y="0" width="{n*scale:.2f}" height="{h}"/>')
            x+=n*scale; black=not black
    w=x+4
    return f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {w:.2f} 48" preserveAspectRatio="none"><rect width="100%" height="100%" fill="white"/>'+''.join(parts)+f'<text x="{w/2:.2f}" y="46" text-anchor="middle" font-family="Arial" font-size="8">{value}</text></svg>'

@app.route("/admin/users/<int:uid>/id-card")
@admin_required
def user_id_card(uid):
    c=db(); u=c.execute("SELECT * FROM users WHERE id=?",(uid,)).fetchone(); c.close()
    if not u: return ("User not found",404)
    return render_template("user_id_card.html",u=u,barcode_svg=code128_svg(u["staff_id"]))

@app.route("/admin/users/<int:uid>/edit",methods=["POST"])
@admin_required
def edit_user(uid):
    c=db(); u=c.execute("SELECT * FROM users WHERE id=?",(uid,)).fetchone()
    if not u: c.close(); flash("User not found.","error"); return redirect(url_for("users"))
    try:
        role=request.form.get("role") if request.form.get("role") in ("STAFF","CONSULTANT","SUPER_ADMIN") else "STAFF"
        username=request.form.get("username","").strip()
        if not username: raise ValueError("Username is required.")
        clash=c.execute("SELECT id FROM users WHERE username=? AND id<>?",(username,uid)).fetchone()
        if clash: raise ValueError("Username already exists.")
        scope=request.form.get('personnel_scope','PROJECT') if request.form.get('personnel_scope') in PERSONNEL_SCOPES else 'PROJECT'
        org_id=request.form.get('org_unit_id') or None
        if scope=='HEAD_OFFICE' and not org_id: raise ValueError('Head Office personnel must have a Head Office department/team.')
        if scope=='PROJECT': org_id=None
        c.execute("UPDATE users SET full_name=?,username=?,department=?,position=?,location=?,phone=?,email=?,role=?,org_unit_id=?,reports_to_user_id=?,personnel_scope=? WHERE id=?",(request.form.get('full_name','').strip(),username,request.form.get('department','Project'),request.form.get('position','Other'),request.form.get('location','').strip(),request.form.get('phone','').strip(),request.form.get('email','').strip(),role,org_id,request.form.get('reports_to_user_id') or None,scope,uid))
        if scope=='HEAD_OFFICE':
            c.execute('UPDATE project_assignments SET active=0 WHERE user_id=?',(uid,))
            c.execute('DELETE FROM user_projects WHERE user_id=?',(uid,))
        c.commit(); flash('✏️ User profile, personnel type, department, position and role updated.','success')
    except Exception as e: c.rollback(); flash("User update failed: "+str(e),"error")
    c.close(); return redirect(url_for("users"))

@app.route("/admin/users/<int:uid>/reset",methods=["POST"])
@admin_required
def reset_user(uid):
    c=db();p=request.form.get("password","").strip()
    if not p:flash("Password cannot be empty.","error")
    else:c.execute("UPDATE users SET password_hash=?,active=1 WHERE id=?",(generate_password_hash(p),uid));c.commit();flash("🔑 User password reset.","success")
    c.close();return redirect(url_for("users"))

@app.route("/admin/users/<int:uid>/toggle",methods=["POST"])
@admin_required
def toggle_user(uid):
    if uid==current_user()["id"]:flash("You cannot disable your own Super Admin account.","error")
    else:
        c=db();c.execute("UPDATE users SET active=CASE WHEN active=1 THEN 0 ELSE 1 END WHERE id=?",(uid,));c.commit();c.close();flash("User status updated.","success")
    return redirect(url_for("users"))

@app.route("/admin/users/<int:uid>/projects",methods=["POST"])
@admin_required
def assign_projects(uid):
    c=db()
    c.execute("DELETE FROM user_projects WHERE user_id=?",(uid,))
    c.execute("UPDATE project_assignments SET active=0 WHERE user_id=?",(uid,))
    user=c.execute("SELECT position FROM users WHERE id=?",(uid,)).fetchone()
    for pid in request.form.getlist("project_ids"):
        c.execute("INSERT OR IGNORE INTO user_projects(user_id,project_id) VALUES(?,?)",(uid,pid))
        c.execute("INSERT INTO project_assignments(user_id,project_id,position,active) VALUES(?,?,?,1) ON CONFLICT(user_id,project_id) DO UPDATE SET active=1",(uid,pid,(user['position'] if user else '') or ''))
    c.execute("UPDATE responsibilities SET active=0 WHERE subordinate_user_id=? AND scope_type='PROJECT'",(uid,))
    c.execute("INSERT OR IGNORE INTO responsibilities(supervisor_user_id,subordinate_user_id,scope_type,project_id,source) SELECT manager_user_id,user_id,'PROJECT',project_id,'Project Assignment' FROM project_assignments WHERE user_id=? AND active=1 AND manager_user_id IS NOT NULL",(uid,))
    c.commit();c.close();flash("🏗️ Project access and project assignment status updated.","success");return redirect(url_for("users"))

@app.route("/admin/users/<int:uid>/project-contacts", methods=["POST"])
@admin_required
def assign_head_office_project_contacts(uid):
    c=db()
    try:
        ho=c.execute("SELECT * FROM users WHERE id=? AND active=1 AND personnel_scope='HEAD_OFFICE'",(uid,)).fetchone()
        if not ho: raise ValueError("Selected account is not an active Head Office personnel account.")
        pid=int(request.form.get('project_id') or 0)
        area=(request.form.get('responsibility_area') or 'General Project').strip()
        if not c.execute("SELECT id FROM projects WHERE id=?",(pid,)).fetchone(): raise ValueError("Project not found.")
        selected={int(x) for x in request.form.getlist('project_user_ids') if str(x).isdigit()}
        # The HO user gets project responsibility/access; selected project staff become direct contacts.
        c.execute("INSERT INTO project_responsibilities(project_id,user_id,responsibility_area,source,assigned_by,active) VALUES(?,?,?,?,?,1) ON CONFLICT(project_id,user_id,responsibility_area) DO UPDATE SET active=1,source=excluded.source,assigned_by=excluded.assigned_by",(pid,uid,area,'User Contact Assignment',current_user()['id']))
        c.execute("UPDATE personnel_project_contacts SET active=0 WHERE project_id=? AND head_office_user_id=? AND responsibility_area=?",(pid,uid,area))
        for puid in selected:
            valid=c.execute("SELECT id FROM users u JOIN project_assignments pa ON pa.user_id=u.id AND pa.project_id=? AND pa.active=1 WHERE u.id=? AND u.active=1 AND u.personnel_scope='PROJECT'",(pid,puid)).fetchone()
            if not valid: continue
            c.execute("INSERT INTO personnel_project_contacts(project_id,head_office_user_id,project_user_id,responsibility_area,active,assigned_by) VALUES(?,?,?,?,1,?) ON CONFLICT(project_id,head_office_user_id,project_user_id,responsibility_area) DO UPDATE SET active=1,assigned_by=excluded.assigned_by",(pid,uid,puid,area,current_user()['id']))
        c.commit(); flash(f"🤝 Head Office contact assignment saved for {ho['full_name']} on the selected project team.","success")
    except Exception as e:
        c.rollback(); flash("Could not save Head Office/project contact assignment: "+str(e),"error")
    c.close(); return redirect(url_for('users'))

@app.route("/admin/projects",methods=["GET","POST"])
@admin_required
def projects_admin():
    c=db()
    if request.method=="POST":
        try:c.execute("INSERT INTO projects(name,code,location,client,consultant,status,start_date,end_date,contractor_role,contract_sign_date,commencement_date,contract_end_date,contract_days,planned_income,planned_physical_pct,contract_value) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",(request.form["name"],request.form["code"],request.form["location"],request.form.get("client",""),request.form.get("consultant",""),request.form.get("status","Active"),request.form.get("start_date","") ,request.form.get("end_date",""),request.form.get("contractor_role","Main Contractor"),request.form.get("contract_sign_date",""),request.form.get("commencement_date",request.form.get("start_date","")),request.form.get("contract_end_date",request.form.get("end_date","")),int(parse_float(request.form.get("contract_days"))),parse_float(request.form.get("planned_income")),parse_float(request.form.get("planned_physical_pct")),parse_float(request.form.get("contract_value"))));c.commit();flash("🏗️ Project created.","success")
        except Exception as e:c.rollback();flash(str(e),"error")
    projects=c.execute("SELECT * FROM projects ORDER BY name").fetchall();c.close();return render_template("projects_admin.html",projects=projects)



@app.route("/admin/projects/<int:pid>/edit",methods=["GET","POST"])
@admin_required
def edit_project(pid):
    c=db(); p=c.execute("SELECT * FROM projects WHERE id=?",(pid,)).fetchone()
    if not p: c.close(); flash("Project not found.","error"); return redirect(url_for("projects_admin"))
    if request.method=="POST":
        c.execute("""UPDATE projects SET name=?,code=?,location=?,client=?,consultant=?,status=?,contractor_role=?,contract_sign_date=?,commencement_date=?,contract_end_date=?,contract_days=?,planned_income=?,planned_physical_pct=?,contract_value=?,start_date=?,end_date=? WHERE id=?""",(request.form["name"],request.form["code"],request.form.get("location",""),request.form.get("client",""),request.form.get("consultant",""),request.form.get("status","Active"),request.form.get("contractor_role","Main Contractor"),request.form.get("contract_sign_date",""),request.form.get("commencement_date",""),request.form.get("contract_end_date",""),int(parse_float(request.form.get("contract_days"))),parse_float(request.form.get("planned_income")),parse_float(request.form.get("planned_physical_pct")),parse_float(request.form.get("contract_value")),request.form.get("start_date",""),request.form.get("end_date",""),pid)); c.commit(); flash("🏗️ Project baseline updated.","success"); p=c.execute("SELECT * FROM projects WHERE id=?",(pid,)).fetchone()
    c.close(); return render_template("project_edit.html",p=p)

def _norm_header(v):
    if v is None: return ""
    x=str(v).strip().lower().replace("\n"," ")
    for ch in [".",":","/","\\","(",")","-","_"]:
        x=x.replace(ch," ")
    return " ".join(x.split())

def _excel_number(v):
    if v is None or v=="": return 0.0
    if isinstance(v,(int,float)): return float(v)
    x=str(v).strip().replace(",","").replace("ETB","").replace("Birr","").strip()
    if x.startswith("="): return 0.0
    try: return float(x)
    except Exception: return 0.0

def import_boq_xlsx(path,pid):
    """Import real-world BOQ spreadsheets without requiring exact column wording.
    Finds the header row anywhere in each sheet and accepts common Ethiopian/contract BOQ labels.
    """
    wb_values=load_workbook(path,data_only=True,read_only=True)
    wb_formulas=load_workbook(path,data_only=False,read_only=True)
    c=db(); count=0; sheets=[]; scanned=0
    try:
        for ws_val, ws_formula in zip(wb_values.worksheets, wb_formulas.worksheets):
            max_rows=min(ws_val.max_row or 1,10000)
            rows=list(ws_val.iter_rows(min_row=1,max_row=max_rows,values_only=True))
            header_idx=None; header=None
            for idx,row in enumerate(rows):
                vals=[_norm_header(v) for v in row]
                joined=" | ".join(vals)
                has_item=any(v in vals or v.startswith("item no") or v.startswith("item number") or v in ("no","s no","s n","bill item","bill no") for v in vals)
                has_desc=any(("description" in v) or ("activity" in v) or ("work description" in v) or ("particular" in v) or ("item description" in v) for v in vals)
                has_unit=any(v=="unit" or v.endswith(" unit") or v=="uom" for v in vals)
                has_qty=any(("quantity" in v) or v in ("qty","contract qty","contract quantity","total quantity") for v in vals)
                has_rate=any(("rate" in v) or ("unit price" in v) or ("unit cost" in v) or ("price"==v) for v in vals)
                if has_desc and has_unit and has_rate and (has_item or has_qty):
                    header_idx=idx; header=vals; break
            if header_idx is None: continue
            def find_col(kind):
                exact={
                  "item": ["item no","item number","item no.","bill item","bill no","no","s no","s n","item"],
                  "desc": ["description","description of work","description of works","work description","activity","particular","particulars","item description"],
                  "unit": ["unit","uom","unit of measurement"],
                  "qty": ["contract quantity","contract qty","total contract quantity","quantity","qty","total quantity","original quantity"],
                  "rate": ["contract rate","unit rate","unit price","unit cost","rate","price"],
                }[kind]
                for i,v in enumerate(header):
                    if v in exact: return i
                for i,v in enumerate(header):
                    if any(x in v for x in exact): return i
                return None
            ci,cd,cu,cq,cr=[find_col(k) for k in ("item","desc","unit","qty","rate")]
            cs,ct=find_col("series") if False else (None,None)
            for i,v in enumerate(header):
                if v in ("series","section","bill section","section no","series no","series number") or "series" in v: cs=i; break
            for i,v in enumerate(header):
                if v in ("title","item title","work title") or "title" in v: ct=i; break
            if cd is None or cu is None or cr is None or (ci is None and cq is None): continue
            sheet_count=0
            for row_idx,row in enumerate(rows[header_idx+1:], start=header_idx+2):
                scanned+=1
                def cell(i): return row[i] if i is not None and i<len(row) else None
                item=cell(ci); desc=cell(cd); unit=cell(cu)
                # Skip repeated headers, section totals and empty lines.
                item_s="" if item is None else str(item).strip()
                desc_s="" if desc is None else str(desc).strip()
                if not desc_s or not item_s: continue
                if _norm_header(item_s) in {"no","item","item no","item number","total","subtotal"}: continue
                lowdesc=_norm_header(desc_s)
                if lowdesc in {"description","description of works","description of work","activity"}: continue
                qty=_excel_number(cell(cq)); rate=_excel_number(cell(cr))
                unit_s="" if unit is None else str(unit).strip()
                # Some BOQs have formulas in the data-only workbook. Try the formula workbook's cached-looking numeric cells.
                if rate==0 and cr is not None and row_idx<=ws_formula.max_row:
                    try:
                        fv=ws_formula.cell(row=row_idx,column=cr+1).value
                        if isinstance(fv,(int,float)): rate=float(fv)
                    except Exception: pass
                if qty==0 and cq is not None and row_idx<=ws_formula.max_row:
                    try:
                        qv=ws_formula.cell(row=row_idx,column=cq+1).value
                        if isinstance(qv,(int,float)): qty=float(qv)
                    except Exception: pass
                # A valid BOQ row should contain at least a unit or a numeric quantity/rate.
                if not unit_s and qty==0 and rate==0: continue
                series_s="" if cs is None or cell(cs) is None else str(cell(cs)).strip()
                title_s="" if ct is None or cell(ct) is None else str(cell(ct)).strip()
                c.execute("INSERT INTO boq(project_id,item_no,description,unit,rate,contract_qty,source_sheet,series,title) VALUES(?,?,?,?,?,?,?,?,?) ON CONFLICT(project_id,item_no) DO UPDATE SET description=excluded.description,unit=excluded.unit,rate=excluded.rate,contract_qty=excluded.contract_qty,source_sheet=excluded.source_sheet,series=excluded.series,title=excluded.title",(pid,item_s,desc_s,unit_s,rate,qty,ws_val.title,series_s,title_s))
                count+=1; sheet_count+=1
            if sheet_count: sheets.append(f"{ws_val.title} ({sheet_count})")
        c.commit()
    except Exception:
        c.rollback(); raise
    finally:
        c.close(); wb_values.close(); wb_formulas.close()
    if not sheets:
        raise ValueError("No BOQ rows were detected. The file must contain columns similar to Item No/No., Description of Works, Unit, Quantity and Rate/Unit Price.")
    return count, ", ".join(sheets)

@app.route("/admin/boq/<int:pid>",methods=["GET","POST"])
@login_required
def boq_admin(pid):
    if not project_admin(pid): return redirect(url_for("project",pid=pid))
    c=db()
    if request.method=="POST":
        action=request.form.get("action")
        try:
            if action=="add":
                c.execute("INSERT INTO boq(project_id,item_no,description,unit,rate,contract_qty,series,title) VALUES(?,?,?,?,?,?,?,?)",(pid,request.form["item_no"].strip(),request.form["description"].strip(),request.form.get("unit","").strip(),parse_float(request.form.get("rate")),parse_float(request.form.get("contract_qty")),request.form.get("series","").strip(),request.form.get("title","").strip()))
            elif action=="settings":
                c.execute("INSERT INTO boq_settings(project_id,title,revision,effective_date) VALUES(?,?,?,?) ON CONFLICT(project_id) DO UPDATE SET title=excluded.title,revision=excluded.revision,effective_date=excluded.effective_date",(pid,request.form.get("boq_title","").strip(),request.form.get("revision","").strip(),request.form.get("effective_date") or None))
            elif action=="edit":
                bid=int(request.form["boq_id"]); c.execute("UPDATE boq SET item_no=?,description=?,unit=?,rate=?,contract_qty=?,series=?,title=? WHERE id=? AND project_id=?",(request.form["item_no"].strip(),request.form["description"].strip(),request.form.get("unit","").strip(),parse_float(request.form.get("rate")),parse_float(request.form.get("contract_qty")),request.form.get("series","").strip(),request.form.get("title","").strip(),bid,pid))
            elif action=="upload":
                f=request.files.get("file")
                if not f or not f.filename.lower().endswith((".xlsx",".xlsm")): raise ValueError("Upload an .xlsx or .xlsm BOQ file.")
                name=secure_filename(f.filename); path=os.path.join(UPLOADS,f"{dt.datetime.now().strftime('%Y%m%d%H%M%S')}_{name}"); f.save(path)
                c.commit(); c.close(); c=None
                try:
                    count,sheet=import_boq_xlsx(path,pid)
                    c=db(); c.execute("INSERT INTO boq_uploads(project_id,filename,uploaded_at,user_id,rows_imported) VALUES(?,?,?,?,?)",(pid,name,dt.datetime.now().isoformat(timespec="seconds"),current_user()["id"],count)); c.commit(); flash(f"📑 BOQ imported successfully: {count} items from {sheet}.","success")
                except Exception as e:
                    c=db(); c.rollback(); flash("BOQ import error: "+str(e),"error")
            if c is not None: c.commit()
        except Exception as e:
            if c is not None:
                try: c.rollback()
                except Exception: pass
            flash("BOQ import error: "+str(e),"error")
    rows=c.execute("SELECT * FROM boq WHERE project_id=? ORDER BY item_no",(pid,)).fetchall();uploads=c.execute("SELECT * FROM boq_uploads WHERE project_id=? ORDER BY uploaded_at DESC LIMIT 10",(pid,)).fetchall();settings=c.execute("SELECT * FROM boq_settings WHERE project_id=?",(pid,)).fetchone();c.close();return render_template("boq.html",pid=pid,rows=rows,uploads=uploads,settings=settings)

@app.route("/admin")
@admin_required
def admin():return redirect(url_for("users"))
@app.route("/health")
def health():return "OK"

@app.after_request
def security_headers(resp):
    resp.headers.setdefault("Cache-Control", "no-store")
    return resp

init_db()
if __name__=="__main__":app.run(host="0.0.0.0",port=int(os.environ.get("PORT",5000)),debug=True)


# ==================== V27 PROJECT ADMIN / HEAD OFFICE ROUTING ====================
@app.route('/requests/<int:rid>/send-head-office', methods=['POST'])
@login_required
def send_request_head_office(rid):
    me=current_user(); c=db(); row=c.execute("SELECT * FROM resource_requests WHERE id=?",(rid,)).fetchone()
    if not row or not row['project_id']: c.close(); flash('Project request not found.','error'); return redirect(url_for('resource_requests'))
    if me['role']!='SUPER_ADMIN' and not project_admin(row['project_id']): c.close(); flash('Only the Project Manager can route this request to Head Office.','error'); return redirect(url_for('resource_requests',pid=row['project_id']))
    try:
        uid=int(request.form.get('head_office_user_id') or 0); target=c.execute("SELECT * FROM users WHERE id=? AND active=1 AND personnel_scope='HEAD_OFFICE'",(uid,)).fetchone()
        if not target: raise ValueError('Select a Head Office recipient. Only Head Office personnel can receive this stage.')
        maxstep=c.execute("SELECT COALESCE(MAX(step_order),0) n FROM request_steps WHERE request_id=?",(rid,)).fetchone()['n']
        c.execute("INSERT INTO request_steps(request_id,step_order,stage,assigned_user_id,to_org_unit_id,department,status,action,comments) VALUES(?,?,?,?,?,?,?,?,?)",(rid,maxstep+1,'HEAD_OFFICE_REVIEW',uid,target['org_unit_id'],target['department'],'PENDING','SENT',request.form.get('comments','Sent by Project Manager')))
        c.execute("UPDATE resource_requests SET next_approver_user_id=?,current_stage='HEAD_OFFICE_REVIEW',status='HEAD_OFFICE_REVIEW',head_office_sent_at=CURRENT_TIMESTAMP WHERE id=?",(uid,rid)); c.commit(); flash(f'📤 {row["request_no"]} sent to {target["full_name"]} in Head Office.','success')
    except Exception as e: c.rollback(); flash('Head Office routing failed: '+str(e),'error')
    c.close(); return redirect(url_for('resource_requests',pid=row['project_id']))

@app.route('/requests/<int:rid>/forward-head-office', methods=['POST'])
@login_required
def forward_request_head_office(rid):
    me=current_user(); c=db(); row=c.execute("SELECT * FROM resource_requests WHERE id=?",(rid,)).fetchone()
    if not row or not user_can_approve_request(me,row): c.close(); flash('Only the current Head Office recipient can forward this request.','error'); return redirect(url_for('resource_requests',pid=row['project_id']) if row and row['project_id'] else url_for('resource_requests'))
    try:
        uid=int(request.form.get('head_office_user_id') or 0); target=c.execute("SELECT * FROM users WHERE id=? AND active=1 AND personnel_scope='HEAD_OFFICE'",(uid,)).fetchone()
        if not target: raise ValueError('Select another Head Office recipient. Only Head Office personnel can receive this stage.')
        maxstep=c.execute("SELECT COALESCE(MAX(step_order),0) n FROM request_steps WHERE request_id=?",(rid,)).fetchone()['n']
        oldstep=current_request_step(c,rid)
        if oldstep: c.execute("UPDATE request_steps SET status='FORWARDED',action='FORWARDED',comments=?,acted_at=CURRENT_TIMESTAMP WHERE id=?",(request.form.get('comments','Forwarded by Head Office'),oldstep['id']))
        c.execute("INSERT INTO request_steps(request_id,step_order,stage,assigned_user_id,to_org_unit_id,department,status,action,comments) VALUES(?,?,?,?,?,?,?,?,?)",(rid,maxstep+1,'HEAD_OFFICE_REVIEW',uid,target['org_unit_id'],target['department'],'PENDING','FORWARDED',request.form.get('comments','Forwarded by Head Office')))
        c.execute("UPDATE resource_requests SET next_approver_user_id=?,current_stage='HEAD_OFFICE_REVIEW',status='HEAD_OFFICE_REVIEW' WHERE id=?",(uid,rid)); c.commit(); flash(f'➡️ {row["request_no"]} forwarded to {target["full_name"]}.','success')
    except Exception as e: c.rollback(); flash('Forward failed: '+str(e),'error')
    c.close(); return redirect(url_for('resource_requests',pid=row['project_id']) if row['project_id'] else url_for('resource_requests'))

@app.route('/projects/<int:pid>/team/responsibilities', methods=['POST'])
@login_required
def update_project_responsibilities(pid):
    me=current_user()
    if not project_admin(pid): return redirect(url_for('project_team',pid=pid))
    c=db()
    try:
        selected_project={int(x) for x in request.form.getlist('project_responsible_user_ids') if str(x).isdigit()}
        selected_head={int(x) for x in request.form.getlist('head_office_responsible_user_ids') if str(x).isdigit()}
        area=request.form.get('responsibility_area','General Project') or 'General Project'
        selected=selected_project|selected_head
        c.execute("UPDATE project_responsibilities SET active=0 WHERE project_id=? AND responsibility_area=?",(pid,area))
        if selected:
            marks=','.join('?'*len(selected)); valid=c.execute(f"SELECT id,personnel_scope FROM users WHERE active=1 AND id IN ({marks})",tuple(selected)).fetchall()
        else: valid=[]
        for usr in valid:
            c.execute("INSERT INTO project_responsibilities(project_id,user_id,responsibility_area,source,assigned_by,active) VALUES(?,?,?,?,?,1) ON CONFLICT(project_id,user_id,responsibility_area) DO UPDATE SET active=1,source=excluded.source,assigned_by=excluded.assigned_by",(pid,usr['id'],area,'Project Admin',me['id']))
        c.commit(); flash(f'👥 {area} responsibility list updated.','success')
    except Exception as e: c.rollback(); flash('Responsibility update failed: '+str(e),'error')
    c.close(); return redirect(url_for('project_team',pid=pid))

@app.route('/projects/<int:pid>/manpower/import', methods=['POST'])
@login_required
def import_manpower_excel(pid):
    if not (project_admin(pid) or can_module('HR')): return redirect(url_for('manpower',pid=pid))
    f=request.files.get('excel_file')
    if not f or not f.filename.lower().endswith(('.xlsx','.xlsm')): flash('Please select an Excel .xlsx/.xlsm file.','error'); return redirect(url_for('manpower',pid=pid))
    path=os.path.join(WORKFLOW_FILES,'import_'+secure_filename(f.filename)); f.save(path)
    try:
        wb=load_workbook(path,data_only=True); ws=wb.active; headers=[str(x.value or '').strip().lower() for x in ws[1]]; idx={h:i for i,h in enumerate(headers)}
        def val(row,names,default=''):
            for n in names:
                if n in idx:return row[idx[n]].value
            return default
        c=db(); count=0
        for row in ws.iter_rows(min_row=2):
            name=val(row,['name','full name','employee name','staff name'])
            if not name: continue
            c.execute("INSERT INTO manpower(project_id,date,name,employment,position,present,working_hours,hourly_rate,daily_rate,notes,user_id) VALUES(?,?,?,?,?,?,?,?,?,?,?)",(pid,val(row,['date']) or dt.date.today().isoformat(),str(name),val(row,['employment','type'],'Temporary') or 'Temporary',val(row,['position','job title'],'Other') or 'Other',parse_float(val(row,['present','qty'],1)),parse_float(val(row,['working hours','hours'],8)),parse_float(val(row,['hourly rate','rate/hour'],0)),parse_float(val(row,['daily rate','rate/day'],0)),'Imported from Excel',current_user()['id'])); count+=1
        c.commit(); c.close(); flash(f'📥 Imported {count} manpower records from Excel.','success')
    except Exception as e: flash('Manpower Excel import failed: '+str(e),'error')
    try: os.remove(path)
    except: pass
    return redirect(url_for('manpower',pid=pid))

@app.route('/projects/<int:pid>/machinery/import', methods=['POST'])
@login_required
def import_machinery_excel(pid):
    if not (project_admin(pid) or can_module('Machinery')): return redirect(url_for('machinery',pid=pid))
    f=request.files.get('excel_file')
    if not f or not f.filename.lower().endswith(('.xlsx','.xlsm')): flash('Please select an Excel .xlsx/.xlsm file.','error'); return redirect(url_for('machinery',pid=pid))
    path=os.path.join(WORKFLOW_FILES,'import_'+secure_filename(f.filename)); f.save(path)
    try:
        wb=load_workbook(path,data_only=True); ws=wb.active; headers=[str(x.value or '').strip().lower() for x in ws[1]]; idx={h:i for i,h in enumerate(headers)}
        def val(row,names,default=''):
            for n in names:
                if n in idx:return row[idx[n]].value
            return default
        c=db(); count=0
        for row in ws.iter_rows(min_row=2):
            code=val(row,['code','fleet code','machine code']); plate=val(row,['plate','plate no','plate number']); engine=val(row,['engine','engine no','engine number'])
            if not (code or plate or engine): continue
            dup=c.execute("SELECT id FROM machines WHERE project_id=? AND (code=? OR (plate_no<>'' AND plate_no=?) OR (engine_no<>'' AND engine_no=?))",(pid,str(code),str(plate),str(engine))).fetchone()
            if dup: continue
            c.execute("INSERT INTO machines(project_id,machine_type,code,plate_no,engine_no,ownership,hourly_rate,rate_unit,expected_fuel,fuel_price,active) VALUES(?,?,?,?,?,?,?,?,?,?,1)",(pid,val(row,['machine type','type'],'Other') or 'Other',str(code or ''),str(plate or ''),str(engine or ''),val(row,['ownership','owner'],'Owned') or 'Owned',parse_float(val(row,['rate','hourly rate'],0)),val(row,['rate unit','basis'],'hr') or 'hr',parse_float(val(row,['expected fuel','fuel l/hr'],0)),parse_float(val(row,['fuel price','fuel price/l'],0)))); count+=1
        c.commit(); c.close(); flash(f'📥 Imported {count} machinery records from Excel.','success')
    except Exception as e: flash('Machinery Excel import failed: '+str(e),'error')
    try: os.remove(path)
    except: pass
    return redirect(url_for('machinery',pid=pid))
